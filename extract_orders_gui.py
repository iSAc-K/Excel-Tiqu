from __future__ import annotations

import json
import os
import queue
import site
import sys
import threading
import traceback
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox
from typing import Any

import tkinter as tk
from tkinter import ttk


def add_dependency_paths() -> None:
    base_dir = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
    candidates: list[Path] = [base_dir / ".codex_test_deps"]
    try:
        candidates.append(Path(site.getusersitepackages()))
    except Exception:
        pass
    for candidate in candidates:
        if candidate.exists():
            text = str(candidate)
            if text not in sys.path:
                sys.path.insert(0, text)


add_dependency_paths()

import customtkinter as ctk


DEFAULT_OUTPUT_NAME = "订单汇总.xlsx"
DEFAULT_SETTINGS = {
    "last_input_path": "",
    "last_output_path": "",
    "workers": 4,
    "clear": False,
    "dry_run": False,
    "detect_duplicate_orders": True,
    "skip_exact_duplicates": True,
    "input_mode": "archives",
    "enable_hc_filter": False,
    "excel_group_mode": "single",
    "window_geometry": "1160x760",
    "selected_page": "start",
}

PAGE_TITLES = {
    "start": "开始处理",
    "progress": "进度日志",
    "reports": "报告与配置",
}

INPUT_MODE_LABELS = {
    "只处理压缩包": "archives",
    "只处理文件夹Excel": "folders",
    "混合模式": "mixed",
}
INPUT_MODE_VALUES = {value: label for label, value in INPUT_MODE_LABELS.items()}

EXCEL_GROUP_LABELS = {
    "单文件订单模式": "single",
    "多文件汇总模式": "multi",
}
EXCEL_GROUP_VALUES = {value: label for label, value in EXCEL_GROUP_LABELS.items()}


def ensure_xlsx_suffix(path_text: str) -> str:
    path_text = path_text.strip()
    if not path_text:
        return ""

    path = Path(path_text)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(path.suffix + ".xlsx") if path.suffix else path.with_suffix(".xlsx")
    return str(path)


def runtime_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def app_settings_path() -> Path:
    return runtime_base_dir() / "app_settings.json"


def category_config_path() -> Path:
    return runtime_base_dir() / "category_config.json"


def load_app_settings() -> tuple[dict[str, Any], str]:
    path = app_settings_path()
    if not path.exists():
        save_app_settings(DEFAULT_SETTINGS)
        return dict(DEFAULT_SETTINGS), ""
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            raise ValueError("设置文件根节点必须是对象")
        settings = dict(DEFAULT_SETTINGS)
        settings.update(data)
        return settings, ""
    except Exception as exc:
        return dict(DEFAULT_SETTINGS), f"设置文件读取失败：{exc}，已使用默认设置"


def save_app_settings(settings: dict[str, Any]) -> None:
    path = app_settings_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")


def unique_timestamped_path(folder: Path, stem: str, suffix: str) -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = folder / f"{stem}_{timestamp}{suffix}"
    if not candidate.exists():
        return candidate
    index = 1
    while True:
        numbered = folder / f"{stem}_{timestamp}_{index}{suffix}"
        if not numbered.exists():
            return numbered
        index += 1


def export_error_rows_to_excel(rows: list[dict[str, Any]], output_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = ["压缩包名称", "压缩包路径", "异常类型", "异常原因", "相关文件", "处理状态", "处理时间"]
    workbook = Workbook()
    ws = workbook.active
    ws.title = "异常压缩包列表"

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row_index, row in enumerate(rows, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row_index, column=col, value=row.get(header, "")).alignment = center

    ws.freeze_panes = "A2"
    for col_index in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_index)
        max_length = 0
        for cell in ws[column_letter]:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 80)

    workbook.save(output_path)
    workbook.close()


def format_final_logs(result: dict[str, Any]) -> str:
    exception_logs = result.get("exception_logs") or []
    process_logs = result.get("process_logs") or []
    log_file_path = result.get("log_file_path") or ""
    output_path = result.get("output_path") or ""
    error_report_path = result.get("error_report_path") or ""
    duplicate_report_file_path = result.get("duplicate_report_path") or result.get("duplicate_report_file_path") or ""
    process_report_path = result.get("process_report_path") or ""
    debug_report_path = result.get("debug_report_path") or result.get("debug_report_file_path") or ""
    backup_path = result.get("backup_path") or ""
    duplicate_report_rows = result.get("duplicate_report_rows") or []
    exact_duplicate_report_rows = result.get("exact_duplicate_report_rows") or []

    sections = ["========== 异常情况汇总 =========="]
    sections.append("\n\n".join(str(item) for item in exception_logs) if exception_logs else "无异常")

    sections.append("")
    sections.append("========== 正常处理日志 ==========")
    sections.append("\n".join(str(item) for item in process_logs) if process_logs else "无正常处理日志")

    sections.append("")
    sections.append("========== 重复订单号提示 ==========")
    if duplicate_report_rows:
        sections.append(f"发现重复订单号：{len(duplicate_report_rows)} 条")
        for index, row in enumerate(duplicate_report_rows[:100], 1):
            sections.append(
                f"{index}. {row.get('亚马逊订单号', '')} | SKU: {row.get('SKU', '')} | "
                f"品类：{row.get('品类', '')} | 来源：{row.get('重复来源', '')}"
            )
    else:
        sections.append("无重复订单号")

    sections.append("")
    sections.append("========== 完全重复行提示 ==========")
    if exact_duplicate_report_rows:
        sections.append(f"发现完全重复行：{len(exact_duplicate_report_rows)} 条")
        for index, row in enumerate(exact_duplicate_report_rows[:100], 1):
            sections.append(
                f"{index}. {row.get('品类', '')} | {row.get('亚马逊订单号', '')} | "
                f"{row.get('SKU', '')} | {row.get('数量', '')} | {row.get('日期', '')} | "
                f"{row.get('重复来源', '')} | {row.get('处理方式', '')}"
            )
    else:
        sections.append("无完全重复行")

    sections.append("")
    sections.append("========== 输出与报告路径 ==========")
    sections.append(f"输出汇总 Excel：{output_path}")
    sections.append(f"异常报告：{error_report_path}")
    sections.append(f"重复报告：{duplicate_report_file_path}")
    sections.append(f"处理报告：{process_report_path}")
    sections.append(f"DEBUG 报告：{debug_report_path}")
    sections.append(f"旧汇总备份：{backup_path}")
    if log_file_path:
        sections.append(f"日志已保存：{log_file_path}")

    return "\n".join(sections)


def format_stats_text(stats: dict[str, Any] | None) -> str:
    if not stats:
        return "还没有处理结果。"

    dry_run = bool(stats.get("dry_run"))
    lines = [
        "处理统计",
        f"当前模式：{'仅预览，不写入汇总 Excel' if dry_run else '正式写入汇总 Excel'}",
        f"总压缩包：{stats.get('total_archives', 0)}",
        f"成功处理：{stats.get('success_archives', 0)}",
        f"异常跳过：{stats.get('skipped_archives', 0)}",
        f"总提取行数：{stats.get('extracted_rows', 0)}",
        f"HC 文件：{stats.get('hc_file_count', 0)}",
        f"HC 复制失败：{stats.get('hc_copy_failed_count', 0)}",
    ]
    if dry_run:
        estimated_written = stats.get("extracted_rows", 0) - stats.get("skipped_exact_duplicate_count", 0)
        lines.append(f"预计可写入行数：{estimated_written}")
    lines.append(f"实际写入行数：{stats.get('written_rows', 0)}")
    lines.append(f"重复订单号：{stats.get('duplicate_order_count', 0)}")
    if dry_run:
        lines.append(f"预计跳过完全重复行：{stats.get('skipped_exact_duplicate_count', 0)}")
    else:
        lines.append(f"跳过完全重复行：{stats.get('skipped_exact_duplicate_count', 0)}")
    lines.append(f"并发数量：{stats.get('workers', 1)}")

    lines.append("")
    lines.append("品类统计：")
    category_counts = stats.get("category_counts") or {}
    if category_counts:
        for category, count in sorted(category_counts.items()):
            lines.append(f"{category}：{count} 行")
    else:
        lines.append("无")

    return "\n".join(lines)


class CategoryConfigWindow(ctk.CTkToplevel):
    def __init__(self, master: ctk.CTk, config_path: Path) -> None:
        super().__init__(master)
        self.title("品类关键词配置")
        self.geometry("820x560")
        self.minsize(760, 500)
        self.transient(master)
        self.config_path = config_path
        self.config: dict[str, list[str]] = {}
        self.current_category: str | None = None
        self.status_var = tk.StringVar(value="")

        self._build_ui()
        self.load_config()
        self.grab_set()

    def _build_ui(self) -> None:
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        header = ctk.CTkFrame(self, fg_color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=20, pady=(18, 8))
        header.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(header, text="品类关键词配置", font=ctk.CTkFont(size=22, weight="bold")).grid(row=0, column=0, sticky="w")
        ctk.CTkLabel(
            header,
            text="配置会写入 category_config.json，下一次提取会自动使用。",
            text_color="#5b6675",
        ).grid(row=1, column=0, sticky="w", pady=(4, 0))

        body = ctk.CTkFrame(self, fg_color="#f6f8fb")
        body.grid(row=1, column=0, sticky="nsew", padx=20, pady=10)
        body.grid_columnconfigure(0, weight=1)
        body.grid_columnconfigure(1, weight=2)
        body.grid_rowconfigure(0, weight=1)

        left = ctk.CTkFrame(body, corner_radius=8)
        left.grid(row=0, column=0, sticky="nsew", padx=(12, 6), pady=12)
        left.grid_columnconfigure(0, weight=1)
        left.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(left, text="品类", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))
        self.category_listbox = tk.Listbox(left, exportselection=False, activestyle="none", relief="flat", highlightthickness=1)
        self.category_listbox.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 10))
        self.category_listbox.bind("<<ListboxSelect>>", self.on_category_select)

        category_buttons = ctk.CTkFrame(left, fg_color="transparent")
        category_buttons.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 12))
        ctk.CTkButton(category_buttons, text="新增", width=66, command=self.add_category).pack(side=tk.LEFT)
        ctk.CTkButton(category_buttons, text="改名", width=66, fg_color="#4b5563", command=self.rename_category).pack(side=tk.LEFT, padx=6)
        ctk.CTkButton(category_buttons, text="删除", width=66, fg_color="#9f1239", command=self.delete_category).pack(side=tk.LEFT)

        right = ctk.CTkFrame(body, corner_radius=8)
        right.grid(row=0, column=1, sticky="nsew", padx=(6, 12), pady=12)
        right.grid_columnconfigure(0, weight=1)
        right.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(right, text="关键词", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, sticky="w", padx=12, pady=(12, 8))
        self.keyword_listbox = tk.Listbox(right, exportselection=False, activestyle="none", relief="flat", highlightthickness=1)
        self.keyword_listbox.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 10))

        keyword_buttons = ctk.CTkFrame(right, fg_color="transparent")
        keyword_buttons.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 12))
        ctk.CTkButton(keyword_buttons, text="新增关键词", width=92, command=self.add_keyword).pack(side=tk.LEFT)
        ctk.CTkButton(keyword_buttons, text="修改关键词", width=92, fg_color="#4b5563", command=self.rename_keyword).pack(side=tk.LEFT, padx=6)
        ctk.CTkButton(keyword_buttons, text="删除关键词", width=92, fg_color="#9f1239", command=self.delete_keyword).pack(side=tk.LEFT)

        footer = ctk.CTkFrame(self, fg_color="transparent")
        footer.grid(row=2, column=0, sticky="ew", padx=20, pady=(2, 18))
        footer.grid_columnconfigure(2, weight=1)
        ctk.CTkButton(footer, text="保存配置", width=112, command=self.save_config).grid(row=0, column=0, sticky="w")
        ctk.CTkButton(footer, text="恢复默认配置", width=132, fg_color="#4b5563", command=self.restore_defaults).grid(row=0, column=1, sticky="w", padx=8)
        ctk.CTkLabel(footer, textvariable=self.status_var, anchor="w", text_color="#5b6675").grid(row=0, column=2, sticky="ew", padx=10)
        ctk.CTkButton(footer, text="关闭", width=88, fg_color="#6b7280", command=self.destroy).grid(row=0, column=3, sticky="e")

    def ask_text(self, title: str, prompt: str, initial: str = "") -> str:
        dialog = ctk.CTkInputDialog(text=prompt, title=title)
        value = dialog.get_input()
        if value is None:
            return ""
        value = value.strip()
        return value if value else initial.strip() if False else value

    def load_config(self) -> None:
        try:
            from extract_orders import ensure_default_category_config, load_category_config

            ensure_default_category_config(self.config_path)
            config, _, error = load_category_config(self.config_path)
            if error:
                messagebox.showwarning("配置读取失败", error, parent=self)
                self.status_var.set(f"已回退默认配置：{self.config_path}")
            else:
                self.status_var.set(f"已加载：{self.config_path}")
            self.config = {category: list(keywords) for category, keywords in config.items()}
        except Exception as exc:
            messagebox.showerror("配置读取失败", f"无法读取品类配置：{exc}", parent=self)
            self.config = {}
            self.status_var.set(f"配置读取失败：{self.config_path}")
        self.refresh_categories()

    def refresh_categories(self) -> None:
        selected = self.current_category
        self.category_listbox.delete(0, tk.END)
        for category in self.config:
            self.category_listbox.insert(tk.END, category)
        if selected in self.config:
            index = list(self.config).index(selected)
            self.category_listbox.selection_set(index)
            self.category_listbox.activate(index)
            self.category_listbox.see(index)
        elif self.config:
            self.current_category = next(iter(self.config))
            self.category_listbox.selection_set(0)
            self.category_listbox.activate(0)
            self.category_listbox.see(0)
        else:
            self.current_category = None
        self.refresh_keywords()

    def refresh_keywords(self) -> None:
        self.keyword_listbox.delete(0, tk.END)
        if not self.current_category:
            self.status_var.set(f"当前没有品类配置：{self.config_path}")
            return
        keywords = self.config.get(self.current_category, [])
        for keyword in keywords:
            self.keyword_listbox.insert(tk.END, keyword)
        self.status_var.set(f"当前品类：{self.current_category}，关键词 {len(keywords)} 个")

    def on_category_select(self, _event: tk.Event) -> None:
        selection = self.category_listbox.curselection()
        if not selection:
            return
        self.current_category = self.category_listbox.get(selection[0])
        self.refresh_keywords()

    def add_category(self) -> None:
        name = self.ask_text("新增品类", "请输入品类名称：")
        if not name:
            return
        if name in self.config:
            messagebox.showwarning("提示", "品类已存在", parent=self)
            return
        self.config[name] = [name]
        self.current_category = name
        self.refresh_categories()

    def rename_category(self) -> None:
        if not self.current_category:
            messagebox.showinfo("提示", "请先选择品类", parent=self)
            return
        new_name = self.ask_text("修改品类名", f"请输入新的品类名称：\n当前：{self.current_category}")
        if not new_name or new_name == self.current_category:
            return
        if new_name in self.config:
            messagebox.showwarning("提示", "品类已存在", parent=self)
            return
        new_config: dict[str, list[str]] = {}
        for category, keywords in self.config.items():
            new_config[new_name if category == self.current_category else category] = keywords
        self.config = new_config
        self.current_category = new_name
        self.refresh_categories()

    def delete_category(self) -> None:
        if not self.current_category:
            messagebox.showinfo("提示", "请先选择品类", parent=self)
            return
        if not messagebox.askyesno("确认删除", f"确定要删除品类“{self.current_category}”吗？", parent=self):
            return
        self.config.pop(self.current_category, None)
        self.current_category = None
        self.refresh_categories()

    def selected_keyword_index(self) -> int | None:
        selection = self.keyword_listbox.curselection()
        if not selection:
            messagebox.showinfo("提示", "请先选择关键词", parent=self)
            return None
        return int(selection[0])

    def add_keyword(self) -> None:
        if not self.current_category:
            messagebox.showinfo("提示", "请先选择品类", parent=self)
            return
        keyword = self.ask_text("新增关键词", "请输入关键词：")
        if not keyword:
            return
        keywords = self.config.setdefault(self.current_category, [])
        if keyword in keywords:
            messagebox.showwarning("提示", "关键词已存在", parent=self)
            return
        keywords.append(keyword)
        self.refresh_keywords()

    def rename_keyword(self) -> None:
        if not self.current_category:
            messagebox.showinfo("提示", "请先选择品类", parent=self)
            return
        index = self.selected_keyword_index()
        if index is None:
            return
        old_keyword = self.config.get(self.current_category, [])[index]
        new_keyword = self.ask_text("修改关键词", f"请输入新的关键词：\n当前：{old_keyword}")
        if not new_keyword:
            return
        self.config[self.current_category][index] = new_keyword
        self.refresh_keywords()
        self.keyword_listbox.selection_set(index)

    def delete_keyword(self) -> None:
        if not self.current_category:
            messagebox.showinfo("提示", "请先选择品类", parent=self)
            return
        index = self.selected_keyword_index()
        if index is None:
            return
        self.config[self.current_category].pop(index)
        self.refresh_keywords()

    def save_config(self) -> None:
        try:
            from extract_orders import save_category_config

            save_category_config(self.config, self.config_path)
        except Exception as exc:
            messagebox.showerror("保存失败", f"保存品类配置失败：{exc}", parent=self)
            return
        if hasattr(self.master, "save_current_settings"):
            self.master.save_current_settings()
        messagebox.showinfo("保存成功", "品类关键词配置已保存", parent=self)

    def restore_defaults(self) -> None:
        if not messagebox.askyesno("确认恢复默认", "确定要恢复默认品类关键词配置吗？当前修改会被覆盖。", parent=self):
            return
        try:
            from extract_orders import copy_default_category_keywords

            self.config = copy_default_category_keywords()
        except Exception as exc:
            messagebox.showerror("恢复失败", f"恢复默认配置失败：{exc}", parent=self)
            return
        self.current_category = next(iter(self.config), None)
        self.refresh_categories()


class ExtractOrdersApp(ctk.CTk):
    def __init__(self) -> None:
        ctk.set_appearance_mode("Light")
        ctk.set_default_color_theme("blue")
        super().__init__()

        self.title("Excel订单数据提取工具")
        self.minsize(1020, 680)
        self.settings, settings_error = load_app_settings()
        self.geometry(str(self.settings.get("window_geometry") or DEFAULT_SETTINGS["window_geometry"]))
        if settings_error:
            messagebox.showwarning("设置读取失败", settings_error)

        self.log_queue: queue.Queue[tuple[str, Any]] = queue.Queue()
        self.worker_thread: threading.Thread | None = None
        self.last_result: dict[str, Any] | None = None
        self.last_error_report_rows: list[dict[str, Any]] = []
        self.selected_page = str(self.settings.get("selected_page") or "start")
        if self.selected_page not in PAGE_TITLES:
            self.selected_page = "start"
        self.options_visible = False

        self.input_var = tk.StringVar(value=str(self.settings.get("last_input_path") or ""))
        self.output_var = tk.StringVar(value=str(self.settings.get("last_output_path") or ""))
        self.clear_var = tk.BooleanVar(value=bool(self.settings.get("clear")))
        self.dry_run_var = tk.BooleanVar(value=bool(self.settings.get("dry_run")))
        self.mode_var = tk.StringVar(value="预览模式" if self.dry_run_var.get() else "正式写入")
        self.detect_duplicates_var = tk.BooleanVar(value=bool(self.settings.get("detect_duplicate_orders", True)))
        self.skip_exact_duplicates_var = tk.BooleanVar(value=bool(self.settings.get("skip_exact_duplicates", True)))
        self.input_mode_var = tk.StringVar(value=INPUT_MODE_VALUES.get(str(self.settings.get("input_mode") or "archives"), "只处理压缩包"))
        self.enable_hc_filter_var = tk.BooleanVar(value=bool(self.settings.get("enable_hc_filter", False)))
        self.excel_group_mode_var = tk.StringVar(value=EXCEL_GROUP_VALUES.get(str(self.settings.get("excel_group_mode") or "single"), "单文件订单模式"))
        self.workers_var = tk.IntVar(value=self.normalize_workers(self.settings.get("workers") or 4))
        self.status_var = tk.StringVar(value="请选择路径后开始提取")
        self.mode_status_var = tk.StringVar(value="")
        self.worker_status_var = tk.StringVar(value="")
        self.report_time_var = tk.StringVar(value="最后报告：-")
        self.progress_text_var = tk.StringVar(value="0 / 0")
        self.current_file_var = tk.StringVar(value="当前处理：-")
        self.recent_file_var = tk.StringVar(value="最近完成：-")
        self.total_archives_var = tk.StringVar(value="0")
        self.completed_archives_var = tk.StringVar(value="0")
        self.failed_archives_var = tk.StringVar(value="0")
        self.active_workers_var = tk.StringVar(value="0")
        self.report_status_var = tk.StringVar(value="还没有处理结果")

        self.nav_buttons: dict[str, ctk.CTkButton] = {}
        self.pages: dict[str, ctk.CTkFrame] = {}
        self.report_buttons: dict[str, ctk.CTkButton] = {}
        self.start_controls: list[Any] = []

        self._build_ui()
        self.ensure_default_category_config()
        self.update_top_status()
        self.switch_page(self.selected_page)
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def _build_ui(self) -> None:
        self.configure(fg_color="#eef2f6")
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.sidebar = ctk.CTkFrame(self, width=220, corner_radius=0, fg_color="#172033")
        self.sidebar.grid(row=0, column=0, sticky="nsw")
        self.sidebar.grid_propagate(False)
        self.sidebar.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            self.sidebar,
            text="Excel\n订单提取",
            justify="left",
            text_color="#ffffff",
            font=ctk.CTkFont(size=24, weight="bold"),
        ).grid(row=0, column=0, sticky="ew", padx=22, pady=(28, 10))
        ctk.CTkLabel(
            self.sidebar,
            text="压缩包到汇总表\n清晰、可追踪、可回看",
            justify="left",
            text_color="#a7b3c7",
            font=ctk.CTkFont(size=13),
        ).grid(row=1, column=0, sticky="ew", padx=22, pady=(0, 24))

        for row, (page_key, title) in enumerate(PAGE_TITLES.items(), 2):
            button = ctk.CTkButton(
                self.sidebar,
                text=title,
                height=42,
                anchor="w",
                fg_color="transparent",
                hover_color="#24314a",
                text_color="#dbe5f4",
                command=lambda key=page_key: self.switch_page(key),
            )
            button.grid(row=row, column=0, sticky="ew", padx=14, pady=4)
            self.nav_buttons[page_key] = button

        ctk.CTkFrame(self.sidebar, height=1, fg_color="#2d3b55").grid(row=6, column=0, sticky="ew", padx=18, pady=(24, 14))
        ctk.CTkLabel(
            self.sidebar,
            text="业务逻辑仍由\nextract_orders.py 执行",
            justify="left",
            text_color="#8fa1bb",
            font=ctk.CTkFont(size=12),
        ).grid(row=7, column=0, sticky="sw", padx=22, pady=(0, 18))

        self.main = ctk.CTkFrame(self, corner_radius=0, fg_color="#eef2f6")
        self.main.grid(row=0, column=1, sticky="nsew")
        self.main.grid_columnconfigure(0, weight=1)
        self.main.grid_rowconfigure(1, weight=1)

        self._build_status_bar()
        self.page_container = ctk.CTkFrame(self.main, fg_color="transparent")
        self.page_container.grid(row=1, column=0, sticky="nsew", padx=22, pady=(0, 22))
        self.page_container.grid_columnconfigure(0, weight=1)
        self.page_container.grid_rowconfigure(0, weight=1)

        self._build_start_page()
        self._build_progress_page()
        self._build_reports_page()

    def _build_status_bar(self) -> None:
        bar = ctk.CTkFrame(self.main, height=68, fg_color="#ffffff", corner_radius=0)
        bar.grid(row=0, column=0, sticky="ew")
        bar.grid_columnconfigure(1, weight=1)

        self.page_title_label = ctk.CTkLabel(bar, text="", font=ctk.CTkFont(size=24, weight="bold"), text_color="#162033")
        self.page_title_label.grid(row=0, column=0, sticky="w", padx=22, pady=18)

        status_items = ctk.CTkFrame(bar, fg_color="transparent")
        status_items.grid(row=0, column=1, sticky="e", padx=22)
        ctk.CTkLabel(status_items, textvariable=self.mode_status_var, text_color="#334155").pack(side=tk.LEFT, padx=(0, 16))
        ctk.CTkLabel(status_items, textvariable=self.worker_status_var, text_color="#334155").pack(side=tk.LEFT, padx=(0, 16))
        ctk.CTkLabel(status_items, textvariable=self.status_var, text_color="#0f766e").pack(side=tk.LEFT, padx=(0, 16))
        ctk.CTkLabel(status_items, textvariable=self.report_time_var, text_color="#64748b").pack(side=tk.LEFT)

    def _build_start_page(self) -> None:
        page = self.make_page("start")
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(3, weight=1)

        intro = ctk.CTkFrame(page, fg_color="#ffffff", corner_radius=8)
        intro.grid(row=0, column=0, sticky="ew", pady=(0, 14))
        intro.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(intro, text="按顺序选择压缩包文件夹和汇总 Excel 保存位置，然后开始处理。", anchor="w", font=ctk.CTkFont(size=16, weight="bold"), text_color="#172033").grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 4))
        ctk.CTkLabel(intro, text="新手建议先用“预览模式”跑一次，确认报告正常后再正式写入。", anchor="w", text_color="#5b6675").grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 16))

        paths = ctk.CTkFrame(page, fg_color="#ffffff", corner_radius=8)
        paths.grid(row=1, column=0, sticky="ew", pady=(0, 14))
        paths.grid_columnconfigure(1, weight=1)
        self.add_path_row(paths, 0, "输入路径", self.input_var, "选择文件夹", self.choose_input_folder)
        self.add_path_row(paths, 1, "汇总 Excel 保存位置", self.output_var, "选择保存位置", self.choose_output_file)

        controls = ctk.CTkFrame(page, fg_color="#ffffff", corner_radius=8)
        controls.grid(row=2, column=0, sticky="ew", pady=(0, 14))
        controls.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(controls, text="处理模式", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 6))
        self.mode_switch = ctk.CTkSegmentedButton(
            controls,
            values=["预览模式", "正式写入"],
            variable=self.mode_var,
            command=self.on_mode_change,
        )
        self.mode_switch.grid(row=1, column=0, sticky="w", padx=18, pady=(0, 18))
        self.start_controls.append(self.mode_switch)

        self.start_button = ctk.CTkButton(
            controls,
            text="开始提取",
            height=46,
            width=150,
            font=ctk.CTkFont(size=16, weight="bold"),
            command=self.start_extract,
        )
        self.start_button.grid(row=1, column=2, sticky="e", padx=18, pady=(0, 18))
        self.start_controls.append(self.start_button)

        self.options_button = ctk.CTkButton(
            controls,
            text="展开处理选项",
            width=128,
            fg_color="#4b5563",
            command=self.toggle_options,
        )
        self.options_button.grid(row=1, column=1, sticky="e", padx=(0, 12), pady=(0, 18))
        self.start_controls.append(self.options_button)

        self.options_frame = ctk.CTkFrame(page, fg_color="#ffffff", corner_radius=8)
        self.options_frame.grid_columnconfigure(1, weight=1)
        self._build_options(self.options_frame)

        filler = ctk.CTkFrame(page, fg_color="transparent")
        filler.grid(row=3, column=0, sticky="nsew")

    def _build_options(self, frame: ctk.CTkFrame) -> None:
        ctk.CTkLabel(frame, text="处理选项", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, columnspan=3, sticky="w", padx=18, pady=(18, 8))

        self.workers_label = ctk.CTkLabel(frame, text=f"同时处理压缩包数量：{self.workers_var.get()}", text_color="#172033")
        self.workers_label.grid(row=1, column=0, sticky="w", padx=18, pady=(0, 10))
        self.workers_slider = ctk.CTkSlider(frame, from_=1, to=8, number_of_steps=7, command=self.on_workers_change)
        self.workers_slider.set(self.workers_var.get())
        self.workers_slider.grid(row=1, column=1, sticky="ew", padx=14, pady=(0, 10))
        ctk.CTkLabel(frame, text="建议 4；电脑性能较好可选 6-8", text_color="#64748b").grid(row=1, column=2, sticky="w", padx=(0, 18), pady=(0, 10))

        ctk.CTkLabel(frame, text="数据来源", text_color="#172033").grid(row=2, column=0, sticky="w", padx=18, pady=(4, 8))
        self.input_mode_menu = ctk.CTkOptionMenu(frame, values=list(INPUT_MODE_LABELS.keys()), variable=self.input_mode_var, command=lambda _value: self.update_top_status())
        self.input_mode_menu.grid(row=2, column=1, sticky="w", padx=14, pady=(4, 8))

        ctk.CTkLabel(frame, text="Excel 规则", text_color="#172033").grid(row=3, column=0, sticky="w", padx=18, pady=(4, 8))
        self.excel_group_mode_menu = ctk.CTkOptionMenu(frame, values=list(EXCEL_GROUP_LABELS.keys()), variable=self.excel_group_mode_var, command=lambda _value: self.update_top_status())
        self.excel_group_mode_menu.grid(row=3, column=1, sticky="w", padx=14, pady=(4, 8))
        self.enable_hc_filter_check = ctk.CTkCheckBox(frame, text="过滤 HC 文件", variable=self.enable_hc_filter_var, command=self.update_top_status)
        self.enable_hc_filter_check.grid(row=3, column=2, sticky="w", padx=(0, 18), pady=(4, 8))

        self.clear_check = ctk.CTkCheckBox(frame, text="清空旧汇总后重新生成", variable=self.clear_var, command=self.update_top_status)
        self.clear_check.grid(row=4, column=0, sticky="w", padx=18, pady=(4, 12))
        self.detect_duplicates_check = ctk.CTkCheckBox(frame, text="检测重复订单号，只提示不自动跳过", variable=self.detect_duplicates_var, command=self.update_top_status)
        self.detect_duplicates_check.grid(row=4, column=1, sticky="w", padx=14, pady=(4, 12))
        self.skip_exact_duplicates_check = ctk.CTkCheckBox(frame, text="跳过完全重复行", variable=self.skip_exact_duplicates_var, command=self.update_top_status)
        self.skip_exact_duplicates_check.grid(row=4, column=2, sticky="w", padx=(0, 18), pady=(4, 12))
        self.start_controls.extend([
            self.workers_slider,
            self.input_mode_menu,
            self.excel_group_mode_menu,
            self.enable_hc_filter_check,
            self.clear_check,
            self.detect_duplicates_check,
            self.skip_exact_duplicates_check,
        ])

    def _build_progress_page(self) -> None:
        page = self.make_page("progress")
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(2, weight=1)

        cards = ctk.CTkFrame(page, fg_color="transparent")
        cards.grid(row=0, column=0, sticky="ew", pady=(0, 14))
        for col in range(4):
            cards.grid_columnconfigure(col, weight=1)
        self.add_metric_card(cards, 0, "总压缩包", self.total_archives_var)
        self.add_metric_card(cards, 1, "已完成", self.completed_archives_var)
        self.add_metric_card(cards, 2, "异常/跳过", self.failed_archives_var)
        self.add_metric_card(cards, 3, "活动线程", self.active_workers_var)

        progress = ctk.CTkFrame(page, fg_color="#ffffff", corner_radius=8)
        progress.grid(row=1, column=0, sticky="ew", pady=(0, 14))
        progress.grid_columnconfigure(0, weight=1)
        self.progress_bar = ctk.CTkProgressBar(progress, height=14)
        self.progress_bar.grid(row=0, column=0, sticky="ew", padx=18, pady=(18, 8))
        self.progress_bar.set(0)
        ctk.CTkLabel(progress, textvariable=self.progress_text_var, text_color="#334155").grid(row=1, column=0, sticky="w", padx=18)
        ctk.CTkLabel(progress, textvariable=self.current_file_var, text_color="#334155").grid(row=2, column=0, sticky="w", padx=18, pady=(4, 0))
        ctk.CTkLabel(progress, textvariable=self.recent_file_var, text_color="#64748b").grid(row=3, column=0, sticky="w", padx=18, pady=(4, 18))

        log_panel = ctk.CTkFrame(page, fg_color="#111827", corner_radius=8)
        log_panel.grid(row=2, column=0, sticky="nsew")
        log_panel.grid_columnconfigure(0, weight=1)
        log_panel.grid_rowconfigure(1, weight=1)
        log_header = ctk.CTkFrame(log_panel, fg_color="transparent")
        log_header.grid(row=0, column=0, sticky="ew", padx=14, pady=(12, 8))
        log_header.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(log_header, text="实时处理日志", text_color="#f8fafc", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, sticky="w")
        ctk.CTkButton(log_header, text="清空日志", width=90, fg_color="#374151", command=self.clear_logs).grid(row=0, column=1, sticky="e")
        self.log_text = ctk.CTkTextbox(log_panel, wrap=tk.WORD, font=ctk.CTkFont(family="Consolas", size=12), fg_color="#0b1120", text_color="#dbeafe")
        self.log_text.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 14))
        self.log_text.configure(state=tk.DISABLED)

    def _build_reports_page(self) -> None:
        page = self.make_page("reports")
        page.grid_columnconfigure(0, weight=1)
        page.grid_rowconfigure(1, weight=1)

        actions = ctk.CTkFrame(page, fg_color="#ffffff", corner_radius=8)
        actions.grid(row=0, column=0, sticky="ew", pady=(0, 14))
        actions.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(actions, text="本次处理结果", font=ctk.CTkFont(size=17, weight="bold"), text_color="#172033").grid(row=0, column=0, sticky="w", padx=18, pady=(18, 4))
        ctk.CTkLabel(actions, textvariable=self.report_status_var, text_color="#64748b").grid(row=1, column=0, sticky="w", padx=18, pady=(0, 14))

        button_row = ctk.CTkFrame(actions, fg_color="transparent")
        button_row.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        specs = [
            ("output_path", "打开输出 Excel"),
            ("process_report_path", "打开处理报告"),
            ("log_file_path", "打开日志文件"),
            ("duplicate_report_path", "打开重复报告"),
            ("debug_report_path", "打开 DEBUG 报告"),
        ]
        for index, (key, text) in enumerate(specs):
            button = ctk.CTkButton(button_row, text=text, width=116, command=lambda path_key=key: self.open_result_path(path_key))
            button.grid(row=0, column=index, sticky="w", padx=(0, 8))
            self.report_buttons[key] = button
        self.export_errors_button = ctk.CTkButton(button_row, text="导出异常列表", width=116, fg_color="#4b5563", command=self.export_error_report)
        self.export_errors_button.grid(row=0, column=len(specs), sticky="w")

        body = ctk.CTkFrame(page, fg_color="transparent")
        body.grid(row=1, column=0, sticky="nsew")
        body.grid_columnconfigure(0, weight=2)
        body.grid_columnconfigure(1, weight=1)
        body.grid_rowconfigure(0, weight=1)

        summary = ctk.CTkFrame(body, fg_color="#ffffff", corner_radius=8)
        summary.grid(row=0, column=0, sticky="nsew", padx=(0, 14))
        summary.grid_columnconfigure(0, weight=1)
        summary.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(summary, text="处理摘要", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 8))
        self.stats_text = ctk.CTkTextbox(summary, wrap=tk.WORD, font=ctk.CTkFont(family="Microsoft YaHei UI", size=13), fg_color="#f8fafc", text_color="#172033")
        self.stats_text.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 18))
        self.stats_text.configure(state=tk.DISABLED)

        config_panel = ctk.CTkFrame(body, fg_color="#ffffff", corner_radius=8)
        config_panel.grid(row=0, column=1, sticky="nsew")
        config_panel.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(config_panel, text="配置与文件", font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 8))
        ctk.CTkLabel(config_panel, text="常用配置会自动记忆。品类关键词会影响订单归类，修改后下一次处理生效。", wraplength=260, justify="left", text_color="#64748b").grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 16))
        self.category_config_button = ctk.CTkButton(config_panel, text="品类关键词配置", height=38, command=self.open_category_config)
        self.category_config_button.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 10))
        ctk.CTkButton(config_panel, text="打开程序 logs 文件夹", height=38, fg_color="#4b5563", command=self.open_logs_folder).grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 10))
        ctk.CTkButton(config_panel, text="打开输出文件夹", height=38, fg_color="#4b5563", command=self.open_output_folder).grid(row=4, column=0, sticky="ew", padx=18, pady=(0, 10))
        ctk.CTkLabel(config_panel, text=f"配置文件：\n{app_settings_path()}", wraplength=260, justify="left", text_color="#64748b").grid(row=5, column=0, sticky="ew", padx=18, pady=(16, 0))

        self.update_report_buttons()
        self.replace_stats("还没有处理结果。")

    def make_page(self, key: str) -> ctk.CTkFrame:
        page = ctk.CTkFrame(self.page_container, fg_color="transparent")
        page.grid(row=0, column=0, sticky="nsew")
        self.pages[key] = page
        return page

    def add_path_row(self, parent: ctk.CTkFrame, row: int, label: str, variable: tk.StringVar, button_text: str, command: Any) -> None:
        ctk.CTkLabel(parent, text=label, font=ctk.CTkFont(size=14, weight="bold"), text_color="#172033").grid(row=row, column=0, sticky="w", padx=18, pady=(18 if row == 0 else 8, 18))
        entry = ctk.CTkEntry(parent, textvariable=variable, height=36)
        entry.grid(row=row, column=1, sticky="ew", padx=(0, 10), pady=(18 if row == 0 else 8, 18))
        button = ctk.CTkButton(parent, text=button_text, width=116, command=command)
        button.grid(row=row, column=2, sticky="e", padx=(0, 18), pady=(18 if row == 0 else 8, 18))
        self.start_controls.extend([entry, button])

    def add_metric_card(self, parent: ctk.CTkFrame, column: int, title: str, variable: tk.StringVar) -> None:
        card = ctk.CTkFrame(parent, fg_color="#ffffff", corner_radius=8)
        card.grid(row=0, column=column, sticky="ew", padx=(0 if column == 0 else 10, 0))
        ctk.CTkLabel(card, text=title, text_color="#64748b").grid(row=0, column=0, sticky="w", padx=16, pady=(14, 2))
        ctk.CTkLabel(card, textvariable=variable, font=ctk.CTkFont(size=26, weight="bold"), text_color="#172033").grid(row=1, column=0, sticky="w", padx=16, pady=(0, 14))

    def switch_page(self, key: str) -> None:
        if key not in self.pages:
            key = "start"
        self.selected_page = key
        self.page_title_label.configure(text=PAGE_TITLES[key])
        for page_key, page in self.pages.items():
            if page_key == key:
                page.tkraise()
            button = self.nav_buttons.get(page_key)
            if button:
                if page_key == key:
                    button.configure(fg_color="#2f65d5", text_color="#ffffff")
                else:
                    button.configure(fg_color="transparent", text_color="#dbe5f4")
        self.save_current_settings()

    def normalize_workers(self, value: Any) -> int:
        try:
            workers = int(value)
        except (TypeError, ValueError):
            workers = 4
        return max(1, min(workers, 8))

    def on_workers_change(self, value: float) -> None:
        workers = self.normalize_workers(round(value))
        self.workers_var.set(workers)
        self.workers_slider.set(workers)
        self.workers_label.configure(text=f"同时处理压缩包数量：{workers}")
        self.update_top_status()

    def on_mode_change(self, value: str) -> None:
        self.dry_run_var.set(value == "预览模式")
        self.update_top_status()

    def toggle_options(self) -> None:
        self.options_visible = not self.options_visible
        if self.options_visible:
            self.options_frame.grid(row=3, column=0, sticky="ew", pady=(0, 14))
            self.options_button.configure(text="收起处理选项")
        else:
            self.options_frame.grid_forget()
            self.options_button.configure(text="展开处理选项")

    def update_top_status(self) -> None:
        mode = "预览模式" if self.dry_run_var.get() else "正式写入"
        self.mode_var.set(mode)
        self.mode_status_var.set(f"模式：{mode}")
        hc_text = "过滤HC" if self.enable_hc_filter_var.get() else "不过滤HC"
        self.worker_status_var.set(f"并发：{self.workers_var.get()} | {self.input_mode_var.get()} | {self.excel_group_mode_var.get()} | {hc_text}")

    def ensure_default_category_config(self) -> None:
        try:
            from extract_orders import ensure_default_category_config

            ensure_default_category_config(category_config_path())
        except Exception as exc:
            messagebox.showwarning("品类配置提示", f"品类配置初始化失败：{exc}")

    def save_current_settings(self) -> None:
        settings = {
            "last_input_path": self.input_var.get().strip(),
            "last_output_path": self.output_var.get().strip(),
            "workers": self.normalize_workers(self.workers_var.get()),
            "clear": self.clear_var.get(),
            "dry_run": self.dry_run_var.get(),
            "detect_duplicate_orders": self.detect_duplicates_var.get(),
            "skip_exact_duplicates": self.skip_exact_duplicates_var.get(),
            "input_mode": self.selected_input_mode(),
            "enable_hc_filter": self.enable_hc_filter_var.get(),
            "excel_group_mode": self.selected_excel_group_mode(),
            "window_geometry": self.geometry(),
            "selected_page": self.selected_page,
        }
        save_app_settings(settings)

    def on_close(self) -> None:
        try:
            self.save_current_settings()
        except Exception:
            pass
        self.destroy()

    def open_category_config(self) -> None:
        CategoryConfigWindow(self, category_config_path())

    def choose_input_folder(self) -> None:
        folder = filedialog.askdirectory(title="选择压缩包所在文件夹")
        if folder:
            self.input_var.set(folder)
            self.save_current_settings()

    def choose_output_file(self) -> None:
        path = filedialog.asksaveasfilename(
            title="选择汇总 Excel 保存位置",
            defaultextension=".xlsx",
            initialfile=DEFAULT_OUTPUT_NAME,
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if path:
            self.output_var.set(ensure_xlsx_suffix(path))
            self.save_current_settings()

    def selected_input_mode(self) -> str:
        return INPUT_MODE_LABELS.get(self.input_mode_var.get(), "archives")

    def selected_excel_group_mode(self) -> str:
        return EXCEL_GROUP_LABELS.get(self.excel_group_mode_var.get(), "single")

    def validate_inputs(self) -> tuple[bool, str, str]:
        input_path = self.input_var.get().strip()
        output_path = ensure_xlsx_suffix(self.output_var.get())
        if output_path:
            self.output_var.set(output_path)

        if not input_path:
            messagebox.showwarning("提示", "请选择输入路径")
            return False, "", ""
        input_item = Path(input_path).expanduser()
        if not input_item.exists():
            messagebox.showwarning("提示", "输入路径不存在")
            return False, "", ""
        if not input_item.is_dir() and input_item.suffix.lower() not in {".zip", ".rar", ".7z", ".xlsx", ".xlsm"}:
            messagebox.showwarning("提示", "输入路径必须是文件夹、压缩包或 Excel 文件")
            return False, "", ""
        if not output_path:
            messagebox.showwarning("提示", "请选择汇总 Excel 保存位置")
            return False, "", ""
        output_folder = Path(output_path).expanduser().parent
        if not output_folder.exists() or not output_folder.is_dir():
            messagebox.showwarning("提示", "输出文件夹不存在")
            return False, "", ""
        return True, str(input_item), str(Path(output_path).expanduser())

    def get_workers(self) -> int:
        workers = self.normalize_workers(self.workers_var.get())
        self.workers_var.set(workers)
        if hasattr(self, "workers_slider"):
            self.workers_slider.set(workers)
        return workers

    def start_extract(self) -> None:
        valid, input_path, output_path = self.validate_inputs()
        if not valid:
            return

        workers = self.get_workers()
        self.save_current_settings()
        self.clear_logs()
        self.clear_stats()
        self.reset_progress()
        self.last_result = None
        self.last_error_report_rows = []
        self.update_report_buttons()
        self.set_running_state(True)
        self.status_var.set("处理中...")
        self.report_status_var.set("正在处理，请到“进度日志”查看实时反馈")
        self.switch_page("progress")

        self.worker_thread = threading.Thread(
            target=self.run_extract_worker,
            args=(
                input_path,
                output_path,
                self.clear_var.get(),
                self.dry_run_var.get(),
                workers,
                self.detect_duplicates_var.get(),
                self.skip_exact_duplicates_var.get(),
                self.selected_input_mode(),
                self.enable_hc_filter_var.get(),
                self.selected_excel_group_mode(),
            ),
            daemon=True,
        )
        self.worker_thread.start()
        self.after(100, self.poll_log_queue)

    def run_extract_worker(
        self,
        input_path: str,
        output_path: str,
        clear: bool,
        dry_run: bool,
        workers: int,
        detect_duplicates: bool,
        skip_exact_duplicates: bool,
        input_mode: str,
        enable_hc_filter: bool,
        excel_group_mode: str,
    ) -> None:
        def log_callback(message: str) -> None:
            self.log_queue.put(("log", message))

        def progress_callback(payload: dict[str, Any]) -> None:
            self.log_queue.put(("progress", payload))

        try:
            from extract_orders import run_extract

            result = run_extract(
                input_path,
                output_path,
                clear=clear,
                dry_run=dry_run,
                workers=workers,
                detect_duplicate_orders=detect_duplicates,
                skip_exact_duplicates=skip_exact_duplicates,
                category_config_path=category_config_path(),
                report_dir=runtime_base_dir() / "logs",
                backup_dir=runtime_base_dir() / "backups",
                log_callback=log_callback,
                progress_callback=progress_callback,
                input_mode=input_mode,
                enable_hc_filter=enable_hc_filter,
                excel_group_mode=excel_group_mode,
            )
            self.log_queue.put(("done", result))
        except Exception as exc:
            error_text = f"异常：处理失败，原因：{exc}\n{traceback.format_exc()}"
            result = {
                "success": False,
                "total_rows": 0,
                "rows": [],
                "stats": {
                    "total_archives": 0,
                    "success_archives": 0,
                    "skipped_archives": 0,
                    "extracted_rows": 0,
                    "written_rows": 0,
                    "category_counts": {},
                    "dry_run": dry_run,
                    "workers": workers,
                    "duplicate_order_count": 0,
                    "skipped_exact_duplicate_count": 0,
                },
                "exception_logs": [error_text],
                "process_logs": [],
                "error_report_rows": [],
                "duplicate_report_rows": [],
                "exact_duplicate_report_rows": [],
                "duplicate_report_file_path": "",
                "duplicate_report_path": "",
                "error_report_path": "",
                "process_report_path": "",
                "debug_report_path": "",
                "backup_path": "",
                "log_dir": str((runtime_base_dir() / "logs").resolve()),
                "written_rows": 0,
                "error_count": 0,
                "duplicate_count": 0,
                "exact_duplicate_count": 0,
                "skipped_archives": 0,
                "processed_archives": 0,
                "log_file_path": "",
                "output_path": output_path,
            }
            self.log_queue.put(("log", error_text))
            self.log_queue.put(("done", result))

    def poll_log_queue(self) -> None:
        completed_result: dict[str, Any] | None = None
        while True:
            try:
                item_type, payload = self.log_queue.get_nowait()
            except queue.Empty:
                break
            if item_type == "log":
                self.append_log(str(payload))
            elif item_type == "progress":
                self.update_progress(payload)
            elif item_type == "done":
                completed_result = payload

        if completed_result is not None:
            self.finish_extract(completed_result)
            return
        if self.worker_thread and self.worker_thread.is_alive():
            self.after(100, self.poll_log_queue)

    def finish_extract(self, result: dict[str, Any]) -> None:
        self.last_result = result
        self.last_error_report_rows = list(result.get("error_report_rows") or [])
        self.replace_logs(format_final_logs(result))
        self.replace_stats(format_stats_text(result.get("stats")))
        self.set_running_state(False)
        self.update_report_buttons()

        success = bool(result.get("success"))
        total_rows = int(result.get("total_rows") or 0)
        output_path = str(result.get("output_path") or "")
        exception_logs = result.get("exception_logs") or []
        stats = result.get("stats") or {}
        dry_run = bool(stats.get("dry_run"))
        self.report_time_var.set(f"最后报告：{datetime.now().strftime('%H:%M:%S')}")

        if success:
            self.status_var.set("处理完成")
            self.report_status_var.set(f"处理完成：提取 {total_rows} 行，写入 {stats.get('written_rows', 0)} 行")
            self.switch_page("reports")
            if dry_run:
                messagebox.showinfo("预览完成", f"当前为预览模式，未实际写入 Excel\n共提取 {total_rows} 行数据")
            elif exception_logs:
                messagebox.showwarning("处理完成", f"处理完成，但存在异常情况，请查看报告\n共提取 {total_rows} 行数据")
            else:
                messagebox.showinfo("处理完成", f"处理完成\n共提取 {total_rows} 行数据\n输出文件：{output_path}")
        else:
            self.status_var.set("处理失败，请查看日志")
            self.report_status_var.set("处理失败，请查看日志和异常信息")
            self.switch_page("reports")
            messagebox.showerror("处理失败", "处理失败，请查看日志")

    def set_running_state(self, running: bool) -> None:
        state = tk.DISABLED if running else tk.NORMAL
        for control in self.start_controls:
            try:
                control.configure(state=state)
            except Exception:
                pass
        self.start_button.configure(text="处理中..." if running else "开始提取")
        self.category_config_button.configure(state=state)
        self.export_errors_button.configure(state=state)

    def update_progress(self, payload: dict[str, Any]) -> None:
        total = int(payload.get("total") or 0)
        current = int(payload.get("current") or 0)
        archive_name = str(payload.get("archive_name") or "-")
        status = str(payload.get("status") or "")
        active_workers = int(payload.get("active_workers") or 0)
        completed_archives = int(payload.get("completed_archives") or current)
        failed_archives = int(payload.get("failed_archives") or 0)

        self.total_archives_var.set(str(total))
        self.completed_archives_var.set(str(completed_archives))
        self.failed_archives_var.set(str(failed_archives))
        self.active_workers_var.set(str(active_workers))
        self.progress_text_var.set(f"处理进度：{current} / {total}")
        if total > 0:
            self.progress_bar.set(max(0, min(current / total, 1)))
        else:
            self.progress_bar.set(0)
        if status == "processing":
            self.current_file_var.set(f"当前处理：{archive_name}")
        elif status in {"done", "error"}:
            self.recent_file_var.set(f"最近完成：{archive_name}")
            self.current_file_var.set("当前处理：等待其他压缩包完成" if current < total else "当前处理：-")
        else:
            self.current_file_var.set(f"当前处理：{archive_name}")

    def reset_progress(self) -> None:
        self.progress_bar.set(0)
        self.progress_text_var.set("处理进度：0 / 0")
        self.current_file_var.set("当前处理：-")
        self.recent_file_var.set("最近完成：-")
        self.total_archives_var.set("0")
        self.completed_archives_var.set("0")
        self.failed_archives_var.set("0")
        self.active_workers_var.set("0")

    def append_log(self, message: str) -> None:
        self.log_text.configure(state=tk.NORMAL)
        if self.log_text.index("end-1c") != "1.0":
            self.log_text.insert(tk.END, "\n")
        self.log_text.insert(tk.END, message)
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def replace_logs(self, content: str) -> None:
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete("1.0", tk.END)
        self.log_text.insert(tk.END, content)
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def clear_logs(self) -> None:
        self.replace_logs("")

    def replace_stats(self, content: str) -> None:
        self.stats_text.configure(state=tk.NORMAL)
        self.stats_text.delete("1.0", tk.END)
        self.stats_text.insert(tk.END, content)
        self.stats_text.configure(state=tk.DISABLED)

    def clear_stats(self) -> None:
        self.replace_stats("还没有处理结果。")
        self.report_status_var.set("正在处理，请到“进度日志”查看实时反馈")

    def update_report_buttons(self) -> None:
        result = self.last_result or {}
        for key, button in self.report_buttons.items():
            path_text = str(result.get(key) or "")
            if key == "debug_report_path" and not path_text:
                path_text = str(result.get("debug_report_file_path") or "")
            exists = bool(path_text and Path(path_text).expanduser().exists())
            button.configure(state=tk.NORMAL if exists else tk.DISABLED)
        has_errors = bool(self.last_error_report_rows)
        if hasattr(self, "export_errors_button"):
            self.export_errors_button.configure(state=tk.NORMAL if has_errors else tk.DISABLED)

    def export_error_report(self) -> None:
        rows = self.last_error_report_rows
        if not rows:
            messagebox.showinfo("提示", "本次没有异常")
            return
        output_path = ensure_xlsx_suffix(self.output_var.get())
        default_folder = Path(output_path).expanduser().parent if output_path else runtime_base_dir()
        if not default_folder.exists():
            default_folder = runtime_base_dir()
        export_path = unique_timestamped_path(default_folder, "异常压缩包列表", ".xlsx")
        try:
            export_error_rows_to_excel(rows, export_path)
        except Exception as exc:
            messagebox.showerror("导出失败", f"导出异常列表失败：{exc}")
            return
        if self.last_result is not None:
            self.last_result["error_report_path"] = str(export_path.resolve())
        self.append_log(f"异常列表已导出：{export_path}")
        self.update_report_buttons()
        messagebox.showinfo("导出完成", f"异常列表已导出：{export_path}")

    def open_result_path(self, key: str) -> None:
        if not self.last_result:
            messagebox.showinfo("提示", "还没有处理结果")
            return
        path_text = str(self.last_result.get(key) or "")
        if key == "debug_report_path" and not path_text:
            path_text = str(self.last_result.get("debug_report_file_path") or "")
        self.open_path(path_text, "报告文件不存在，请先运行处理或查看日志")

    def open_logs_folder(self) -> None:
        self.open_path(str((runtime_base_dir() / "logs").resolve()), "logs 文件夹不存在")

    def open_output_folder(self) -> None:
        output_path = ensure_xlsx_suffix(self.output_var.get())
        if output_path:
            self.output_var.set(output_path)
        if not output_path:
            messagebox.showwarning("提示", "请选择汇总 Excel 保存位置")
            return
        output_folder = Path(output_path).expanduser().parent
        self.open_path(str(output_folder), "输出文件夹不存在")

    def open_path(self, path_text: str, missing_message: str) -> None:
        if not path_text:
            messagebox.showinfo("提示", missing_message)
            return
        path = Path(path_text).expanduser()
        if not path.exists():
            messagebox.showinfo("提示", missing_message)
            return
        os.startfile(path)


def main() -> None:
    app = ExtractOrdersApp()
    app.mainloop()


if __name__ == "__main__":
    main()
