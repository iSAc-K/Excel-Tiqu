from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ARCHIVE_EXTENSIONS = {".zip", ".rar", ".7z"}
EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
HC_FOLDER_NAME = "HC"
UNCLASSIFIED_FOLDER_NAME = "未分类Excel"
SKIP_FILE_PREFIXES = ("修改", "售后")
INPUT_MODE_ARCHIVES = "archives"
INPUT_MODE_FOLDERS = "folders"
INPUT_MODE_MIXED = "mixed"
INPUT_MODES = {INPUT_MODE_ARCHIVES, INPUT_MODE_FOLDERS, INPUT_MODE_MIXED}
EXCEL_GROUP_SINGLE = "single"
EXCEL_GROUP_MULTI = "multi"
EXCEL_GROUP_MODES = {EXCEL_GROUP_SINGLE, EXCEL_GROUP_MULTI}
OUTPUT_HEADERS = ["亚马逊订单号", "SKU", "数量", "日期"]
OUTPUT_DATE_YEAR = datetime.now().year
OUTPUT_DATE_NUMBER_FORMAT = 'm"月"d"日"'
CORE_HEADER_LABELS = {
    "order_id": "亚马逊订单号",
    "sku": "SKU",
    "quantity": "数量",
}
CORE_HEADER_KEYS = tuple(CORE_HEADER_LABELS.keys())
EXPECTED_QUANTITY_UNITS = (
    "数量",
    "pieces",
    "piece",
    "pcs",
    "pc",
    "个",
    "件",
    "套",
    "只",
    "条",
    "张",
    "份",
    "包",
    "箱",
    "袋",
    "盒",
    "对",
    "双",
    "支",
    "根",
    "瓶",
    "罐",
    "卷",
    "片",
    "台",
    "把",
    "枚",
    "块",
    "组",
    "本",
    "部",
    "副",
    "串",
)
DUPLICATE_REPORT_HEADERS = [
    "类型",
    "压缩包名",
    "Excel 文件名",
    "Sheet 名",
    "源数据行号",
    "亚马逊订单号",
    "SKU",
    "数量",
    "日期",
    "处理方式",
]

DEFAULT_CATEGORY_KEYWORDS = {
    "军牌钥匙扣": ["军牌钥匙扣"],
    "军牌项链": ["军牌项链"],
    "钢片军牌钥匙扣": ["钢片军牌钥匙扣"],
    "小钢片": ["小钢片"],
    "刀叉": ["刀叉", "刀铲"],
    "纯木名片架": ["纯木名片架"],
    "方黑名片架": ["方黑名片架"],
    "雕刻手链": ["雕刻手链"],
    "NP图片项链": ["NP图片项链", "诺派旋转图片项链"],
    "翅膀图片项链": ["翅膀图片项链", "银翅膀图片项链", "金翅膀图片项链", "翅膀项链", "热转印-图片项链"],
    "双面钥匙扣": ["双图热转钥匙扣", "双面爱心钥匙扣", "双面圆形钥匙扣", "双面心形钥匙扣"],
    "旋转钥匙扣": ["旋转钥匙扣"],
    "心形刻字钥匙扣": ["心形刻字钥匙扣"],
    "MA88钥匙扣": ["MA88钥匙扣"],
    "四图钥匙扣": ["四图钥匙扣", "诺派四图钥匙扣", "四图诺派钥匙扣"],
    "热转印骨灰项链": ["骨灰项链"],
}
CATEGORY_KEYWORDS = dict(DEFAULT_CATEGORY_KEYWORDS)

exception_logs: list[str] = []
process_logs: list[str] = []
debug_logs: list[str] = []
error_report_rows: list[dict[str, Any]] = []
skipped_temp_excel_logs: dict[Path, list[str]] = {}
_log_callback: Callable[[str], None] | None = None
_thread_context = threading.local()
_hc_copy_lock = threading.Lock()

SHEET_TAB_COLORS = {
    "刀叉": "4F81BD",
    "军牌钥匙扣": "9BBB59",
    "军牌项链": "8064A2",
    "钢片军牌钥匙扣": "F79646",
    "小钢片": "92CDDC",
    "纯木名片架": "C0504D",
    "方黑名片架": "1F4E79",
    "雕刻手链": "948A54",
    "NP图片项链": "7030A0",
    "翅膀图片项链": "B1A0C7",
    "双面钥匙扣": "00B0F0",
    "旋转钥匙扣": "00B050",
    "心形刻字钥匙扣": "FF66CC",
    "MA88钥匙扣": "FFC000",
    "四图钥匙扣": "C00000",
    "热转印骨灰项链": "808080",
    "未分类": "A6A6A6",
}
FALLBACK_SHEET_TAB_COLORS = ["5B9BD5", "70AD47", "FFC000", "ED7D31", "A5A5A5", "4472C4", "997300"]


def default_category_config_path() -> Path:
    return get_runtime_base_dir() / "category_config.json"


def default_app_settings_path() -> Path:
    return get_runtime_base_dir() / "app_settings.json"


def copy_default_category_keywords() -> dict[str, list[str]]:
    return {category: list(keywords) for category, keywords in DEFAULT_CATEGORY_KEYWORDS.items()}


@dataclass
class CategoryConfigData:
    categories: dict[str, list[str]]
    prefixes: list[str]

    def copy(self) -> "CategoryConfigData":
        return CategoryConfigData(
            categories={category: list(keywords) for category, keywords in self.categories.items()},
            prefixes=list(self.prefixes),
        )


def copy_default_category_config_data() -> CategoryConfigData:
    return CategoryConfigData(copy_default_category_keywords(), [])


def validate_category_config(data: Any) -> dict[str, list[str]]:
    if not isinstance(data, dict):
        raise ValueError("配置文件根节点必须是对象")

    config: dict[str, list[str]] = {}
    for category, keywords in data.items():
        category_text = str(category).strip()
        if not category_text:
            continue
        if not isinstance(keywords, list):
            raise ValueError(f"品类 {category_text} 的关键词必须是列表")
        cleaned_keywords = []
        for keyword in keywords:
            keyword_text = str(keyword).strip()
            if keyword_text:
                cleaned_keywords.append(keyword_text)
        config[category_text] = cleaned_keywords

    if not config:
        raise ValueError("配置文件没有可用的品类")
    return config


def validate_prefixes(data: Any) -> list[str]:
    if data is None:
        return []
    if not isinstance(data, list):
        raise ValueError("prefixes 必须是列表")
    prefixes: list[str] = []
    seen: set[str] = set()
    for item in data:
        prefix = str(item).strip()
        if prefix and prefix not in seen:
            prefixes.append(prefix)
            seen.add(prefix)
    return prefixes


def validate_category_config_data(data: Any) -> CategoryConfigData:
    if not isinstance(data, dict):
        raise ValueError("配置文件根节点必须是对象")
    if isinstance(data.get("categories"), dict):
        categories = validate_category_config(data.get("categories", {}))
        prefixes = validate_prefixes(data.get("prefixes", []))
        return CategoryConfigData(categories=categories, prefixes=prefixes)
    return CategoryConfigData(categories=validate_category_config(data), prefixes=[])


def save_category_config(config: dict[str, list[str]], config_path: str | Path | None = None) -> Path:
    path = Path(config_path).expanduser() if config_path else default_category_config_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    validated = validate_category_config(config)
    path.write_text(json.dumps(validated, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def ensure_default_category_config(config_path: str | Path | None = None) -> Path:
    path = Path(config_path).expanduser() if config_path else default_category_config_path()
    if not path.exists():
        save_category_config(copy_default_category_keywords(), path)
    return path


def save_category_config_data(config_data: CategoryConfigData, config_path: str | Path | None = None) -> Path:
    path = Path(config_path).expanduser() if config_path else default_category_config_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    validated = validate_category_config_data(
        {
            "categories": config_data.categories,
            "prefixes": config_data.prefixes,
        }
    )
    payload = {
        "prefixes": validated.prefixes,
        "categories": validated.categories,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def load_category_config_data(
    config_path: str | Path | None = None,
    create_if_missing: bool = False,
) -> tuple[CategoryConfigData, str, str]:
    path = Path(config_path).expanduser() if config_path else default_category_config_path()
    if create_if_missing and not path.exists():
        try:
            save_category_config(copy_default_category_keywords(), path)
        except Exception as exc:
            return copy_default_category_config_data(), str(path), f"创建品类配置失败：{exc}"

    if not path.exists():
        return copy_default_category_config_data(), str(path), "品类配置文件不存在，已使用内置默认配置"

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return validate_category_config_data(data), str(path), ""
    except Exception as exc:
        return copy_default_category_config_data(), str(path), f"品类配置文件读取失败：{exc}，已使用内置默认配置"


def save_confirmed_category_candidate(
    config_path: str | Path | None,
    *,
    prefix: str = "",
    category: str,
    keyword: str = "",
) -> Path:
    category_text = str(category).strip()
    if not category_text:
        raise ValueError("候选品类不能为空")
    keyword_text = str(keyword or category_text).strip()
    prefix_text = str(prefix).strip()

    path = Path(config_path).expanduser() if config_path else default_category_config_path()
    existed_before = path.exists()
    config_data, _, error = load_category_config_data(path, create_if_missing=True)
    if error and existed_before:
        raise ValueError(error)
    if error and not existed_before and "已使用内置默认配置" not in error:
        raise ValueError(error)

    if prefix_text and prefix_text not in config_data.prefixes:
        config_data.prefixes.append(prefix_text)

    keywords = config_data.categories.setdefault(category_text, [])
    if keyword_text and keyword_text not in keywords:
        keywords.append(keyword_text)
    if category_text not in keywords:
        keywords.insert(0, category_text)

    return save_category_config_data(config_data, path)


def save_candidate_keyword_to_existing_category(
    config_path: str | Path | None,
    *,
    prefix: str = "",
    target_category: str,
    keyword: str,
) -> Path:
    target_text = str(target_category).strip()
    if not target_text:
        raise ValueError("已有品类不能为空")
    keyword_text = str(keyword).strip()
    if not keyword_text:
        raise ValueError("候选关键词不能为空")
    prefix_text = str(prefix).strip()

    path = Path(config_path).expanduser() if config_path else default_category_config_path()
    existed_before = path.exists()
    config_data, _, error = load_category_config_data(path, create_if_missing=True)
    if error and existed_before:
        raise ValueError(error)
    if error and not existed_before and "已使用内置默认配置" not in error:
        raise ValueError(error)
    if target_text not in config_data.categories:
        raise ValueError(f"已有品类不存在：{target_text}")

    if prefix_text and prefix_text not in config_data.prefixes:
        config_data.prefixes.append(prefix_text)

    keywords = config_data.categories[target_text]
    if keyword_text not in keywords:
        keywords.append(keyword_text)

    return save_category_config_data(config_data, path)


def load_category_config(
    config_path: str | Path | None = None,
    create_if_missing: bool = False,
) -> tuple[dict[str, list[str]], str, str]:
    config_data, path, error = load_category_config_data(config_path, create_if_missing)
    return config_data.categories, path, error


def set_log_callback(log_callback: Callable[[str], None] | None) -> None:
    global _log_callback
    _log_callback = log_callback


def emit_log(message: str) -> None:
    if _log_callback is None:
        return
    try:
        _log_callback(message)
    except Exception:
        pass


def reset_logs() -> None:
    exception_logs.clear()
    process_logs.clear()
    debug_logs.clear()
    error_report_rows.clear()
    skipped_temp_excel_logs.clear()


def add_exception(message: str) -> None:
    current_exception_logs = getattr(_thread_context, "exception_logs", None)
    if current_exception_logs is not None:
        current_exception_logs.append(message)
    else:
        exception_logs.append(message)
    emit_log(message)


def add_log(message: str) -> None:
    current_process_logs = getattr(_thread_context, "process_logs", None)
    if current_process_logs is not None:
        current_process_logs.append(message)
    else:
        process_logs.append(message)
    emit_log(message)


def add_debug_log(message: str) -> None:
    add_log(message)
    current_debug_logs = getattr(_thread_context, "debug_logs", None)
    if current_debug_logs is not None:
        current_debug_logs.append(message)
    else:
        debug_logs.append(message)


def set_thread_log_context(
    archive_path: Path,
    archive_process_logs: list[str],
    archive_exception_logs: list[str],
    archive_error_report_rows: list[dict[str, Any]],
    archive_debug_logs: list[str],
) -> None:
    _thread_context.archive_path = archive_path
    _thread_context.process_logs = archive_process_logs
    _thread_context.exception_logs = archive_exception_logs
    _thread_context.error_report_rows = archive_error_report_rows
    _thread_context.debug_logs = archive_debug_logs


def clear_thread_log_context() -> None:
    for name in ("archive_path", "process_logs", "exception_logs", "error_report_rows", "debug_logs"):
        if hasattr(_thread_context, name):
            delattr(_thread_context, name)


def add_structured_error(
    archive_path: Path | None,
    exception_type: str,
    reason: str,
    related_files: str = "",
    status: str = "已跳过",
) -> None:
    actual_archive_path = archive_path or getattr(_thread_context, "archive_path", None)
    archive_name = actual_archive_path.name if isinstance(actual_archive_path, Path) else ""
    archive_text = str(actual_archive_path) if isinstance(actual_archive_path, Path) else ""
    row = {
        "压缩包名称": archive_name,
        "压缩包路径": archive_text,
        "异常类型": exception_type,
        "异常原因": reason,
        "相关文件": related_files,
        "处理状态": status,
        "处理时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    current_error_rows = getattr(_thread_context, "error_report_rows", None)
    if current_error_rows is not None:
        current_error_rows.append(row)
    else:
        error_report_rows.append(row)


def find_winrar() -> Path | None:
    """
    自动查找电脑上的 WinRAR.exe 或 Rar.exe。
    """
    candidates = [
        Path(r"C:\Program Files\WinRAR\WinRAR.exe"),
        Path(r"C:\Program Files\WinRAR\Rar.exe"),
        Path(r"C:\Program Files (x86)\WinRAR\WinRAR.exe"),
        Path(r"C:\Program Files (x86)\WinRAR\Rar.exe"),
    ]

    for candidate in candidates:
        if candidate.exists():
            return candidate

    for executable in ("WinRAR.exe", "Rar.exe"):
        found = shutil.which(executable)
        if found:
            return Path(found)

    return None


def find_archive_files(input_path: str | Path) -> list[Path]:
    """
    如果 input_path 是单个压缩包文件，则返回这个文件。
    如果 input_path 是文件夹，则查找文件夹下所有支持的压缩包文件。
    支持 .zip、.rar、.7z。
    """
    path = Path(input_path).expanduser()

    if path.is_file():
        if path.suffix.lower() in ARCHIVE_EXTENSIONS:
            add_log(f"输入为单个压缩包：{path}")
            return [path]
        add_exception(f"异常：输入文件不是支持的压缩包格式：{path}")
        add_structured_error(path, "不支持的压缩包格式", f"输入文件不是支持的压缩包格式：{path.suffix}", status="已跳过")
        return []

    if path.is_dir():
        add_log(f"开始扫描压缩包文件夹：{path}")
        excluded_folder_names = {HC_FOLDER_NAME, build_skip_dir(path).name, "跳过不处理"}
        archives = sorted(
            file
            for file in path.rglob("*")
            if file.is_file() and file.suffix.lower() in ARCHIVE_EXTENSIONS
            and not set(file.relative_to(path).parts[:-1]).intersection(excluded_folder_names)
        )
        add_log(f"找到压缩包：{len(archives)} 个")
        return archives

    add_exception(f"异常：输入路径不存在：{path}")
    return []


def extract_archive(archive_path: Path) -> Path | None:
    """
    解压单个压缩包到临时目录。
    优先使用 WinRAR。
    如果没有 WinRAR，且文件是 .zip，则使用 zipfile 兜底解压。
    如果没有 WinRAR，且文件不是 .zip，则记录异常日志并跳过。
    """
    extracted_root = Path(tempfile.mkdtemp(prefix="extract_orders_"))
    winrar_path = find_winrar()

    if winrar_path:
        add_log("使用 WinRAR 解压")
        command = [
            str(winrar_path),
            "x",
            "-y",
            str(archive_path),
            str(extracted_root) + os.sep,
        ]
        if winrar_path.name.lower() == "winrar.exe":
            command.insert(2, "-ibck")

        try:
            result = subprocess.run(
                command,
                check=False,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
            )
        except OSError as exc:
            shutil.rmtree(extracted_root, ignore_errors=True)
            add_exception(f"异常：解压失败 {archive_path.name}，原因：{exc}")
            add_structured_error(archive_path, "解压失败", f"解压失败，原因：{exc}")
            return None

        if result.returncode != 0:
            shutil.rmtree(extracted_root, ignore_errors=True)
            output = (result.stderr or result.stdout or "").strip()
            add_exception(f"异常：解压失败 {archive_path.name}，WinRAR 返回码：{result.returncode}\n{output}")
            add_structured_error(archive_path, "解压失败", f"WinRAR 返回码：{result.returncode}；{output}")
            return None

        return extracted_root

    if archive_path.suffix.lower() == ".zip":
        add_log("未找到 WinRAR，使用 Python zipfile 解压 .zip")
        try:
            with zipfile.ZipFile(archive_path) as zip_file:
                zip_file.extractall(extracted_root)
        except (zipfile.BadZipFile, OSError) as exc:
            shutil.rmtree(extracted_root, ignore_errors=True)
            add_exception(f"异常：解压失败 {archive_path.name}，原因：{exc}")
            add_structured_error(archive_path, "解压失败", f"解压失败，原因：{exc}")
            return None
        return extracted_root

    shutil.rmtree(extracted_root, ignore_errors=True)
    add_exception(
        f"异常：未找到 WinRAR，无法解压：{archive_path.name}\n"
        "请安装 WinRAR 或检查 WinRAR 路径"
    )
    add_structured_error(archive_path, "未找到 WinRAR", "未找到 WinRAR，无法解压 rar/7z 压缩包")
    return None


def find_excel_files_in_extracted_dir(root_dir: Path) -> list[Path]:
    """
    在解压目录里递归查找 Excel 文件。
    忽略 ~$ 开头的临时文件。
    支持 .xlsx 和 .xlsm。
    遇到 .xls 日志提示跳过。
    """
    excel_files: list[Path] = []
    skipped_temp_excel_logs[root_dir] = []

    for file_path in sorted(root_dir.rglob("*")):
        if not file_path.is_file():
            continue

        suffix = file_path.suffix.lower()
        if suffix not in EXCEL_EXTENSIONS and suffix != ".xls":
            continue

        if file_path.name.startswith("~$"):
            skipped_temp_excel_logs[root_dir].append(str(file_path.relative_to(root_dir)))
            continue

        if suffix == ".xls":
            add_log(f"跳过文件：{file_path.name}，暂不支持 .xls")
            continue

        excel_files.append(file_path)

    return excel_files


def find_folder_excel_files(input_path: str | Path) -> list[Path]:
    path = Path(input_path).expanduser()
    if path.is_file():
        suffix = path.suffix.lower()
        if path.name.startswith("~$"):
            return []
        if suffix == ".xls":
            add_log(f"跳过文件：{path.name}，暂不支持 .xls")
            return []
        if suffix in EXCEL_EXTENSIONS:
            add_log(f"输入为单个 Excel 文件：{path}")
            return [path]
        return []

    if not path.is_dir():
        return []

    add_log(f"开始扫描文件夹里的 Excel：{path}")
    excluded_folder_names = {HC_FOLDER_NAME, build_skip_dir(path).name, UNCLASSIFIED_FOLDER_NAME}
    excel_files: list[Path] = []
    for file_path in sorted(path.rglob("*")):
        if not file_path.is_file():
            continue
        relative_parts = set(file_path.relative_to(path).parts[:-1])
        if relative_parts.intersection(excluded_folder_names):
            continue
        if file_path.name.startswith("~$"):
            continue
        suffix = file_path.suffix.lower()
        if suffix == ".xls":
            add_log(f"跳过文件：{file_path.name}，暂不支持 .xls")
            continue
        if suffix in EXCEL_EXTENSIONS:
            excel_files.append(file_path)
    add_log(f"找到文件夹 Excel：{len(excel_files)} 个")
    return excel_files


def is_hc_excel_file(path: Path) -> bool:
    return "hc" in path.name.casefold()


def split_hc_excel_files(excel_files: list[Path]) -> tuple[list[Path], list[Path]]:
    hc_files: list[Path] = []
    normal_files: list[Path] = []
    for excel_file in excel_files:
        if is_hc_excel_file(excel_file):
            hc_files.append(excel_file)
        else:
            normal_files.append(excel_file)
    return hc_files, normal_files


def safe_path_part(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", value).strip(" .")
    return cleaned or "_"


def unique_hc_target_path(hc_dir: Path, archive_path: Path, extracted_root: Path, hc_file: Path) -> Path:
    direct_target = hc_dir / hc_file.name
    if not direct_target.exists():
        return direct_target

    relative_parent = hc_file.parent.relative_to(extracted_root)
    folder_parts = [safe_path_part(archive_path.stem)]
    folder_parts.extend(safe_path_part(part) for part in relative_parent.parts)
    target_folder = hc_dir.joinpath(*folder_parts)
    target_path = target_folder / hc_file.name
    if not target_path.exists():
        return target_path

    index = 1
    while True:
        numbered_folder = hc_dir / f"{safe_path_part(archive_path.stem)}_{index}"
        numbered_path = numbered_folder.joinpath(*(safe_path_part(part) for part in relative_parent.parts)) / hc_file.name
        if not numbered_path.exists():
            return numbered_path
        index += 1


def build_hc_report_row(
    archive_path: Path,
    extracted_root: Path,
    hc_file: Path,
    target_path: Path,
    status: str,
    reason: str = "",
) -> dict[str, Any]:
    relative_path = hc_file.relative_to(extracted_root)
    relative_parent = relative_path.parent
    return {
        "外层压缩包名": archive_path.name,
        "子文件夹名": "" if str(relative_parent) == "." else str(relative_parent),
        "Excel文件名": hc_file.name,
        "压缩包内路径": str(relative_path),
        "处理状态": status,
        "目标路径": str(target_path.resolve()),
        "失败原因": reason,
    }


def copy_or_preview_hc_excel(archive_path: Path, extracted_root: Path, hc_file: Path, hc_dir: Path, dry_run: bool) -> dict[str, Any]:
    relative_path = hc_file.relative_to(extracted_root)
    target_path = hc_dir / hc_file.name
    if dry_run:
        target_path = unique_hc_target_path(hc_dir, archive_path, extracted_root, hc_file)
        add_log(f"[{archive_path.name}] dry-run 模式：HC 文件将复制到：{target_path}；来源：{relative_path}")
        return build_hc_report_row(archive_path, extracted_root, hc_file, target_path, "dry-run 预计复制")

    try:
        with _hc_copy_lock:
            target_path = unique_hc_target_path(hc_dir, archive_path, extracted_root, hc_file)
            target_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(hc_file, target_path)
        add_log(f"[{archive_path.name}] 已排除 HC 文件并复制到：{target_path}；来源：{relative_path}")
        return build_hc_report_row(archive_path, extracted_root, hc_file, target_path, "已复制")
    except OSError as exc:
        reason = str(exc)
        add_exception(f"异常：HC 文件复制失败但已排除：{archive_path.name} / {relative_path}，原因：{reason}")
        return build_hc_report_row(archive_path, extracted_root, hc_file, target_path, "复制失败，已排除", reason)


def build_hc_only_archive_detail(archive_path: Path, extracted_root: Path, hc_files: list[Path], reason: str) -> dict[str, Any]:
    related = "；".join(str(file.relative_to(extracted_root)) for file in hc_files)
    return {
        "外层压缩包名": archive_path.name,
        "子文件夹名": "",
        "压缩包名": archive_path.name,
        "压缩包路径": str(archive_path),
        "处理状态": "跳过",
        "正式Excel数量": 0,
        "正式Excel文件": "",
        "提取行数": 0,
        "识别品类": "",
        "识别日期": "",
        "是否跳过": "是",
        "异常原因": f"{reason}：{related}" if related else reason,
        "已复制到": "",
    }


def validate_excel_count(excel_files: list[Path], archive_path: Path, extracted_root: Path) -> bool:
    """
    判断正式 Excel 数量。
    如果数量为 1，返回 True。
    如果数量为 0，记录异常日志，返回 False。
    如果数量大于 1，记录异常日志，列出文件夹和文件名，返回 False，并跳过该压缩包。
    """
    if len(excel_files) == 1:
        return True

    temp_files = skipped_temp_excel_logs.get(extracted_root, [])
    if len(excel_files) == 0:
        all_excel_like = [
            file
            for file in extracted_root.rglob("*")
            if file.is_file() and file.suffix.lower() in EXCEL_EXTENSIONS.union({".xls"})
        ]
        if all_excel_like:
            detail = f"异常：{archive_path.name} 未找到正式 Excel 文件，已跳过"
            if temp_files:
                detail += "\n已跳过临时文件：\n" + "\n".join(f"{index}. {name}" for index, name in enumerate(temp_files, 1))
                add_structured_error(
                    archive_path,
                    "只找到临时Excel",
                    "未找到正式 Excel 文件，只找到临时文件",
                    "；".join(temp_files),
                )
            else:
                related = "；".join(str(file.relative_to(extracted_root)) for file in all_excel_like)
                add_structured_error(archive_path, "未找到正式Excel", "没有找到正式 Excel 文件", related)
            add_exception(detail)
        else:
            add_exception(f"异常：{archive_path.name} 未找到 Excel 文件，已跳过")
            add_structured_error(archive_path, "未找到Excel", "没有找到 Excel 文件")
        return False

    relative_files = [str(file.relative_to(extracted_root)) for file in excel_files]
    message = (
        f"异常：压缩包 {archive_path.name} 中识别到多个正式 Excel，已跳过该压缩包\n"
        f"识别位置：{extracted_root}\n"
        "正式 Excel 文件：\n"
        + "\n".join(f"{index}. {name}" for index, name in enumerate(relative_files, 1))
    )
    add_exception(message)
    add_structured_error(archive_path, "多个正式Excel", "识别到多个正式 Excel，已跳过该压缩包", "；".join(relative_files))
    return False


def group_excel_files_by_parent(excel_files: list[Path]) -> list[tuple[Path, list[Path]]]:
    groups: dict[Path, list[Path]] = {}
    for file_path in excel_files:
        groups.setdefault(file_path.parent, []).append(file_path)
    return [(folder, sorted(files)) for folder, files in sorted(groups.items(), key=lambda item: str(item[0]))]


def build_excel_processing_groups(extracted_root: Path, excel_files: list[Path]) -> list[tuple[Path, list[Path]]]:
    groups = group_excel_files_by_parent(excel_files)
    top_level_dirs = sorted(path for path in extracted_root.iterdir() if path.is_dir())
    if len(top_level_dirs) <= 1:
        return groups

    group_folders = {folder for folder, _ in groups}
    for folder in top_level_dirs:
        has_group_inside = any(existing == folder or folder in existing.parents for existing in group_folders)
        if not has_group_inside:
            groups.append((folder, []))
    return sorted(groups, key=lambda item: str(item[0]))

def build_recognition_names(
    excel_file: Path,
    extracted_root: Path,
    unit_folder: Path,
    archive_path: Path,
) -> list[str]:
    names: list[str] = []
    try:
        relative_parts = list(unit_folder.relative_to(extracted_root).parts)
    except ValueError:
        relative_parts = [unit_folder.name] if unit_folder.name else []
    names.extend(reversed(relative_parts))
    if archive_path.name != unit_folder.name:
        names.append(archive_path.name)
    if archive_path.parent.name and archive_path.parent != extracted_root:
        names.append(archive_path.parent.name)
    names.append(excel_file.name)
    return unique_names(names)


def normalize_header(value: Any) -> str:
    """
    标准化表头文字：
    去空格、去换行、去全角空格、英文转小写。
    """
    if value is None:
        return ""
    text = str(value)
    for item in (" ", "\u3000", "\n", "\r", "\t"):
        text = text.replace(item, "")
    return text.lower()


def identify_header_type(value: Any) -> str | None:
    normalized = normalize_header(value)
    if not normalized:
        return None

    if normalized in {"亚马逊订单号", "亚马逊", "订单号"}:
        return "order_id"
    if "亚马逊" in normalized and "订单号" in normalized:
        return "order_id"
    if normalized == "sku":
        return "sku"
    if normalized in {"数量", "qty", "quantity"}:
        return "quantity"
    return None


def build_merged_cell_value_map(ws) -> dict[tuple[int, int], Any]:
    merged_value_map: dict[tuple[int, int], Any] = {}
    for merged_range in getattr(ws.merged_cells, "ranges", []):
        top_left_value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        for row_index in range(merged_range.min_row, merged_range.max_row + 1):
            for col_index in range(merged_range.min_col, merged_range.max_col + 1):
                merged_value_map[(row_index, col_index)] = top_left_value
    return merged_value_map


def get_cell_value(ws, row_index: int, col_index: int, merged_value_map: dict[tuple[int, int], Any] | None = None) -> Any:
    if merged_value_map and (row_index, col_index) in merged_value_map:
        return merged_value_map[(row_index, col_index)]
    return ws.cell(row=row_index, column=col_index).value


def sheet_has_any_value(ws, merged_value_map: dict[tuple[int, int], Any] | None = None) -> bool:
    for row_index in range(1, (ws.max_row or 0) + 1):
        for col_index in range(1, (ws.max_column or 0) + 1):
            if has_value(get_cell_value(ws, row_index, col_index, merged_value_map)):
                return True
    return False


def find_header_row(ws, merged_value_map: dict[tuple[int, int], Any] | None = None) -> dict[str, int] | None:
    """
    扫描前 10 行，只要识别到任意核心表头就返回。
    """
    max_row = min(ws.max_row or 0, 10)
    for row_index in range(1, max_row + 1):
        found: dict[str, int] = {}
        for col_index in range(1, (ws.max_column or 0) + 1):
            header_type = identify_header_type(get_cell_value(ws, row_index, col_index, merged_value_map))
            if header_type and header_type not in found:
                found[header_type] = col_index
        if any(header_key in found for header_key in CORE_HEADER_KEYS):
            found["header_row"] = row_index
            return found
    return None


def format_debug_value(value: Any) -> str:
    if value in ("", None):
        return ""
    return str(value)


def format_sheet_debug_log(
    file_name: str,
    sheet_name: str,
    is_blank: bool,
    header: dict[str, int] | None,
    classification: str,
    extracted_rows_count: int,
    fail_reason: str = "",
) -> str:
    return "\n".join(
        [
            "========== DEBUG 表头识别 ==========",
            f"Excel 文件：{file_name}",
            f"Sheet 名：{sheet_name}",
            f"是否空白：{'是' if is_blank else '否'}",
            "识别到的表头：",
            f"- order_column: {format_debug_value((header or {}).get('order_id'))}",
            f"- sku_column: {format_debug_value((header or {}).get('sku'))}",
            f"- qty_column: {format_debug_value((header or {}).get('quantity'))}",
            f"最终 Sheet 分类：{classification}",
            f"提取到的数据行数：{extracted_rows_count}",
            "最终 Excel 是否被判定为失败：",
            f"- fail_reason: {fail_reason}",
        ]
    )


def make_output_date(month: int, day: int) -> date | None:
    try:
        return date(OUTPUT_DATE_YEAR, month, day)
    except ValueError:
        return None


def format_output_date_display(value: Any) -> str:
    if isinstance(value, datetime):
        return f"{value.month}月{value.day}日"
    if isinstance(value, date):
        return f"{value.month}月{value.day}日"
    return str(value)


def format_mmdd_text(value: str) -> str | None:
    month = int(value[:2])
    day = int(value[2:])
    if 1 <= month <= 12 and 1 <= day <= 31:
        return f"{month}月{day}日"
    return None


def parse_mmdd_date(value: str) -> date | None:
    return make_output_date(int(value[:2]), int(value[2:]))


def parse_month_day_date(month_text: str, day_text: str) -> date | None:
    month = int(month_text)
    day = int(day_text)
    return make_output_date(month, day)


def strip_leading_sequence_prefix(stem: str) -> str:
    """
    去掉文件名前面的批次/序号前缀，让 45-4.17 或 33~35-0418 仍能识别真实日期。
    """
    stripped = re.sub(r"^\s*\d{1,3}\s*(?:[~～]\s*\d{1,3})?\s*[-_]\s*", "", stem, count=1)
    stripped = re.sub(r"^\s*\d{1,3}\s*\.\s*(?=\d{1,2}\.\d{1,2}(?!\d))", "", stripped, count=1)
    return stripped


def parse_date_from_filename(filename: str) -> date | str:
    """
    从 Excel 文件名识别日期。
    例如 0507 -> date(OUTPUT_DATE_YEAR, 5, 7)。
    日期范围保留为文本，例如 0501-0503 -> 5月1日-5月3日。
    识别不到则返回空字符串。
    """
    stem = Path(filename).stem
    search_stems = [stem]
    stripped_stem = strip_leading_sequence_prefix(stem)
    if stripped_stem != stem:
        search_stems.append(stripped_stem)

    range_match = re.search(r"(?<!\d)(\d{4})\s*[-_至到]\s*(\d{4})(?!\d)", stem)
    if range_match:
        start = format_mmdd_text(range_match.group(1))
        end = format_mmdd_text(range_match.group(2))
        if start and end:
            return f"{start}-{end}"

    year_match = re.search(r"(?<!\d)(?:20\d{2})[\._\-年](\d{1,2})[\._\-月](\d{1,2})(?:日)?(?!\d)", stem)
    if year_match:
        parsed = parse_month_day_date(year_match.group(1), year_match.group(2))
        if parsed:
            return parsed

    for candidate_stem in search_stems:
        separated_match = re.search(r"(?<!\d)(\d{1,2})[\._\-月](\d{1,2})(?:日)?(?!\d)", candidate_stem)
        if separated_match:
            parsed = parse_month_day_date(separated_match.group(1), separated_match.group(2))
            if parsed:
                return parsed

    start_match = re.match(r"^(\d{4})(?!\d)", stem)
    if start_match:
        parsed = parse_mmdd_date(start_match.group(1))
        if parsed:
            return parsed

    any_match = re.search(r"(?<!\d)(\d{4})(?!\d)", stem)
    if any_match:
        parsed = parse_mmdd_date(any_match.group(1))
        if parsed:
            return parsed

    return ""


def first_detected_date_from_names(names: list[str]) -> date | str:
    for name in names:
        parsed = parse_date_from_filename(name)
        if parsed:
            return parsed
    return ""


def detect_category_from_filename(filename: str, category_keywords: dict[str, list[str]] | None = None) -> str:
    """
    根据 Excel 文件名中的关键词识别品类。
    如果没有命中任何关键词，返回 '未分类'。
    如果命中多个，按关键词长度和配置顺序决定最终品类。
    """
    matches: list[tuple[int, int, str, str]] = []
    keywords_config = category_keywords or CATEGORY_KEYWORDS
    for category_index, (category, keywords) in enumerate(keywords_config.items()):
        for keyword in keywords:
            if keyword in filename:
                matches.append((len(keyword), category_index, category, keyword))

    if not matches:
        add_log("未命中品类关键词，写入工作表：未分类")
        return "未分类"

    matches.sort(key=lambda item: (-item[0], item[1]))
    selected = matches[0]

    unique_keywords = []
    seen_keywords = set()
    for _, _, _, keyword in matches:
        if keyword not in seen_keywords:
            seen_keywords.add(keyword)
            unique_keywords.append(keyword)

    if len(matches) > 1:
        add_log(f"文件名命中多个品类关键词：{'、'.join(unique_keywords)}")
        add_log(f"最终选择品类：{selected[2]}")
    else:
        add_log(f"命中品类关键词：{selected[3]}")

    return selected[2]


def strip_known_source_suffix(name: str) -> str:
    path = Path(str(name))
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls", ".zip", ".rar", ".7z"}:
        return path.with_suffix("").name
    return path.name


def split_candidate_name_parts(name: str) -> list[str]:
    stem = strip_known_source_suffix(name)
    normalized = re.sub(r"[＿_]+", "-", stem)
    normalized = re.sub(r"\s+", "", normalized)
    return [part.strip(" -—–~") for part in re.split(r"[-—–]+", normalized) if part.strip(" -—–~")]


def is_sequence_part(part: str) -> bool:
    return bool(re.fullmatch(r"\d{1,3}(?:[~～]\d{1,3})?", part))


def is_date_part(part: str) -> bool:
    if parse_date_from_filename(part):
        return True
    return bool(re.fullmatch(r"\d{1,2}[.月]\d{1,2}(?:日)?", part))


def is_quantity_part(part: str) -> bool:
    return bool(re.fullmatch(r".*\d+\s*(?:单|个|件|pcs?|orders?|qty).*", part, flags=re.IGNORECASE))


def is_candidate_noise_part(part: str) -> bool:
    text = part.strip()
    if not text:
        return True
    return is_sequence_part(text) or is_date_part(text) or is_quantity_part(text)


def clean_candidate_with_prefix(raw_candidate: str, prefixes: list[str]) -> tuple[str, str]:
    candidate = raw_candidate.strip(" -—–")
    for prefix in sorted((item.strip() for item in prefixes if item.strip()), key=len, reverse=True):
        separated = f"{prefix}-"
        if candidate.startswith(separated):
            cleaned = candidate[len(separated):].strip(" -—–")
            return prefix, cleaned
        next_char = candidate[len(prefix):len(prefix) + 1]
        if candidate.startswith(prefix) and next_char and not next_char.isascii():
            cleaned = candidate[len(prefix):].strip(" -—–")
            return prefix, cleaned
    return "", candidate


def build_category_candidate_from_name(name: str, prefixes: list[str] | None = None) -> dict[str, str] | None:
    parts = split_candidate_name_parts(name)
    kept = [part for part in parts if not is_candidate_noise_part(part)]
    if not kept:
        return None
    raw_candidate = "-".join(kept).strip(" -—–")
    if not raw_candidate or not re.search(r"[^\d\s~～.\-—–]", raw_candidate):
        return None
    prefix, category = clean_candidate_with_prefix(raw_candidate, prefixes or [])
    if not category:
        return None
    return {
        "source_name": str(name),
        "raw_candidate": raw_candidate,
        "prefix": prefix,
        "category": category,
    }


def first_category_candidate_from_names(names: list[str], prefixes: list[str] | None = None) -> dict[str, str] | None:
    for name in names:
        candidate = build_category_candidate_from_name(name, prefixes or [])
        if candidate:
            return candidate
    return None


def first_detected_category_from_names(
    names: list[str],
    category_keywords: dict[str, list[str]] | None = None,
) -> tuple[str, str]:
    for name in names:
        category = detect_category_from_filename(name, category_keywords)
        if category != "未分类":
            return category, name
    return "未分类", names[0] if names else ""


def should_skip_named_file(path: Path) -> bool:
    return path.stem.startswith(SKIP_FILE_PREFIXES)


def skip_named_file_reason(path: Path) -> str:
    for prefix in SKIP_FILE_PREFIXES:
        if path.stem.startswith(prefix):
            return f"文件名前两个字为“{prefix}”"
    return ""


def unique_names(names: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for name in names:
        text = str(name or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def parse_expected_counts_from_filename(filename: str) -> tuple[int | None, int | None, str]:
    """
    从文件名中提取“预计单量”和“预计数量”。
    只识别带单位的数字，避免把 0507 这类日期误判为单量或数量。
    """
    stem = Path(filename).stem
    quantity_units_pattern = "|".join(re.escape(unit) for unit in EXPECTED_QUANTITY_UNITS)
    order_match = re.search(r"(?<!\d)(\d+)\s*单(?:量)?", stem)
    quantity_match = re.search(rf"(?<!\d)(\d+)\s*(?:{quantity_units_pattern})(?![A-Za-z])", stem, re.IGNORECASE)
    expected_orders = int(order_match.group(1)) if order_match else None
    expected_quantity = int(quantity_match.group(1)) if quantity_match else None

    if expected_orders is None and expected_quantity is None:
        note = "未识别到文件名单量 / 数量"
    elif expected_orders is None:
        note = "未识别到文件名单量"
    elif expected_quantity is None:
        note = "未识别到文件名数量"
    else:
        note = ""
    return expected_orders, expected_quantity, note


def parse_quantity_integer(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        number = None
    if number is not None and number.is_integer():
        return int(number)
    match = re.search(r"[-+]?\d+", text)
    if match:
        return int(match.group(0))
    return None


def quantity_to_number(value: Any) -> float | None:
    parsed = parse_quantity_integer(value)
    if parsed is None:
        return None
    return float(parsed)


def format_quantity_total(value: float) -> int | float:
    if float(value).is_integer():
        return int(value)
    return value


def build_filename_validation(
    archive_name: str,
    excel_name: str,
    rows: list[dict[str, Any]],
    category: str,
    date_text: date | str,
    subfolder_name: str = "",
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    expected_orders, expected_quantity, note = parse_expected_counts_from_filename(excel_name)
    actual_orders = len({row_order_id(row) for row in rows if row_order_id(row)})

    actual_quantity_total = 0.0
    invalid_quantity_rows: list[dict[str, Any]] = []
    for row in rows:
        raw_quantity = row.get("数量")
        number = quantity_to_number(raw_quantity)
        if number is None:
            if has_value(raw_quantity):
                invalid_quantity_rows.append(row)
            continue
        actual_quantity_total += number

    actual_quantity = format_quantity_total(actual_quantity_total)
    if expected_orders is None:
        order_result = "未识别"
    else:
        order_result = "匹配" if expected_orders == actual_orders else "不匹配"

    if expected_quantity is None:
        quantity_result = "未识别"
    elif invalid_quantity_rows:
        quantity_result = "无法计算"
    else:
        quantity_result = "匹配" if expected_quantity == actual_quantity else "不匹配"

    notes = [note] if note else []
    if invalid_quantity_rows:
        notes.append(f"发现 {len(invalid_quantity_rows)} 行数量无法转成数字，未参与数量求和")
    if order_result == "不匹配" or quantity_result in {"不匹配", "无法计算"}:
        add_log(
            f"警告：{excel_name} 文件名预计 "
            f"{expected_orders if expected_orders is not None else '未识别'} 单 "
            f"{expected_quantity if expected_quantity is not None else '未识别'} 个，实际提取 "
            f"{actual_orders} 单 {actual_quantity} 个，"
            f"单量校验：{order_result}，数量校验：{quantity_result}"
        )

    validation = {
        "外层压缩包名": archive_name,
        "子文件夹名": subfolder_name,
        "压缩包名": archive_name,
        "Excel文件名": excel_name,
        "品类": category,
        "日期": format_output_date_display(date_text) if date_text else "",
        "文件名预计单量": expected_orders if expected_orders is not None else "",
        "实际提取单量": actual_orders,
        "单量校验结果": order_result,
        "文件名预计数量": expected_quantity if expected_quantity is not None else "",
        "实际提取数量": actual_quantity,
        "数量校验结果": quantity_result,
        "提取行数": len(rows),
        "备注": "；".join(notes),
    }
    return validation, invalid_quantity_rows


def normalize_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip() != ""
    return True


def normalize_quantity(value: Any, excel_name: str, row_index: int) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        add_log(f"数量无法转成整数，保留原值：{excel_name} 第 {row_index} 行，值={value}")
        return value
    if isinstance(value, str) and value.strip() == "":
        return ""
    parsed = parse_quantity_integer(value)
    if parsed is not None:
        return parsed
    add_log(f"数量无法转成整数，保留原值：{excel_name} 第 {row_index} 行，值={value}")
    return value


def extract_rows_from_workbook(
    file_path: Path,
    category_keywords: dict[str, list[str]] | None = None,
    recognition_names: list[str] | None = None,
    category_prefixes: list[str] | None = None,
) -> tuple[list[dict[str, Any]], bool, str, date | str, list[dict[str, Any]], dict[str, str] | None]:
    """
    从单个 Excel 文件中提取数据。
    """
    names = unique_names(list(recognition_names or []) + [file_path.name])
    date_text = first_detected_date_from_names(names)
    if date_text:
        add_log(f"识别日期：{format_output_date_display(date_text)}")
    else:
        add_log(f"未识别到日期：{file_path.name}，日期列留空")

    category, category_source_name = first_detected_category_from_names(names, category_keywords)
    if category_source_name and category_source_name != file_path.name and category != "未分类":
        add_log(f"品类按文件夹或外层名称识别：{category_source_name}")
    add_log(f"识别品类：{category}")
    category_candidate = None
    if category == "未分类":
        category_candidate = first_category_candidate_from_names(names, category_prefixes or [])
        if category_candidate and not any(not char.isascii() for char in category_candidate.get("category", "")):
            category_candidate = None
        if category_candidate:
            add_log(f"发现待确认品类候选：{category_candidate.get('category', '')}（来源：{category_candidate.get('source_name', '')}）")

    try:
        workbook = load_workbook(file_path, data_only=True, read_only=False)
    except Exception as exc:
        add_exception(f"异常：打开 Excel 失败 {file_path.name}，原因：{exc}")
        add_structured_error(None, "打开Excel失败", f"打开 Excel 失败 {file_path.name}，原因：{exc}", file_path.name)
        return [], False, category, date_text, [], category_candidate

    rows: list[dict[str, Any]] = []
    header_report_rows: list[dict[str, Any]] = []
    sheet_debug_entries: list[dict[str, Any]] = []
    header_found = False
    nonempty_sheets_without_headers: list[str] = []
    try:
        for ws in workbook.worksheets:
            merged_value_map = build_merged_cell_value_map(ws)
            if not sheet_has_any_value(ws, merged_value_map):
                add_log(f"跳过空白 Sheet：{ws.title}")
                sheet_debug_entries.append(
                    {
                        "file_name": file_path.name,
                        "sheet_name": ws.title,
                        "is_blank": True,
                        "header": None,
                        "classification": "空白 Sheet",
                        "extracted_rows_count": 0,
                    }
                )
                continue

            header = find_header_row(ws, merged_value_map)
            if not header:
                add_log(f"未识别到任何核心表头：{file_path.name} / {ws.title}")
                nonempty_sheets_without_headers.append(ws.title)
                sheet_debug_entries.append(
                    {
                        "file_name": file_path.name,
                        "sheet_name": ws.title,
                        "is_blank": False,
                        "header": None,
                        "classification": "未识别核心表头",
                        "extracted_rows_count": 0,
                    }
                )
                continue
            header_found = True
            sheet_row_count_before = len(rows)
            recognized_headers = [label for key, label in CORE_HEADER_LABELS.items() if key in header]
            missing_headers = [label for key, label in CORE_HEADER_LABELS.items() if key not in header]
            if missing_headers:
                add_log(
                    f"警告：缺少 {' / '.join(missing_headers)} 表头，已按空值处理："
                    f"{file_path.name} / {ws.title}；"
                    f"已识别：{' / '.join(recognized_headers)}"
                )
                header_report_rows.append(
                    {
                        "Excel文件名": file_path.name,
                        "Sheet名": ws.title,
                        "已识别表头": " / ".join(recognized_headers),
                        "缺失表头": " / ".join(missing_headers),
                        "处理方式": "缺失字段已留空，不跳过整个文件",
                    }
                )

            for row_index in range(header["header_row"] + 1, (ws.max_row or 0) + 1):
                order_id = (
                    get_cell_value(ws, row_index, header["order_id"], merged_value_map)
                    if "order_id" in header
                    else ""
                )
                sku = get_cell_value(ws, row_index, header["sku"], merged_value_map) if "sku" in header else ""
                quantity = (
                    get_cell_value(ws, row_index, header["quantity"], merged_value_map)
                    if "quantity" in header
                    else ""
                )

                recognized_values = []
                if "order_id" in header:
                    recognized_values.append(order_id)
                if "sku" in header:
                    recognized_values.append(sku)
                if "quantity" in header:
                    recognized_values.append(quantity)
                if not any(has_value(value) for value in recognized_values):
                    continue

                rows.append(
                    {
                        "category": category,
                        "压缩包名": getattr(_thread_context, "archive_path", Path("")).name,
                        "Excel 文件名": file_path.name,
                        "Sheet 名": ws.title,
                        "源数据行号": row_index,
                        "亚马逊订单号": normalize_cell_value(order_id),
                        "SKU": normalize_cell_value(sku),
                        "原始数量": normalize_cell_value(quantity),
                        "数量": normalize_quantity(quantity, file_path.name, row_index),
                        "日期": date_text,
                    }
                )
            if len(rows) == sheet_row_count_before:
                add_log(f"未提取到有效数据：{file_path.name} / {ws.title}")
            sheet_debug_entries.append(
                {
                    "file_name": file_path.name,
                    "sheet_name": ws.title,
                    "is_blank": False,
                    "header": dict(header),
                    "classification": "可处理 Sheet",
                    "extracted_rows_count": len(rows) - sheet_row_count_before,
                }
            )
    finally:
        workbook.close()

    fail_reason = ""
    if not header_found:
        if not nonempty_sheets_without_headers:
            for entry in sheet_debug_entries:
                add_debug_log(format_sheet_debug_log(**entry, fail_reason=fail_reason))
            add_log(f"未提取到有效数据：{file_path.name}")
            return [], True, category, date_text, header_report_rows, category_candidate
        fail_reason = "未识别到任何核心表头"
        for entry in sheet_debug_entries:
            add_debug_log(format_sheet_debug_log(**entry, fail_reason=fail_reason))
        related = f"{file_path.name} / " + "；".join(nonempty_sheets_without_headers)
        add_exception(f"异常：{file_path.name} 未识别到任何核心表头，已跳过")
        add_structured_error(None, "未识别到任何核心表头", "未识别到任何核心表头：亚马逊订单号 / SKU / 数量", related)
        return [], False, category, date_text, header_report_rows, category_candidate

    if nonempty_sheets_without_headers:
        fail_reason = "部分 Sheet 未识别到任何核心表头"
        related = f"{file_path.name} / " + "；".join(nonempty_sheets_without_headers)
        add_exception(f"异常：{file_path.name} 部分工作表未识别到任何核心表头，已跳过相关 Sheet")
        add_structured_error(None, "未识别到任何核心表头", "未识别到任何核心表头：亚马逊订单号 / SKU / 数量", related)

    for entry in sheet_debug_entries:
        add_debug_log(format_sheet_debug_log(**entry, fail_reason=fail_reason))
    add_log(f"提取数据：{len(rows)} 行")
    return rows, True, category, date_text, header_report_rows, category_candidate


def process_excel_unit(
    archive_path: Path,
    extracted_root: Path,
    excel_files: list[Path],
    unit_folder: Path,
    category_keywords: dict[str, list[str]] | None = None,
    category_prefixes: list[str] | None = None,
    skip_dir: Path | None = None,
    unclassified_dir: Path | None = None,
    dry_run: bool = False,
    excel_group_mode: str = EXCEL_GROUP_SINGLE,
    copy_skipped_excel_file: bool = False,
) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    quantity_error_rows: list[dict[str, Any]] = []
    header_report_rows: list[dict[str, Any]] = []
    filename_validations: list[dict[str, Any]] = []
    category_candidates: list[dict[str, str]] = []
    excel_count = len(excel_files)
    selected_excel_name = "；".join(str(file.relative_to(extracted_root)) for file in excel_files)
    category = ""
    date_text = ""
    filename_validation: dict[str, Any] | None = None
    subfolder_name = "" if unit_folder == extracted_root else unit_folder.name

    def build_result(
        success: bool,
        status: str,
        skip: bool = False,
        reason: str = "",
        ignored: bool = False,
        copied_to: str = "",
    ) -> dict[str, Any]:
        archive_detail = {
            "外层压缩包名": archive_path.name,
            "子文件夹名": subfolder_name,
            "压缩包名": archive_path.name,
            "压缩包路径": str(archive_path),
            "处理状态": status,
            "正式Excel数量": excel_count,
            "正式Excel文件": selected_excel_name,
            "提取行数": len(rows),
            "识别品类": category,
            "识别日期": format_output_date_display(date_text) if date_text else "",
            "是否跳过": "是" if skip else "否",
            "异常原因": reason,
            "已复制到": copied_to,
        }
        if filename_validation:
            archive_detail.update(
                {
                    "文件名预计单量": filename_validation.get("文件名预计单量", ""),
                    "实际提取单量": filename_validation.get("实际提取单量", ""),
                    "单量校验结果": filename_validation.get("单量校验结果", ""),
                    "文件名预计数量": filename_validation.get("文件名预计数量", ""),
                    "实际提取数量": filename_validation.get("实际提取数量", ""),
                    "数量校验结果": filename_validation.get("数量校验结果", ""),
                }
            )
        return {
            "success": success,
            "rows": list(rows),
            "archive_detail": archive_detail,
            "filename_validation": dict(filename_validation or {}),
            "filename_validations": list(filename_validations),
            "quantity_error_rows": list(quantity_error_rows),
            "header_report_rows": list(header_report_rows),
            "category_candidates": list(category_candidates),
            "ignored": ignored,
        }

    def record_category_candidate(candidate: dict[str, str] | None, excel_file: Path) -> None:
        if not candidate:
            return
        enriched = dict(candidate)
        enriched.update(
            {
                "status": "待确认",
                "archive_name": archive_path.name,
                "excel_file": excel_file.name,
                "source_path": str(excel_file),
            }
        )
        category_candidates.append(enriched)
        copied_to = copy_skipped_source(excel_file, unclassified_dir, dry_run=dry_run)
        if copied_to:
            add_log(f"[{archive_path.name}] 未分类 Excel 已复制到：{copied_to}")

    def skip_excel_file(excel_file: Path, reason: str, skip_type: str) -> str:
        skipped_source = excel_file if copy_skipped_excel_file else archive_path
        copied_to = copy_skipped_source(skipped_source, skip_dir, dry_run=dry_run)
        add_log(f"[{archive_path.name}] 已跳过：{excel_file.name}，原因：{reason}")
        add_structured_error(
            archive_path,
            skip_type,
            f"{reason}；原文件名：{excel_file.name}；已复制到：{copied_to}",
            excel_file.name,
            status="已跳过",
        )
        return copied_to

    if excel_count == 0:
        reason = "子文件夹未找到正式 Excel"
        add_exception(f"异常：{archive_path.name} / {subfolder_name or '.'} 未找到正式 Excel，已跳过")
        add_structured_error(archive_path, "子文件夹未找到正式Excel", reason, subfolder_name)
        return build_result(False, "跳过", skip=True, reason=reason)

    if excel_count > 1 and excel_group_mode == EXCEL_GROUP_SINGLE:
        reason = "子文件夹中识别到多个正式 Excel"
        related = "；".join(str(file.relative_to(extracted_root)) for file in excel_files)
        add_exception(f"异常：{archive_path.name} / {subfolder_name or '.'} 中识别到多个正式 Excel，已跳过该子文件夹\n{related}")
        add_structured_error(archive_path, "子文件夹多个正式Excel", reason, related)
        return build_result(False, "跳过", skip=True, reason=reason)

    if excel_group_mode == EXCEL_GROUP_MULTI:
        copied_to = ""
        skipped_reasons: list[str] = []
        for excel_file in excel_files:
            if subfolder_name:
                add_log(f"子文件夹：{subfolder_name}")
            add_log(f"[{archive_path.name}] 找到正式 Excel：{excel_file.name}")
            recognition_names = build_recognition_names(excel_file, extracted_root, unit_folder, archive_path)
            if should_skip_named_file(excel_file):
                reason = skip_named_file_reason(excel_file)
                skipped_reasons.append(reason)
                copied_to = skip_excel_file(excel_file, reason, "前缀文件跳过")
                continue

            extracted_rows, workbook_success, current_category, current_date_text, current_header_rows, current_category_candidate = extract_rows_from_workbook(
                excel_file,
                category_keywords,
                recognition_names,
                category_prefixes,
            )
            record_category_candidate(current_category_candidate, excel_file)
            for row in current_header_rows:
                row["压缩包名"] = archive_path.name
                row["子文件夹名"] = subfolder_name
            header_report_rows.extend(current_header_rows)
            for row in extracted_rows:
                row["外层压缩包名"] = archive_path.name
                row["子文件夹名"] = subfolder_name
                row["source_path"] = str(excel_file)
            if not workbook_success:
                continue
            if current_category == "未分类":
                copy_unclassified_excel(excel_file, unclassified_dir, dry_run=dry_run)

            current_validation, current_quantity_errors = build_filename_validation(
                archive_path.name,
                excel_file.name,
                extracted_rows,
                current_category,
                current_date_text,
                subfolder_name,
            )
            filename_validations.append(current_validation)
            if filename_validation is None:
                filename_validation = current_validation
            quantity_error_rows.extend(current_quantity_errors)
            if not category:
                category = current_category
            if not date_text:
                date_text = current_date_text
            rows.extend(extracted_rows)

        if rows:
            add_log(f"[{archive_path.name}] 目标工作表：{rows[0]['category']}")
            return build_result(True, "成功", copied_to=copied_to)
        if skipped_reasons:
            return build_result(False, "跳过", skip=True, reason="；".join(skipped_reasons), ignored=True, copied_to=copied_to)
        return build_result(False, "异常", skip=True, reason="Excel 读取失败或未识别到任何核心表头")

    excel_file = excel_files[0]
    selected_excel_name = excel_file.name
    if subfolder_name:
        add_log(f"子文件夹：{subfolder_name}")
    add_log(f"[{archive_path.name}] 找到正式 Excel：{excel_file.name}")
    recognition_names = build_recognition_names(excel_file, extracted_root, unit_folder, archive_path)
    if should_skip_named_file(excel_file):
        reason = skip_named_file_reason(excel_file)
        copied_to = skip_excel_file(excel_file, reason, "前缀文件跳过")
        return build_result(False, "跳过", skip=True, reason=reason, ignored=True, copied_to=copied_to)

    rows, workbook_success, category, date_text, header_report_rows, category_candidate = extract_rows_from_workbook(
        excel_file,
        category_keywords,
        recognition_names,
        category_prefixes,
    )
    record_category_candidate(category_candidate, excel_file)
    for row in header_report_rows:
        row["压缩包名"] = archive_path.name
        row["子文件夹名"] = subfolder_name
    for row in rows:
        row["外层压缩包名"] = archive_path.name
        row["子文件夹名"] = subfolder_name
        row["source_path"] = str(excel_file)
    if not workbook_success:
        return build_result(False, "异常", skip=True, reason="Excel 读取失败或未识别到任何核心表头")
    if category == "未分类":
        copy_unclassified_excel(excel_file, unclassified_dir, dry_run=dry_run)

    filename_validation, quantity_error_rows = build_filename_validation(
        archive_path.name,
        excel_file.name,
        rows,
        category,
        date_text,
        subfolder_name,
    )
    filename_validations.append(filename_validation)
    if rows:
        add_log(f"[{archive_path.name}] 目标工作表：{rows[0]['category']}")
    return build_result(True, "成功")


def process_archive(
    archive_path: Path,
    category_keywords: dict[str, list[str]] | None = None,
    category_prefixes: list[str] | None = None,
    skip_dir: Path | None = None,
    unclassified_dir: Path | None = None,
    hc_dir: Path | None = None,
    dry_run: bool = False,
    enable_hc_filter: bool = False,
    excel_group_mode: str = EXCEL_GROUP_SINGLE,
) -> dict[str, Any]:
    """
    处理单个压缩包：
    1. 解压压缩包
    2. 递归查找正式 Excel
    3. 忽略 ~$ 临时 Excel
    4. 如果正式 Excel 数量不是 1，则记录异常并跳过
    5. 从唯一正式 Excel 中提取订单数据
    6. 返回处理结果
    """
    archive_process_logs: list[str] = []
    archive_exception_logs: list[str] = []
    archive_error_report_rows: list[dict[str, Any]] = []
    archive_debug_logs: list[str] = []
    set_thread_log_context(
        archive_path,
        archive_process_logs,
        archive_exception_logs,
        archive_error_report_rows,
        archive_debug_logs,
    )
    rows: list[dict[str, Any]] = []
    archive_details: list[dict[str, Any]] = []
    filename_validations: list[dict[str, Any]] = []
    quantity_error_rows: list[dict[str, Any]] = []
    header_report_rows: list[dict[str, Any]] = []
    hc_report_rows: list[dict[str, Any]] = []
    category_candidates: list[dict[str, str]] = []
    ignored_units = 0
    success_units = 0

    def build_result(success: bool, ignored: bool = False, reason: str = "") -> dict[str, Any]:
        if not archive_details:
            archive_details.append(
                {
                    "外层压缩包名": archive_path.name,
                    "子文件夹名": "",
                    "压缩包名": archive_path.name,
                    "压缩包路径": str(archive_path),
                    "处理状态": "异常" if not ignored else "跳过",
                    "正式Excel数量": 0,
                    "正式Excel文件": "",
                    "提取行数": 0,
                    "识别品类": "",
                    "识别日期": "",
                    "是否跳过": "是",
                    "异常原因": reason or (archive_exception_logs[-1] if archive_exception_logs else ""),
                    "已复制到": "",
                }
            )
        return {
            "archive_path": archive_path,
            "archive_name": archive_path.name,
            "success": success,
            "rows": list(rows),
            "exception_logs": list(archive_exception_logs),
            "process_logs": list(archive_process_logs),
            "debug_logs": list(archive_debug_logs),
            "error_report_rows": list(archive_error_report_rows),
            "archive_details": list(archive_details),
            "filename_validations": list(filename_validations),
            "quantity_error_rows": list(quantity_error_rows),
            "header_report_rows": list(header_report_rows),
            "hc_report_rows": list(hc_report_rows),
            "category_candidates": list(category_candidates),
            "ignored": ignored,
        }

    try:
        if should_skip_named_file(archive_path):
            reason = skip_named_file_reason(archive_path)
            copied_to = copy_skipped_source(archive_path, skip_dir, dry_run=dry_run)
            add_log(f"[{archive_path.name}] 已跳过压缩包：{reason}；已复制到：{copied_to}")
            add_structured_error(
                archive_path,
                "前缀压缩包跳过",
                f"{reason}；已复制到：{copied_to}",
                archive_path.name,
                status="已跳过",
            )
            archive_details.append(
                {
                    "外层压缩包名": archive_path.name,
                    "子文件夹名": "",
                    "压缩包名": archive_path.name,
                    "压缩包路径": str(archive_path),
                    "处理状态": "跳过",
                    "正式Excel数量": 0,
                    "正式Excel文件": "",
                    "提取行数": 0,
                    "识别品类": "",
                    "识别日期": "",
                    "是否跳过": "是",
                    "异常原因": reason,
                    "已复制到": copied_to,
                }
            )
            return build_result(False, ignored=True, reason=reason)

        add_log(f"[{archive_path.name}] 正在处理压缩包：{archive_path.name}")
        extracted_root = extract_archive(archive_path)
        if extracted_root is None:
            return build_result(False)

        try:
            excel_files = find_excel_files_in_extracted_dir(extracted_root)
            if enable_hc_filter:
                hc_files, normal_excel_files = split_hc_excel_files(excel_files)
            else:
                hc_files, normal_excel_files = [], excel_files
            if hc_files and hc_dir is not None:
                for hc_file in hc_files:
                    hc_report_rows.append(copy_or_preview_hc_excel(archive_path, extracted_root, hc_file, hc_dir, dry_run))

            if not normal_excel_files:
                if hc_files:
                    reason = "仅发现 HC 文件，已排除"
                    add_log(f"[{archive_path.name}] {reason}")
                    archive_details.append(build_hc_only_archive_detail(archive_path, extracted_root, hc_files, reason))
                    return build_result(False, ignored=True, reason=reason)
                if not validate_excel_count(normal_excel_files, archive_path, extracted_root):
                    return build_result(False)

            hc_parent_folders = {file.parent for file in hc_files}
            groups = build_excel_processing_groups(extracted_root, normal_excel_files)
            if len(groups) > 1:
                add_log("发现压缩包内包含多个订单子文件夹，已按子文件夹拆分处理")

            for unit_folder, unit_excel_files in groups:
                if not unit_excel_files and unit_folder in hc_parent_folders:
                    folder_hc_files = [file for file in hc_files if file.parent == unit_folder]
                    reason = "子文件夹仅发现 HC 文件，已排除"
                    add_log(f"[{archive_path.name}] {unit_folder.name}：{reason}")
                    archive_details.append(build_hc_only_archive_detail(archive_path, extracted_root, folder_hc_files, reason))
                    ignored_units += 1
                    continue
                unit_result = process_excel_unit(
                    archive_path,
                    extracted_root,
                    unit_excel_files,
                    unit_folder,
                    category_keywords,
                    category_prefixes,
                    skip_dir,
                    unclassified_dir,
                    dry_run,
                    excel_group_mode=excel_group_mode,
                )
                rows.extend(unit_result.get("rows") or [])
                if unit_result.get("archive_detail"):
                    archive_details.append(unit_result["archive_detail"])
                if unit_result.get("filename_validations"):
                    filename_validations.extend(unit_result["filename_validations"])
                elif unit_result.get("filename_validation"):
                    filename_validations.append(unit_result["filename_validation"])
                quantity_error_rows.extend(unit_result.get("quantity_error_rows") or [])
                header_report_rows.extend(unit_result.get("header_report_rows") or [])
                category_candidates.extend(unit_result.get("category_candidates") or [])
                if unit_result.get("ignored"):
                    ignored_units += 1
                elif unit_result.get("success"):
                    success_units += 1

            return build_result(success_units > 0, ignored=(ignored_units > 0 and success_units == 0))
        finally:
            shutil.rmtree(extracted_root, ignore_errors=True)
    except Exception as exc:
        add_exception(f"异常：处理压缩包失败 {archive_path.name}，原因：{exc}")
        add_structured_error(archive_path, "处理失败", f"处理压缩包失败，原因：{exc}")
        return build_result(False, reason=f"处理压缩包失败，原因：{exc}")
    finally:
        clear_thread_log_context()


def build_folder_processing_groups(input_root: Path, excel_files: list[Path]) -> list[tuple[Path, list[Path]]]:
    grouped: dict[Path, list[Path]] = {}
    for excel_file in excel_files:
        grouped.setdefault(excel_file.parent, []).append(excel_file)
    return [(folder, sorted(files)) for folder, files in sorted(grouped.items(), key=lambda item: str(item[0]))]


def process_folder_excel_group(
    input_root: Path,
    unit_folder: Path,
    excel_files: list[Path],
    category_keywords: dict[str, list[str]] | None = None,
    category_prefixes: list[str] | None = None,
    skip_dir: Path | None = None,
    unclassified_dir: Path | None = None,
    hc_dir: Path | None = None,
    dry_run: bool = False,
    enable_hc_filter: bool = False,
    excel_group_mode: str = EXCEL_GROUP_SINGLE,
) -> dict[str, Any]:
    virtual_archive = unit_folder
    folder_process_logs: list[str] = []
    folder_exception_logs: list[str] = []
    folder_error_report_rows: list[dict[str, Any]] = []
    folder_debug_logs: list[str] = []
    set_thread_log_context(
        virtual_archive,
        folder_process_logs,
        folder_exception_logs,
        folder_error_report_rows,
        folder_debug_logs,
    )
    archive_details: list[dict[str, Any]] = []
    hc_report_rows: list[dict[str, Any]] = []

    def build_result(result: dict[str, Any], ignored: bool = False) -> dict[str, Any]:
        return {
            "archive_path": virtual_archive,
            "archive_name": virtual_archive.name,
            "success": bool(result.get("success")),
            "rows": list(result.get("rows") or []),
            "exception_logs": list(folder_exception_logs),
            "process_logs": list(folder_process_logs),
            "debug_logs": list(folder_debug_logs),
            "error_report_rows": list(folder_error_report_rows),
            "archive_details": list(archive_details or ([result["archive_detail"]] if result.get("archive_detail") else [])),
            "filename_validations": list(result.get("filename_validations") or ([result["filename_validation"]] if result.get("filename_validation") else [])),
            "quantity_error_rows": list(result.get("quantity_error_rows") or []),
            "header_report_rows": list(result.get("header_report_rows") or []),
            "hc_report_rows": list(hc_report_rows),
            "category_candidates": list(result.get("category_candidates") or []),
            "ignored": ignored or bool(result.get("ignored")),
        }

    try:
        if enable_hc_filter:
            hc_files, normal_excel_files = split_hc_excel_files(excel_files)
        else:
            hc_files, normal_excel_files = [], excel_files

        if hc_files and hc_dir is not None:
            for hc_file in hc_files:
                hc_report_rows.append(copy_or_preview_hc_excel(virtual_archive, input_root, hc_file, hc_dir, dry_run))

        if hc_files and not normal_excel_files:
            reason = "仅发现 HC 文件，已排除"
            add_log(f"[{virtual_archive.name}] {reason}")
            archive_details.append(build_hc_only_archive_detail(virtual_archive, input_root, hc_files, reason))
            return build_result({"success": False, "rows": []}, ignored=True)

        result = process_excel_unit(
            virtual_archive,
            input_root,
            normal_excel_files,
            unit_folder,
            category_keywords,
            category_prefixes,
            skip_dir,
            unclassified_dir,
            dry_run,
            excel_group_mode=excel_group_mode,
            copy_skipped_excel_file=True,
        )
        return build_result(result)
    except Exception as exc:
        add_exception(f"异常：处理文件夹 Excel 失败 {virtual_archive.name}，原因：{exc}")
        add_structured_error(virtual_archive, "处理失败", f"处理文件夹 Excel 失败，原因：{exc}")
        return build_result({"success": False, "rows": []})
    finally:
        clear_thread_log_context()


def clean_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "_", name).strip()
    if not cleaned:
        cleaned = "未分类"
    return cleaned[:31]


def ensure_sheet_headers(ws) -> None:
    if ws.max_row == 1 and all(ws.cell(row=1, column=col).value is None for col in range(1, len(OUTPUT_HEADERS) + 1)):
        for col, header in enumerate(OUTPUT_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)
        return

    first_row = [ws.cell(row=1, column=col).value for col in range(1, len(OUTPUT_HEADERS) + 1)]
    if first_row != OUTPUT_HEADERS:
        ws.insert_rows(1)
        for col, header in enumerate(OUTPUT_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)


def style_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center

    ws.freeze_panes = "A2"

    for col_index in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_index)
        max_length = 0
        for cell in ws[column_letter]:
            value = cell.value
            if value is None:
                length = 0
            else:
                length = len(str(value))
            max_length = max(max_length, length)
        ws.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 60)


def apply_output_date_format(ws) -> None:
    date_col = OUTPUT_HEADERS.index("日期") + 1
    for row_index in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_index, column=date_col)
        if isinstance(cell.value, (date, datetime)):
            cell.number_format = OUTPUT_DATE_NUMBER_FORMAT


def apply_sheet_tab_color(ws, index: int) -> None:
    color = SHEET_TAB_COLORS.get(ws.title)
    if not color:
        color = FALLBACK_SHEET_TAB_COLORS[index % len(FALLBACK_SHEET_TAB_COLORS)]
    ws.sheet_properties.tabColor = color


def write_to_output(rows: list[dict[str, Any]], output_path: str | Path, clear: bool = False) -> None:
    """
    根据 category 写入不同 Sheet。
    每个 Sheet 表头：
    亚马逊订单号 | SKU | 数量 | 日期
    """
    output = Path(output_path).expanduser()
    output.parent.mkdir(parents=True, exist_ok=True)

    if output.exists() and not clear:
        workbook = load_workbook(output)
        add_log(f"输出文件已存在，追加写入：{output}")
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        if output.exists() and clear:
            add_log(f"已启用 --clear，清空旧输出文件并重新生成：{output}")
        else:
            add_log(f"输出文件不存在，自动创建：{output}")

    for row in rows:
        sheet_name = clean_sheet_name(str(row.get("category", "未分类")))
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook.create_sheet(sheet_name)
            for col, header in enumerate(OUTPUT_HEADERS, 1):
                ws.cell(row=1, column=col, value=header)

        ensure_sheet_headers(ws)
        ws.append([row.get(header, "") for header in OUTPUT_HEADERS])

    if not workbook.sheetnames:
        ws = workbook.create_sheet("未分类")
        for col, header in enumerate(OUTPUT_HEADERS, 1):
            ws.cell(row=1, column=col, value=header)

    for index, ws in enumerate(workbook.worksheets):
        ensure_sheet_headers(ws)
        style_sheet(ws)
        apply_output_date_format(ws)
        apply_sheet_tab_color(ws, index)

    workbook.save(output)
    workbook.close()
    add_log(f"已保存输出 Excel：{output}")


def make_empty_stats(dry_run: bool = False) -> dict[str, Any]:
    return {
        "total_archives": 0,
        "success_archives": 0,
        "skipped_archives": 0,
        "extracted_rows": 0,
        "written_rows": 0,
        "duplicate_order_count": 0,
        "skipped_exact_duplicate_count": 0,
        "hc_file_count": 0,
        "hc_copy_failed_count": 0,
        "hc_dir": "",
        "category_counts": {},
        "dry_run": dry_run,
        "workers": 4,
    }


def build_category_counts(rows: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        category = str(row.get("category") or "未分类")
        counts[category] = counts.get(category, 0) + 1
    return counts


def normalize_cell_for_compare(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return format_output_date_display(value)
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def row_order_id(row: dict[str, Any]) -> str:
    return normalize_cell_for_compare(row.get("亚马逊订单号"))


def row_order_sku_key(row: dict[str, Any]) -> tuple[str, str]:
    return row_order_id(row), normalize_cell_for_compare(row.get("SKU"))


def exact_duplicate_key(row: dict[str, Any]) -> tuple[str, str, str, str, str]:
    return (
        normalize_cell_for_compare(row.get("category")),
        normalize_cell_for_compare(row.get("亚马逊订单号")),
        normalize_cell_for_compare(row.get("SKU")),
        normalize_cell_for_compare(row.get("数量")),
        normalize_cell_for_compare(row.get("日期")),
    )


def read_existing_output_rows(output_path: str | Path) -> list[dict[str, Any]]:
    output = Path(output_path).expanduser()
    if not output.exists():
        return []

    try:
        workbook = load_workbook(output, data_only=True, read_only=True)
    except Exception as exc:
        add_exception(f"异常：读取已有输出 Excel 失败，重复检测将只检查本次导入，原因：{exc}")
        return []

    rows: list[dict[str, Any]] = []
    try:
        for ws in workbook.worksheets:
            headers = [ws.cell(row=1, column=col).value for col in range(1, len(OUTPUT_HEADERS) + 1)]
            if headers != OUTPUT_HEADERS:
                continue
            for row_index in range(2, (ws.max_row or 0) + 1):
                values = [ws.cell(row=row_index, column=col).value for col in range(1, len(OUTPUT_HEADERS) + 1)]
                if not any(has_value(value) for value in values):
                    continue
                rows.append(
                    {
                        "category": ws.title,
                        "压缩包名": "",
                        "Excel 文件名": output.name,
                        "Sheet 名": ws.title,
                        "源数据行号": row_index,
                        "亚马逊订单号": normalize_cell_value(values[0]),
                        "SKU": normalize_cell_value(values[1]),
                        "数量": normalize_cell_value(values[2]),
                        "日期": normalize_cell_value(values[3]),
                    }
                )
    finally:
        workbook.close()
    return rows


def source_label(has_existing: bool, has_current: bool) -> str:
    if has_existing and has_current:
        return "已有输出、本次导入"
    if has_existing:
        return "已有输出"
    return "本次导入"


def detect_duplicates(
    rows: list[dict[str, Any]],
    output_path: str | Path,
    clear: bool,
    detect_duplicate_orders: bool,
    skip_exact_duplicates: bool,
    dry_run: bool,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    existing_rows = [] if clear else read_existing_output_rows(output_path)
    duplicate_report_rows: list[dict[str, Any]] = []
    exact_duplicate_report_rows: list[dict[str, Any]] = []

    existing_order_sku_counts: dict[tuple[str, str], int] = {}
    current_order_sku_counts: dict[tuple[str, str], int] = {}
    existing_exact_counts: dict[tuple[str, str, str, str, str], int] = {}
    seen_current_exact_counts: dict[tuple[str, str, str, str, str], int] = {}

    for row in existing_rows:
        order_id, sku = row_order_sku_key(row)
        if order_id and sku:
            key = (order_id, sku)
            existing_order_sku_counts[key] = existing_order_sku_counts.get(key, 0) + 1
        key = exact_duplicate_key(row)
        existing_exact_counts[key] = existing_exact_counts.get(key, 0) + 1

    for row in rows:
        order_id, sku = row_order_sku_key(row)
        if order_id and sku:
            key = (order_id, sku)
            current_order_sku_counts[key] = current_order_sku_counts.get(key, 0) + 1

    if detect_duplicate_orders:
        for row in rows:
            order_id, sku = row_order_sku_key(row)
            if not order_id or not sku:
                continue
            order_sku_key = (order_id, sku)
            existing_count = existing_order_sku_counts.get(order_sku_key, 0)
            current_count = current_order_sku_counts.get(order_sku_key, 0)
            if existing_count + current_count <= 1:
                continue
            duplicate_report_rows.append(
                {
                    "类型": "重复订单+SKU",
                    "压缩包名": row.get("压缩包名", ""),
                    "Excel 文件名": row.get("Excel 文件名", ""),
                    "Sheet 名": row.get("Sheet 名", ""),
                    "源数据行号": row.get("源数据行号", ""),
                    "亚马逊订单号": row.get("亚马逊订单号", ""),
                    "SKU": row.get("SKU", ""),
                    "数量": row.get("数量", ""),
                    "日期": format_output_date_display(row.get("日期", "")),
                    "品类": row.get("category", "未分类"),
                    "重复来源": source_label(existing_count > 0, current_count > 1),
                    "处理方式": "仅提示",
                }
            )

    rows_to_write: list[dict[str, Any]] = []
    for row in rows:
        key = exact_duplicate_key(row)
        existing_count = existing_exact_counts.get(key, 0)
        current_seen = seen_current_exact_counts.get(key, 0)
        is_duplicate = existing_count > 0 or current_seen > 0
        if is_duplicate:
            duplicate_source = source_label(existing_count > 0, current_seen > 0)
            if skip_exact_duplicates:
                handling = "已跳过"
            elif dry_run:
                handling = "仅提示"
            else:
                handling = "已写入"
            exact_duplicate_report_rows.append(
                {
                    "类型": "完全重复行",
                    "压缩包名": row.get("压缩包名", ""),
                    "Excel 文件名": row.get("Excel 文件名", ""),
                    "Sheet 名": row.get("Sheet 名", ""),
                    "源数据行号": row.get("源数据行号", ""),
                    "品类": row.get("category", "未分类"),
                    "亚马逊订单号": row.get("亚马逊订单号", ""),
                    "SKU": row.get("SKU", ""),
                    "数量": row.get("数量", ""),
                    "日期": format_output_date_display(row.get("日期", "")),
                    "重复来源": duplicate_source,
                    "处理方式": handling,
                }
            )
            if skip_exact_duplicates:
                seen_current_exact_counts[key] = current_seen + 1
                continue

        rows_to_write.append(row)
        seen_current_exact_counts[key] = current_seen + 1

    return rows_to_write, duplicate_report_rows, exact_duplicate_report_rows


def format_stats_logs(stats: dict[str, Any]) -> str:
    lines = [
        "========== 处理统计 ==========",
    ]
    if stats.get("dry_run"):
        lines.append("当前模式：仅预览，不写入 Excel")
    lines.extend(
        [
            f"总压缩包：{stats.get('total_archives', 0)}",
            f"成功处理：{stats.get('success_archives', 0)}",
            f"异常跳过：{stats.get('skipped_archives', 0)}",
            f"总提取行数：{stats.get('extracted_rows', 0)}",
            f"HC 文件：{stats.get('hc_file_count', 0)}",
            f"HC 复制失败：{stats.get('hc_copy_failed_count', 0)}",
        ]
    )
    if stats.get("dry_run"):
        estimated_written = stats.get("extracted_rows", 0) - stats.get("skipped_exact_duplicate_count", 0)
        lines.append(f"预计可写入行数：{estimated_written}")
    lines.append(f"实际写入行数：{stats.get('written_rows', 0)}")
    lines.append(f"重复订单号：{stats.get('duplicate_order_count', 0)}")
    if stats.get("dry_run"):
        lines.append(f"预计跳过完全重复行：{stats.get('skipped_exact_duplicate_count', 0)}")
    else:
        lines.append(f"跳过完全重复行：{stats.get('skipped_exact_duplicate_count', 0)}")
    lines.append(f"并发数量：{stats.get('workers', 1)}")
    lines.append("")
    lines.append("品类统计：")
    category_counts = stats.get("category_counts") or {}
    if category_counts:
        for category, count in sorted(category_counts.items()):
            lines.append(f"{category}：{count} 行")
    else:
        lines.append("无")
    return "\n".join(lines)


def format_duplicate_order_logs(duplicate_report_rows: list[dict[str, Any]] | None = None) -> str:
    rows = duplicate_report_rows or []
    lines = ["========== 重复订单号提示 =========="]
    if not rows:
        lines.append("无重复订单号")
        return "\n".join(lines)
    lines.append(f"发现重复订单号：{len(rows)} 条")
    for index, row in enumerate(rows[:100], 1):
        lines.append(
            f"{index}. {row.get('亚马逊订单号', '')} | SKU: {row.get('SKU', '')} | "
            f"品类：{row.get('品类', '')} | 来源：{row.get('重复来源', '')}"
        )
    if len(rows) > 100:
        lines.append(f"... 其余 {len(rows) - 100} 条已省略")
    return "\n".join(lines)


def format_exact_duplicate_logs(exact_duplicate_report_rows: list[dict[str, Any]] | None = None) -> str:
    rows = exact_duplicate_report_rows or []
    lines = ["========== 完全重复行提示 =========="]
    if not rows:
        lines.append("无完全重复行")
        return "\n".join(lines)
    lines.append(f"发现完全重复行：{len(rows)} 条")
    for index, row in enumerate(rows[:100], 1):
        lines.append(
            f"{index}. {row.get('品类', '')} | {row.get('亚马逊订单号', '')} | "
            f"{row.get('SKU', '')} | {row.get('数量', '')} | {format_output_date_display(row.get('日期', ''))} | "
            f"{row.get('重复来源', '')} | {row.get('处理方式', '')}"
        )
    if len(rows) > 100:
        lines.append(f"... 其余 {len(rows) - 100} 条已省略")
    return "\n".join(lines)


def format_logs_for_output(
    stats: dict[str, Any] | None = None,
    log_file_path: str | None = None,
    duplicate_report_rows: list[dict[str, Any]] | None = None,
    exact_duplicate_report_rows: list[dict[str, Any]] | None = None,
    duplicate_report_file_path: str | None = None,
    process_report_file_path: str | None = None,
    backup_file_path: str | None = None,
) -> str:
    sections = ["========== 异常情况汇总 =========="]
    if exception_logs:
        sections.append("\n\n".join(exception_logs))
    else:
        sections.append("无异常")

    sections.append("")
    sections.append("========== 正常处理日志 ==========")
    if process_logs:
        sections.append("\n".join(process_logs))
    else:
        sections.append("无正常处理日志")

    if stats is not None:
        sections.append("")
        sections.append(format_stats_logs(stats))

    sections.append("")
    sections.append(format_duplicate_order_logs(duplicate_report_rows))
    sections.append("")
    sections.append(format_exact_duplicate_logs(exact_duplicate_report_rows))

    if duplicate_report_file_path:
        sections.append("")
        sections.append(f"重复报告已保存：{duplicate_report_file_path}")

    if process_report_file_path:
        sections.append("")
        sections.append(f"处理报告已保存：{process_report_file_path}")

    if backup_file_path:
        sections.append("")
        sections.append(f"旧汇总备份：{backup_file_path}")

    if log_file_path:
        sections.append("")
        sections.append(f"日志已保存：{log_file_path}")

    return "\n".join(sections)


def print_logs(
    stats: dict[str, Any] | None = None,
    log_file_path: str | None = None,
    duplicate_report_rows: list[dict[str, Any]] | None = None,
    exact_duplicate_report_rows: list[dict[str, Any]] | None = None,
    duplicate_report_file_path: str | None = None,
    process_report_file_path: str | None = None,
    backup_file_path: str | None = None,
) -> None:
    print(
        format_logs_for_output(
            stats,
            log_file_path,
            duplicate_report_rows,
            exact_duplicate_report_rows,
            duplicate_report_file_path,
            process_report_file_path,
            backup_file_path,
        )
    )


def get_runtime_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def save_run_log(
    stats: dict[str, Any],
    duplicate_report_rows: list[dict[str, Any]] | None = None,
    exact_duplicate_report_rows: list[dict[str, Any]] | None = None,
    duplicate_report_file_path: str | None = None,
    process_report_file_path: str | None = None,
    backup_file_path: str | None = None,
    log_dir: str | Path | None = None,
) -> str:
    logs_dir = Path(log_dir).expanduser() if log_dir else get_runtime_base_dir() / "logs"
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    log_file_path = logs_dir / f"{timestamp}_处理日志.txt"
    try:
        logs_dir.mkdir(parents=True, exist_ok=True)
        if log_file_path.exists():
            index = 1
            while True:
                candidate = logs_dir / f"{timestamp}_处理日志_{index}.txt"
                if not candidate.exists():
                    log_file_path = candidate
                    break
                index += 1
        log_file_path.write_text(
            format_logs_for_output(
                stats,
                duplicate_report_rows=duplicate_report_rows,
                exact_duplicate_report_rows=exact_duplicate_report_rows,
                duplicate_report_file_path=duplicate_report_file_path,
                process_report_file_path=process_report_file_path,
                backup_file_path=backup_file_path,
            ),
            encoding="utf-8-sig",
        )
        return str(log_file_path.resolve())
    except Exception as exc:
        add_exception(f"异常：保存日志失败，原因：{exc}")
        return ""


def save_debug_report(debug_rows: list[str], log_dir: str | Path | None = None) -> str:
    logs_dir = Path(log_dir).expanduser() if log_dir else get_runtime_base_dir() / "logs"
    debug_report_path = logs_dir / "debug_report.txt"
    try:
        logs_dir.mkdir(parents=True, exist_ok=True)
        content = "\n\n".join(debug_rows) if debug_rows else "无 DEBUG 日志"
        debug_report_path.write_text(content + "\n", encoding="utf-8-sig")
        add_log(f"DEBUG 报告已保存：{debug_report_path}")
        return str(debug_report_path.resolve())
    except Exception as exc:
        add_exception(f"异常：保存 DEBUG 报告失败，原因：{exc}")
        return ""


def normalize_input_mode(value: str | None) -> str:
    mode = (value or INPUT_MODE_ARCHIVES).strip().lower()
    if mode not in INPUT_MODES:
        raise ValueError(f"input_mode 必须是 archives、folders 或 mixed，当前值：{value}")
    return mode


def normalize_excel_group_mode(value: str | None) -> str:
    mode = (value or EXCEL_GROUP_SINGLE).strip().lower()
    if mode not in EXCEL_GROUP_MODES:
        raise ValueError(f"excel_group_mode 必须是 single 或 multi，当前值：{value}")
    return mode


def clamp_workers(workers: Any) -> int:
    try:
        value = int(workers)
    except (TypeError, ValueError):
        value = 4
    return max(1, min(value, 8))


def unique_report_path(folder: Path, stem: str, suffix: str = ".xlsx") -> Path:
    folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = folder / f"{stem}_{timestamp}{suffix}"
    if not candidate.exists():
        return candidate
    index = 1
    while True:
        numbered = folder / f"{stem}_{timestamp}_{index}{suffix}"
        if not numbered.exists():
            return numbered
        index += 1


def style_report_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center
    ws.freeze_panes = "A2"
    for col_index in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_index)
        max_length = 0
        for cell in ws[column_letter]:
            max_length = max(max_length, len(str(cell.value or "")))
        ws.column_dimensions[column_letter].width = min(max(max_length + 4, 12), 80)


def export_duplicate_report_to_excel(rows: list[dict[str, Any]], output_path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "重复报告"
    for col, header in enumerate(DUPLICATE_REPORT_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)
    for row_index, row in enumerate(rows, 2):
        for col, header in enumerate(DUPLICATE_REPORT_HEADERS, 1):
            ws.cell(row=row_index, column=col, value=row.get(header, ""))
    style_report_sheet(ws)
    workbook.save(output_path)
    workbook.close()


def build_duplicate_report_export_rows(
    duplicate_report_rows: list[dict[str, Any]],
    exact_duplicate_report_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    export_rows: list[dict[str, Any]] = []
    for row in duplicate_report_rows:
        export_rows.append({header: row.get(header, "") for header in DUPLICATE_REPORT_HEADERS})
    for row in exact_duplicate_report_rows:
        export_rows.append({header: row.get(header, "") for header in DUPLICATE_REPORT_HEADERS})
    return export_rows


def save_duplicate_report(
    output_path: str | Path,
    duplicate_report_rows: list[dict[str, Any]],
    exact_duplicate_report_rows: list[dict[str, Any]],
) -> str:
    export_rows = build_duplicate_report_export_rows(duplicate_report_rows, exact_duplicate_report_rows)
    if not export_rows:
        return ""
    output = Path(output_path).expanduser()
    report_path = unique_report_path(output.parent, "重复报告")
    try:
        export_duplicate_report_to_excel(export_rows, report_path)
        add_log(f"重复报告已保存：{report_path}")
        return str(report_path.resolve())
    except Exception as exc:
        add_exception(f"异常：保存重复报告失败，原因：{exc}")
        return ""


def backup_existing_output(output_path: str | Path, backup_dir: str | Path | None = None) -> str:
    output = Path(output_path).expanduser()
    if not output.exists():
        add_log("输出文件不存在，本次无需备份旧汇总 Excel")
        return ""

    backup_folder = Path(backup_dir).expanduser() if backup_dir else get_runtime_base_dir() / "backups"
    backup_folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{output.stem}_backup_{timestamp}{output.suffix or '.xlsx'}"
    backup_path = backup_folder / backup_name
    if backup_path.exists():
        index = 1
        while True:
            candidate = backup_folder / f"{output.stem}_backup_{timestamp}_{index}{output.suffix or '.xlsx'}"
            if not candidate.exists():
                backup_path = candidate
                break
            index += 1
    shutil.copy2(output, backup_path)
    add_log(f"已备份旧汇总 Excel：{backup_path}")
    return str(backup_path.resolve())


def unique_copy_path(folder: Path, source_name: str, create_folder: bool = True) -> Path:
    if create_folder:
        folder.mkdir(parents=True, exist_ok=True)
    candidate = folder / source_name
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    index = 1
    while True:
        numbered = folder / f"{stem}_{index}{suffix}"
        if not numbered.exists():
            return numbered
        index += 1


def copy_skipped_source(source_path: Path, skip_dir: Path | None, dry_run: bool = False) -> str:
    if skip_dir is None:
        return ""
    if dry_run:
        target_path = unique_copy_path(skip_dir, source_path.name, create_folder=False)
        add_log(f"dry-run 模式：将跳过文件复制到：{target_path}")
        return str(target_path.resolve())
    target_path = unique_copy_path(skip_dir, source_path.name)
    shutil.copy2(source_path, target_path)
    add_log(f"已复制跳过文件到：{target_path}")
    return str(target_path.resolve())


def copy_unclassified_excel(source_path: Path, unclassified_dir: Path | None, dry_run: bool = False) -> str:
    if unclassified_dir is None:
        return ""
    if dry_run:
        target_path = unique_copy_path(unclassified_dir, source_path.name, create_folder=False)
        add_log(f"dry-run 模式：未分类 Excel 将复制到：{target_path}；来源：{source_path.name}")
        return str(target_path.resolve())
    target_path = unique_copy_path(unclassified_dir, source_path.name)
    shutil.copy2(source_path, target_path)
    add_log(f"未分类 Excel 已复制到：{target_path}；来源：{source_path.name}")
    return str(target_path.resolve())


def build_skip_dir(input_path: str | Path) -> Path:
    path = Path(input_path).expanduser()
    base = path if path.is_dir() else path.parent
    return base / "未处理压缩包"


def build_unclassified_dir(input_path: str | Path) -> Path:
    path = Path(input_path).expanduser()
    base = path if path.is_dir() else path.parent
    return base / UNCLASSIFIED_FOLDER_NAME


def build_hc_dir(input_path: str | Path) -> Path:
    path = Path(input_path).expanduser()
    base = path if path.is_dir() else path.parent
    return base / HC_FOLDER_NAME


def consolidate_order_sku_rows(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    grouped_rows: list[dict[str, Any]] = []
    group_index: dict[tuple[str, str, str, str], int] = {}
    quantity_error_rows: list[dict[str, Any]] = []

    for row in rows:
        order_id = row_order_id(row)
        sku = normalize_cell_for_compare(row.get("SKU"))
        key = (
            normalize_cell_for_compare(row.get("category")),
            order_id,
            sku,
            normalize_cell_for_compare(row.get("日期")),
        )
        quantity_number = parse_quantity_integer(row.get("数量"))
        if not order_id:
            grouped_rows.append(row)
            if has_value(row.get("数量")) and quantity_number is None:
                quantity_error_rows.append({**row, "处理方式": "数量无法转换，未参与合并"})
            continue
        if quantity_number is None:
            grouped_rows.append(row)
            if has_value(row.get("数量")):
                quantity_error_rows.append({**row, "处理方式": "数量无法转换，未参与合并"})
            continue

        if key not in group_index:
            new_row = dict(row)
            new_row["数量"] = quantity_number
            group_index[key] = len(grouped_rows)
            grouped_rows.append(new_row)
            continue

        target_row = grouped_rows[group_index[key]]
        target_quantity = parse_quantity_integer(target_row.get("数量"))
        if target_quantity is None:
            grouped_rows.append(row)
            quantity_error_rows.append({**target_row, "处理方式": "数量无法转换，未参与合并"})
            continue
        target_row["数量"] = target_quantity + quantity_number
        source_rows = str(target_row.get("源数据行号", ""))
        current_row = str(row.get("源数据行号", ""))
        if current_row and current_row not in source_rows.split("、"):
            target_row["源数据行号"] = f"{source_rows}、{current_row}" if source_rows else current_row

    return grouped_rows, quantity_error_rows


def dedupe_quantity_error_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str, str]] = set()
    for row in rows:
        key = (
            normalize_cell_for_compare(row.get("压缩包名")),
            normalize_cell_for_compare(row.get("Excel 文件名")),
            normalize_cell_for_compare(row.get("Sheet 名")),
            normalize_cell_for_compare(row.get("源数据行号")),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    return deduped


def append_dict_rows(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for row_index, row in enumerate(rows, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row_index, column=col, value=row.get(header, ""))


def build_overview_rows(report_data: dict[str, Any]) -> list[dict[str, Any]]:
    stats = report_data.get("stats") or {}
    return [
        {"项目": "运行开始时间", "内容": report_data.get("start_time", "")},
        {"项目": "运行结束时间", "内容": report_data.get("end_time", "")},
        {"项目": "输入路径", "内容": report_data.get("input_path", "")},
        {"项目": "输出路径", "内容": report_data.get("output_path", "")},
        {"项目": "是否 clear", "内容": "是" if report_data.get("clear") else "否"},
        {"项目": "dry-run 模式", "内容": "是" if report_data.get("dry_run") else "否"},
        {"项目": "是否写入汇总 Excel", "内容": "是" if report_data.get("wrote_output") else "否"},
        {"项目": "是否备份旧汇总 Excel", "内容": "是" if report_data.get("backup_path") else "否"},
        {"项目": "本次无可写入数据", "内容": "是" if report_data.get("no_writable_data") else "否"},
        {"项目": "workers 数量", "内容": stats.get("workers", "")},
        {"项目": "处理压缩包数量", "内容": stats.get("total_archives", 0)},
        {"项目": "成功处理压缩包数量", "内容": stats.get("success_archives", 0)},
        {"项目": "跳过压缩包数量", "内容": stats.get("skipped_archives", 0)},
        {"项目": "HC 文件数量", "内容": stats.get("hc_file_count", 0)},
        {"项目": "HC 复制失败数量", "内容": stats.get("hc_copy_failed_count", 0)},
        {"项目": "HC 文件夹路径", "内容": stats.get("hc_dir", "")},
        {"项目": "异常数量", "内容": len(report_data.get("error_report_rows") or []) + len(report_data.get("quantity_error_rows") or [])},
        {"项目": "表头缺失警告数量", "内容": len(report_data.get("header_report_rows") or [])},
        {"项目": "重复订单数量", "内容": len(report_data.get("duplicate_report_rows") or [])},
        {"项目": "完全重复行数量", "内容": len(report_data.get("exact_duplicate_report_rows") or [])},
        {"项目": "最终写入行数", "内容": stats.get("written_rows", 0)},
        {"项目": "异常报告路径", "内容": report_data.get("error_report_path", "")},
        {"项目": "重复报告路径", "内容": report_data.get("duplicate_report_path", "")},
        {"项目": "处理报告路径", "内容": report_data.get("process_report_path", "")},
        {"项目": "旧汇总备份路径", "内容": report_data.get("backup_path", "")},
    ]


def build_exception_detail_rows(
    error_rows: list[dict[str, Any]],
    quantity_error_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in error_rows:
        rows.append(
            {
                "异常类型": item.get("异常类型", ""),
                "压缩包名": item.get("压缩包名称", ""),
                "Excel文件名": item.get("相关文件", ""),
                "Sheet名": "",
                "行号": "",
                "异常内容": item.get("异常原因", ""),
                "处理方式": item.get("处理状态", ""),
            }
        )
    for item in quantity_error_rows:
        rows.append(
            {
                "异常类型": "数量无法计算",
                "压缩包名": item.get("压缩包名", ""),
                "Excel文件名": item.get("Excel 文件名", ""),
                "Sheet名": item.get("Sheet 名", ""),
                "行号": item.get("源数据行号", ""),
                "原始数量值": item.get("原始数量", item.get("数量", "")),
                "异常内容": f"数量无法转换：{item.get('原始数量', item.get('数量', ''))}",
                "处理方式": item.get("处理方式", "未参与数量求和，原数据仍按规则提取"),
            }
        )
    return rows


def build_duplicate_detail_rows(
    duplicate_report_rows: list[dict[str, Any]],
    exact_duplicate_report_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in duplicate_report_rows + exact_duplicate_report_rows:
        rows.append(
            {
                "重复类型": item.get("类型", ""),
                "压缩包名": item.get("压缩包名", ""),
                "Excel文件名": item.get("Excel 文件名", ""),
                "Sheet名": item.get("Sheet 名", ""),
                "行号": item.get("源数据行号", ""),
                "亚马逊订单号": item.get("亚马逊订单号", ""),
                "SKU": item.get("SKU", ""),
                "数量": item.get("数量", ""),
                "日期": item.get("日期", ""),
                "处理方式": item.get("处理方式", ""),
            }
        )
    return rows


def build_category_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = {}
    for row in rows:
        category = str(row.get("category") or "未分类")
        item = summary.setdefault(
            category,
            {"品类": category, "写入行数": 0, "订单数": 0, "数量合计": 0.0, "涉及压缩包数量": 0, "_orders": set(), "_archives": set()},
        )
        item["写入行数"] += 1
        order_id = row_order_id(row)
        if order_id:
            item["_orders"].add(order_id)
        archive_name = str(row.get("压缩包名") or "")
        if archive_name:
            item["_archives"].add(archive_name)
        number = quantity_to_number(row.get("数量"))
        if number is not None:
            item["数量合计"] += number

    output_rows = []
    for item in summary.values():
        quantity_total = format_quantity_total(float(item["数量合计"]))
        output_rows.append(
            {
                "品类": item["品类"],
                "写入行数": item["写入行数"],
                "订单数": len(item["_orders"]),
                "数量合计": quantity_total,
                "涉及压缩包数量": len(item["_archives"]),
            }
        )
    return output_rows


def build_daily_order_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = {}
    for row in rows:
        raw_date = row.get("日期")
        date_text = format_output_date_display(raw_date) if raw_date else "未识别日期"
        item = summary.setdefault(
            date_text,
            {"日期": date_text, "明细行数": 0, "订单号": set(), "数量合计": 0.0, "品类": set()},
        )
        item["明细行数"] += 1
        order_id = row_order_id(row)
        if order_id:
            item["订单号"].add(order_id)
        category = str(row.get("category") or "")
        if category:
            item["品类"].add(category)
        number = quantity_to_number(row.get("数量"))
        if number is not None:
            item["数量合计"] += number

    output_rows = []
    for item in summary.values():
        output_rows.append(
            {
                "日期": item["日期"],
                "订单数": len(item["订单号"]),
                "数量合计": format_quantity_total(float(item["数量合计"])),
                "明细行数": item["明细行数"],
                "涉及品类": "、".join(sorted(item["品类"])),
            }
        )
    return output_rows


def highlight_report_rows(ws, headers: list[str], yellow_columns: set[str], red_columns: set[str]) -> None:
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
    header_index = {header: index + 1 for index, header in enumerate(headers)}
    for row_index in range(2, ws.max_row + 1):
        row_fill = None
        for column in red_columns:
            col_index = header_index.get(column)
            if col_index and ws.cell(row=row_index, column=col_index).value not in ("", "匹配", "未识别", None):
                row_fill = red_fill
                break
        if row_fill is None:
            for column in yellow_columns:
                col_index = header_index.get(column)
                if col_index and ws.cell(row=row_index, column=col_index).value not in ("", "匹配", "未识别", None):
                    row_fill = yellow_fill
                    break
        if row_fill:
            for col_index in range(1, ws.max_column + 1):
                ws.cell(row=row_index, column=col_index).fill = row_fill


def save_process_report(report_data: dict[str, Any], report_dir: str | Path | None = None) -> str:
    folder = Path(report_dir).expanduser() if report_dir else get_runtime_base_dir() / "logs"
    report_path = unique_report_path(folder, "处理报告")
    report_data["process_report_path"] = str(report_path.resolve())
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    sheets = [
        ("运行概览", ["项目", "内容"], build_overview_rows(report_data)),
        (
            "压缩包明细",
            [
                "序号",
                "外层压缩包名",
                "子文件夹名",
                "压缩包名",
                "压缩包路径",
                "处理状态",
                "正式Excel数量",
                "正式Excel文件",
                "提取行数",
                "识别品类",
                "识别日期",
                "是否跳过",
                "异常原因",
                "文件名预计单量",
                "实际提取单量",
                "单量校验结果",
                "文件名预计数量",
                "实际提取数量",
                "数量校验结果",
                "已复制到",
            ],
            [
                {"序号": index, **row}
                for index, row in enumerate(report_data.get("archive_details") or [], 1)
            ],
        ),
        (
            "文件名校验",
            [
                "序号",
                "外层压缩包名",
                "子文件夹名",
                "压缩包名",
                "Excel文件名",
                "品类",
                "日期",
                "文件名预计单量",
                "实际提取单量",
                "单量校验结果",
                "文件名预计数量",
                "实际提取数量",
                "数量校验结果",
                "提取行数",
                "备注",
            ],
            [
                {"序号": index, **row}
                for index, row in enumerate(report_data.get("filename_validations") or [], 1)
            ],
        ),
        (
            "HC文件明细",
            ["序号", "外层压缩包名", "子文件夹名", "Excel文件名", "压缩包内路径", "处理状态", "目标路径", "失败原因"],
            [
                {"序号": index, **row}
                for index, row in enumerate(report_data.get("hc_report_rows") or [], 1)
            ],
        ),
        (
            "异常明细",
            ["序号", "异常类型", "压缩包名", "Excel文件名", "Sheet名", "行号", "原始数量值", "异常内容", "处理方式"],
            [
                {"序号": index, **row}
                for index, row in enumerate(
                    build_exception_detail_rows(
                        report_data.get("error_report_rows") or [],
                        report_data.get("quantity_error_rows") or [],
                    ),
                    1,
                )
            ],
        ),
        (
            "表头识别",
            ["序号", "压缩包名", "子文件夹名", "Excel文件名", "Sheet名", "已识别表头", "缺失表头", "处理方式"],
            [
                {"序号": index, **row}
                for index, row in enumerate(report_data.get("header_report_rows") or [], 1)
            ],
        ),
        (
            "重复明细",
            ["序号", "重复类型", "压缩包名", "Excel文件名", "Sheet名", "行号", "亚马逊订单号", "SKU", "数量", "日期", "处理方式"],
            [
                {"序号": index, **row}
                for index, row in enumerate(
                    build_duplicate_detail_rows(
                        report_data.get("duplicate_report_rows") or [],
                        report_data.get("exact_duplicate_report_rows") or [],
                    ),
                    1,
                )
            ],
        ),
        (
            "新品类候选",
            ["序号", "原始名称", "原始候选", "识别前缀", "候选品类", "状态", "Excel文件名", "来源路径"],
            [
                {
                    "序号": index,
                    "原始名称": row.get("source_name", ""),
                    "原始候选": row.get("raw_candidate", ""),
                    "识别前缀": row.get("prefix", ""),
                    "候选品类": row.get("category", ""),
                    "状态": row.get("status", "待确认"),
                    "Excel文件名": row.get("excel_file", ""),
                    "来源路径": row.get("source_path", ""),
                }
                for index, row in enumerate(report_data.get("category_candidates") or [], 1)
            ],
        ),
        (
            "品类汇总",
            ["品类", "写入行数", "订单数", "数量合计", "涉及压缩包数量"],
            build_category_summary_rows(report_data.get("rows_to_write") or []),
        ),
        (
            "每日单量汇总",
            ["日期", "订单数", "数量合计", "明细行数", "涉及品类"],
            build_daily_order_summary_rows(report_data.get("rows_to_write") or []),
        ),
    ]

    for sheet_name, headers, rows in sheets:
        ws = workbook.create_sheet(sheet_name)
        append_dict_rows(ws, headers, rows)
        style_report_sheet(ws)
        if sheet_name == "文件名校验":
            highlight_report_rows(ws, headers, {"单量校验结果", "数量校验结果"}, {"数量校验结果"})
        elif sheet_name == "异常明细":
            red_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")
            for row_index in range(2, ws.max_row + 1):
                for col_index in range(1, ws.max_column + 1):
                    ws.cell(row=row_index, column=col_index).fill = red_fill
        elif sheet_name == "表头识别":
            yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
            for row_index in range(2, ws.max_row + 1):
                for col_index in range(1, ws.max_column + 1):
                    ws.cell(row=row_index, column=col_index).fill = yellow_fill

    workbook.save(report_path)
    workbook.close()
    add_log(f"处理报告已保存：{report_path}")
    return str(report_path.resolve())


def run_extract(
    input_path: str,
    output_path: str,
    clear: bool = False,
    dry_run: bool = False,
    workers: int = 4,
    detect_duplicate_orders: bool = True,
    skip_exact_duplicates: bool = True,
    category_config_path: str | Path | None = None,
    report_dir: str | Path | None = None,
    backup_dir: str | Path | None = None,
    log_callback: Callable[[str], None] | None = None,
    progress_callback: Callable[[dict[str, Any]], None] | None = None,
    input_mode: str = INPUT_MODE_ARCHIVES,
    enable_hc_filter: bool = False,
    excel_group_mode: str = EXCEL_GROUP_SINGLE,
) -> dict[str, Any]:
    """
    给 GUI 和 CLI 共用的主入口。

    返回：
    {
        "rows": [...],
        "stats": {...},
        "exception_logs": [...],
        "process_logs": [...],
        "log_file_path": "C:\\xxx\\logs\\2026-05-09_153000_处理日志.txt"
    }
    """
    reset_logs()
    set_log_callback(log_callback)
    start_time = datetime.now()
    output = Path(output_path).expanduser()
    normalized_workers = clamp_workers(workers)
    normalized_input_mode = normalize_input_mode(input_mode)
    normalized_excel_group_mode = normalize_excel_group_mode(excel_group_mode)
    logs_dir = Path(report_dir).expanduser() if report_dir else get_runtime_base_dir() / "logs"
    skip_dir = build_skip_dir(input_path)
    unclassified_dir = build_unclassified_dir(input_path)
    hc_dir = build_hc_dir(input_path)
    all_rows: list[dict[str, Any]] = []
    merged_rows: list[dict[str, Any]] = []
    rows_to_write: list[dict[str, Any]] = []
    duplicate_report_rows: list[dict[str, Any]] = []
    exact_duplicate_report_rows: list[dict[str, Any]] = []
    archive_details: list[dict[str, Any]] = []
    filename_validations: list[dict[str, Any]] = []
    quantity_error_rows: list[dict[str, Any]] = []
    merge_quantity_error_rows: list[dict[str, Any]] = []
    header_report_rows: list[dict[str, Any]] = []
    hc_report_rows: list[dict[str, Any]] = []
    category_candidates: list[dict[str, str]] = []
    archive_results_by_index: dict[int, dict[str, Any]] = {}
    stats = make_empty_stats(dry_run=dry_run)
    stats["workers"] = normalized_workers
    stats["hc_dir"] = str(hc_dir.resolve())
    log_file_path = ""
    duplicate_report_file_path = ""
    process_report_file_path = ""
    debug_report_file_path = ""
    backup_path = ""
    wrote_output = False
    no_writable_data = False
    success = False

    category_config_data, category_config_file_path, category_config_error = load_category_config_data(category_config_path)
    category_keywords = category_config_data.categories
    category_prefixes = category_config_data.prefixes
    if category_config_error:
        add_log(f"警告：{category_config_error}")
    else:
        add_log(f"已加载品类关键词配置：{category_config_file_path}")

    def emit_progress(payload: dict[str, Any]) -> None:
        if progress_callback is None:
            return
        try:
            progress_callback(payload)
        except Exception:
            pass

    try:
        input_item = Path(input_path).expanduser()
        input_root = input_item if input_item.is_dir() else input_item.parent
        archives: list[Path] = []
        folder_groups: list[tuple[Path, list[Path]]] = []
        should_scan_archives = normalized_input_mode == INPUT_MODE_ARCHIVES or (
            normalized_input_mode == INPUT_MODE_MIXED
            and (input_item.is_dir() or input_item.suffix.lower() in ARCHIVE_EXTENSIONS)
        )
        if should_scan_archives:
            archives = find_archive_files(input_path)
        if normalized_input_mode in {INPUT_MODE_FOLDERS, INPUT_MODE_MIXED}:
            folder_excel_files = find_folder_excel_files(input_path)
            folder_groups = build_folder_processing_groups(input_root, folder_excel_files)

        total_tasks = len(archives) + len(folder_groups)
        stats["total_archives"] = total_tasks

        if total_tasks == 0:
            add_log("没有找到可处理的压缩包")
            add_log("处理完成，总计提取：0 行")
            success = not exception_logs
        else:
            with ThreadPoolExecutor(max_workers=normalized_workers) as executor:
                future_map = {}
                task_index = 0
                for archive_path in archives:
                    task_index += 1
                    emit_progress(
                        {
                            "current": 0,
                            "total": total_tasks,
                            "archive_name": archive_path.name,
                            "status": "queued",
                            "active_workers": min(normalized_workers, total_tasks),
                            "completed_archives": 0,
                            "failed_archives": 0,
                        }
                    )
                    future = executor.submit(
                        process_archive,
                        archive_path,
                        category_keywords,
                        category_prefixes,
                        skip_dir,
                        unclassified_dir,
                        hc_dir,
                        dry_run,
                        enable_hc_filter,
                        normalized_excel_group_mode,
                    )
                    future_map[future] = (task_index, archive_path)
                for unit_folder, unit_files in folder_groups:
                    task_index += 1
                    emit_progress(
                        {
                            "current": 0,
                            "total": total_tasks,
                            "archive_name": unit_folder.name,
                            "status": "queued",
                            "active_workers": min(normalized_workers, total_tasks),
                            "completed_archives": 0,
                            "failed_archives": 0,
                        }
                    )
                    future = executor.submit(
                        process_folder_excel_group,
                        input_root,
                        unit_folder,
                        unit_files,
                        category_keywords,
                        category_prefixes,
                        skip_dir,
                        unclassified_dir,
                        hc_dir,
                        dry_run,
                        enable_hc_filter,
                        normalized_excel_group_mode,
                    )
                    future_map[future] = (task_index, unit_folder)

                completed_count = 0
                failed_count = 0
                for future in as_completed(future_map):
                    index, archive_path = future_map[future]
                    result_status = "done"
                    try:
                        result = future.result()
                    except Exception as exc:
                        result_status = "error"
                        result = {
                            "archive_path": archive_path,
                            "archive_name": archive_path.name,
                            "success": False,
                            "rows": [],
                            "process_logs": [],
                            "debug_logs": [],
                            "exception_logs": [f"异常：处理压缩包失败 {archive_path.name}，原因：{exc}"],
                            "error_report_rows": [
                                {
                                    "压缩包名称": archive_path.name,
                                    "压缩包路径": str(archive_path),
                                    "异常类型": "处理失败",
                                    "异常原因": f"处理压缩包失败，原因：{exc}",
                                    "相关文件": "",
                                    "处理状态": "已跳过",
                                    "处理时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                }
                            ],
                            "archive_detail": {
                                "压缩包名": archive_path.name,
                                "压缩包路径": str(archive_path),
                                "处理状态": "异常",
                                "正式Excel数量": 0,
                                "正式Excel文件": "",
                                "提取行数": 0,
                                "识别品类": "",
                                "识别日期": "",
                                "是否跳过": "是",
                                "异常原因": f"处理压缩包失败，原因：{exc}",
                            },
                            "filename_validation": {},
                            "quantity_error_rows": [],
                            "header_report_rows": [],
                            "hc_report_rows": [],
                            "category_candidates": [],
                        }
                    if not result.get("ignored") and not result.get("success"):
                        failed_count += 1
                        result_status = "error"
                    archive_results_by_index[index] = result
                    completed_count += 1
                    emit_progress(
                        {
                            "current": completed_count,
                            "total": total_tasks,
                            "archive_name": archive_path.name,
                            "status": result_status,
                            "active_workers": min(normalized_workers, max(0, total_tasks - completed_count)),
                            "completed_archives": completed_count,
                            "failed_archives": failed_count,
                        }
                    )

            ignored_archives = 0
            for index in range(1, total_tasks + 1):
                result = archive_results_by_index[index]
                process_logs.extend(result.get("process_logs") or [])
                debug_logs.extend(result.get("debug_logs") or [])
                exception_logs.extend(result.get("exception_logs") or [])
                error_report_rows.extend(result.get("error_report_rows") or [])
                if result.get("archive_details"):
                    archive_details.extend(result["archive_details"])
                elif result.get("archive_detail"):
                    archive_details.append(result["archive_detail"])
                if result.get("filename_validations"):
                    filename_validations.extend(result["filename_validations"])
                elif result.get("filename_validation"):
                    filename_validations.append(result["filename_validation"])
                quantity_error_rows.extend(result.get("quantity_error_rows") or [])
                header_report_rows.extend(result.get("header_report_rows") or [])
                hc_report_rows.extend(result.get("hc_report_rows") or [])
                category_candidates.extend(result.get("category_candidates") or [])
                rows = list(result.get("rows") or [])
                all_rows.extend(rows)
                if result.get("ignored"):
                    ignored_archives += 1
                elif result.get("success"):
                    stats["success_archives"] += 1
                else:
                    stats["skipped_archives"] += 1

            if ignored_archives:
                stats["total_archives"] = max(0, total_tasks - ignored_archives)
            stats["extracted_rows"] = len(all_rows)
            stats["hc_file_count"] = len(hc_report_rows)
            stats["hc_copy_failed_count"] = sum(1 for row in hc_report_rows if row.get("处理状态") == "复制失败，已排除")
            stats["category_counts"] = build_category_counts(all_rows)
            def candidate_exclusion_key(row: dict[str, Any]) -> tuple[str, str] | tuple[str, str, str]:
                source_path = str(row.get("source_path") or "")
                if source_path:
                    return ("source_path", os.path.normcase(os.path.abspath(source_path)))
                return (
                    "archive_excel",
                    str(row.get("archive_name") or row.get("压缩包名") or ""),
                    str(row.get("excel_file") or row.get("Excel 文件名") or ""),
                )

            candidate_row_keys = {candidate_exclusion_key(candidate) for candidate in category_candidates}
            writable_source_rows = [
                row for row in all_rows
                if candidate_exclusion_key(row) not in candidate_row_keys
            ]
            merged_rows, merge_quantity_error_rows = consolidate_order_sku_rows(writable_source_rows)
            if len(merged_rows) != len(writable_source_rows):
                add_log(f"同一订单号下相同 SKU 已合并：{len(writable_source_rows)} 行 -> {len(merged_rows)} 行")
            quantity_error_rows.extend(merge_quantity_error_rows)
            quantity_error_rows = dedupe_quantity_error_rows(quantity_error_rows)
            classified_rows = [row for row in merged_rows if str(row.get("category") or "未分类") != "未分类"]
            if len(classified_rows) != len(merged_rows):
                add_log(f"未分类 Excel 已复制到：{unclassified_dir}；未分类行不写入正式汇总")
            rows_to_write, duplicate_report_rows, exact_duplicate_report_rows = detect_duplicates(
                classified_rows,
                output,
                clear=clear,
                detect_duplicate_orders=detect_duplicate_orders,
                skip_exact_duplicates=skip_exact_duplicates,
                dry_run=dry_run,
            )
            stats["duplicate_order_count"] = len(duplicate_report_rows)
            stats["skipped_exact_duplicate_count"] = sum(
                1 for row in exact_duplicate_report_rows if row.get("处理方式") == "已跳过"
            )
            duplicate_report_file_path = save_duplicate_report(output, duplicate_report_rows, exact_duplicate_report_rows)
            if dry_run:
                add_log("dry-run 模式：不会备份旧汇总 Excel，不会写入输出 Excel")
                stats["written_rows"] = 0
            elif not rows_to_write:
                no_writable_data = True
                add_log("本次无可写入数据，不修改汇总 Excel，不备份旧汇总 Excel")
                stats["written_rows"] = 0
            else:
                backup_path = backup_existing_output(output, backup_dir)
                write_to_output(rows_to_write, output, clear=clear)
                stats["written_rows"] = len(rows_to_write)
                wrote_output = True

            add_log(f"处理完成，总计提取：{stats['extracted_rows']} 行")
            success = True
    except Exception as exc:
        add_exception(f"异常：处理失败，原因：{exc}")
    finally:
        stats["extracted_rows"] = len(all_rows)
        stats["category_counts"] = build_category_counts(all_rows)
        if dry_run:
            stats["written_rows"] = 0
        stats["hc_file_count"] = len(hc_report_rows)
        stats["hc_copy_failed_count"] = sum(1 for row in hc_report_rows if row.get("处理状态") == "复制失败，已排除")
        stats["hc_dir"] = str(hc_dir.resolve())
        report_data = {
            "start_time": start_time.strftime("%Y-%m-%d %H:%M:%S"),
            "end_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "input_path": str(Path(input_path).expanduser()),
            "output_path": str(output),
            "clear": clear,
            "dry_run": dry_run,
            "wrote_output": wrote_output,
            "no_writable_data": no_writable_data or (not rows_to_write and not dry_run),
            "stats": dict(stats),
            "error_report_path": "",
            "duplicate_report_path": duplicate_report_file_path,
            "process_report_path": "",
            "debug_report_path": "",
            "backup_path": backup_path,
            "archive_details": list(archive_details),
            "filename_validations": list(filename_validations),
            "error_report_rows": list(error_report_rows),
            "quantity_error_rows": list(quantity_error_rows),
            "header_report_rows": list(header_report_rows),
            "hc_report_rows": list(hc_report_rows),
            "category_candidates": list(category_candidates),
            "duplicate_report_rows": list(duplicate_report_rows),
            "exact_duplicate_report_rows": list(exact_duplicate_report_rows),
            "rows_to_write": list(rows_to_write),
        }
        try:
            process_report_file_path = save_process_report(report_data, logs_dir)
        except Exception as exc:
            add_exception(f"异常：保存处理报告失败，原因：{exc}")
            process_report_file_path = ""
        stats["process_report_path"] = process_report_file_path
        debug_report_file_path = save_debug_report(debug_logs, logs_dir)
        log_file_path = save_run_log(
            stats,
            duplicate_report_rows,
            exact_duplicate_report_rows,
            duplicate_report_file_path,
            process_report_file_path,
            backup_path,
            logs_dir,
        )
        set_log_callback(None)

    return {
        "success": success,
        "total_rows": stats["extracted_rows"],
        "rows": list(all_rows),
        "merged_rows": list(merged_rows),
        "rows_to_write": list(rows_to_write),
        "written_row_records": list(rows_to_write),
        "written_rows": int(stats.get("written_rows", 0) or 0),
        "stats": dict(stats),
        "exception_logs": list(exception_logs),
        "process_logs": list(process_logs),
        "debug_logs": list(debug_logs),
        "error_report_rows": list(error_report_rows),
        "archive_details": list(archive_details),
        "filename_validations": list(filename_validations),
        "header_report_rows": list(header_report_rows),
        "hc_report_rows": list(hc_report_rows),
        "category_candidates": list(category_candidates),
        "duplicate_report_rows": list(duplicate_report_rows),
        "exact_duplicate_report_rows": list(exact_duplicate_report_rows),
        "duplicate_report_file_path": duplicate_report_file_path,
        "duplicate_report_path": duplicate_report_file_path,
        "error_report_path": "",
        "process_report_path": process_report_file_path,
        "debug_report_path": debug_report_file_path,
        "debug_report_file_path": debug_report_file_path,
        "backup_path": backup_path,
        "log_dir": str(logs_dir.resolve()),
        "error_count": len(error_report_rows) + len(quantity_error_rows),
        "duplicate_count": len(duplicate_report_rows),
        "exact_duplicate_count": len(exact_duplicate_report_rows),
        "skipped_archives": int(stats.get("skipped_archives", 0) or 0),
        "processed_archives": int(stats.get("success_archives", 0) or 0),
        "log_file_path": log_file_path,
        "output_path": str(output),
    }


def main() -> None:
    """
    1. 读取 --input
    2. 找到所有支持的压缩包文件
    3. 逐个解压处理
    4. 汇总所有 rows
    5. 按品类写入 output Excel
    6. 最后先打印异常日志，再打印正常处理日志
    """
    parser = argparse.ArgumentParser(description="批量解压 Excel 订单文件，并按品类汇总到不同 Sheet。")
    parser.add_argument("--input", required=True, help="存放多个压缩包的文件夹路径，或单个压缩包路径")
    parser.add_argument("--output", required=True, help="输出汇总 Excel 路径")
    parser.add_argument("--clear", action="store_true", help="清空旧汇总内容重新生成")
    parser.add_argument("--dry-run", action="store_true", help="仅预览处理结果，不写入输出 Excel")
    parser.add_argument("--workers", type=int, default=4, help="同时处理压缩包数量，范围 1-8，默认 4")
    parser.add_argument("--no-detect-duplicates", action="store_true", help="不检测重复订单号")
    parser.add_argument("--no-skip-exact-duplicates", action="store_true", help="不跳过完全重复行，只记录提示")
    parser.add_argument("--category-config", default=None, help="品类关键词配置文件路径，默认使用程序目录下的 category_config.json")
    parser.add_argument("--settings", default=None, help="GUI 设置文件路径，命令行运行时仅保留兼容参数")
    parser.add_argument("--report-dir", default=None, help="处理日志和处理报告保存目录，默认使用程序目录下的 logs")
    parser.add_argument("--backup-dir", default=None, help="旧汇总 Excel 备份目录，默认使用程序目录下的 backups")
    parser.add_argument("--input-mode", choices=sorted(INPUT_MODES), default=INPUT_MODE_ARCHIVES, help="输入来源：archives 只处理压缩包，folders 只处理文件夹 Excel，mixed 混合模式")
    parser.add_argument("--enable-hc-filter", action="store_true", help="启用 HC 文件过滤；默认不启用")
    parser.add_argument("--excel-group-mode", choices=sorted(EXCEL_GROUP_MODES), default=EXCEL_GROUP_SINGLE, help="Excel 处理规则：single 单文件订单模式，multi 多文件汇总模式")
    args = parser.parse_args()

    result = run_extract(
        args.input,
        args.output,
        clear=args.clear,
        dry_run=args.dry_run,
        workers=args.workers,
        detect_duplicate_orders=not args.no_detect_duplicates,
        skip_exact_duplicates=not args.no_skip_exact_duplicates,
        category_config_path=args.category_config,
        report_dir=args.report_dir,
        backup_dir=args.backup_dir,
        input_mode=args.input_mode,
        enable_hc_filter=args.enable_hc_filter,
        excel_group_mode=args.excel_group_mode,
    )
    print_logs(
        result.get("stats"),
        result.get("log_file_path"),
        result.get("duplicate_report_rows"),
        result.get("exact_duplicate_report_rows"),
        result.get("duplicate_report_file_path"),
        result.get("process_report_path"),
        result.get("backup_path"),
    )
    if not result["success"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
