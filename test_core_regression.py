from __future__ import annotations

import json
import shutil
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from extract_orders import run_extract


ORDER_HEADER = "\u4e9a\u9a6c\u900a\u8ba2\u5355\u53f7"
QUANTITY_HEADER = "\u6570\u91cf"
DATE_HEADER = "\u65e5\u671f"
SKIP_FOLDER_NAME = "\u672a\u5904\u7406\u538b\u7f29\u5305"
UNCLASSIFIED_FOLDER_NAME = "\u672a\u5206\u7c7bExcel"
MODIFY_PREFIX = "\u4fee\u6539"
WING_IMAGE_NECKLACE = "\u7fc5\u8180\u56fe\u7247\u9879\u94fe"
SILVER_WING_IMAGE_NECKLACE = "\u94f6\u7fc5\u8180\u56fe\u7247\u9879\u94fe"
DOG_TAG_KEYCHAIN = "\u519b\u724c\u94a5\u5319\u6263"


class DummyStatusVar:
    def __init__(self) -> None:
        self.value = ""

    def set(self, value: str) -> None:
        self.value = value


class DummyMaster:
    def __init__(self) -> None:
        self.saved = False

    def save_current_settings(self) -> None:
        self.saved = True


def make_order_workbook(path: Path, rows: list[tuple[str, str, object]], *, sheet_name: str = "Sheet1") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append([ORDER_HEADER, "SKU", QUANTITY_HEADER])
    for order_id, sku, quantity in rows:
        worksheet.append([order_id, sku, quantity])
    workbook.save(path)
    workbook.close()


def make_zip(zip_path: Path, entries: list[tuple[Path, str]]) -> None:
    with zipfile.ZipFile(zip_path, "w") as archive:
        for source, archive_name in entries:
            archive.write(source, archive_name)


def read_summary_rows(output_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(output_path, data_only=True)
    rows: list[dict[str, object]] = []
    try:
        for worksheet in workbook.worksheets:
            headers = [worksheet.cell(row=1, column=col).value for col in range(1, 5)]
            if headers != [ORDER_HEADER, "SKU", QUANTITY_HEADER, DATE_HEADER]:
                continue
            for row_index in range(2, worksheet.max_row + 1):
                values = [worksheet.cell(row=row_index, column=col).value for col in range(1, 5)]
                if not any(value not in (None, "") for value in values):
                    continue
                rows.append(
                    {
                        "sheet": worksheet.title,
                        "order_id": values[0],
                        "sku": values[1],
                        "quantity": values[2],
                        "date": values[3],
                    }
                )
    finally:
        workbook.close()
    return rows


def read_report_sheet_rows(report_path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = load_workbook(report_path, data_only=True)
    try:
        worksheet = workbook[sheet_name]
        headers = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]
        rows: list[dict[str, object]] = []
        for row_index in range(2, worksheet.max_row + 1):
            values = [worksheet.cell(row=row_index, column=col).value for col in range(1, worksheet.max_column + 1)]
            if not any(value not in (None, "") for value in values):
                continue
            rows.append(dict(zip(headers, values)))
        return rows
    finally:
        workbook.close()


class CoreRegressionTests(unittest.TestCase):
    def setUp(self) -> None:
        self.base = Path(tempfile.mkdtemp(prefix="core_regression_test_"))
        self.input_dir = self.base / "input"
        self.input_dir.mkdir()
        self.work_dir = self.base / "work"
        self.work_dir.mkdir()
        self.logs_dir = self.base / "logs"
        self.backups_dir = self.base / "backups"
        self.output_path = self.base / "summary.xlsx"
        self.category_config_path = self.base / "category_config.json"

    def tearDown(self) -> None:
        shutil.rmtree(self.base, ignore_errors=True)

    def run_tool(
        self,
        *,
        dry_run: bool = False,
        skip_exact_duplicates: bool = True,
        clear: bool = True,
        input_mode: str = "archives",
        excel_group_mode: str = "single",
        enable_hc_filter: bool = False,
    ) -> dict:
        return run_extract(
            str(self.input_dir),
            str(self.output_path),
            clear=clear,
            dry_run=dry_run,
            workers=1,
            report_dir=str(self.logs_dir),
            backup_dir=str(self.backups_dir),
            skip_exact_duplicates=skip_exact_duplicates,
            input_mode=input_mode,
            excel_group_mode=excel_group_mode,
            enable_hc_filter=enable_hc_filter,
        )

    def test_fixture_helpers_create_readable_order_zip(self) -> None:
        workbook_path = self.work_dir / "fixture.xlsx"
        make_order_workbook(workbook_path, [("ORDER-FIXTURE", "SKU-FIXTURE", 1)])
        zip_path = self.input_dir / "fixture.zip"

        make_zip(zip_path, [(workbook_path, workbook_path.name)])

        self.assertTrue(workbook_path.exists())
        self.assertTrue(zip_path.exists())
        with zipfile.ZipFile(zip_path) as archive:
            self.assertEqual(archive.namelist(), [workbook_path.name])

    def test_load_old_category_config_returns_empty_prefixes(self) -> None:
        self.category_config_path.write_text(
            json.dumps({"方黑名片架": ["方黑名片架"]}, ensure_ascii=False),
            encoding="utf-8",
        )

        from extract_orders import load_category_config_data

        config_data, path, error = load_category_config_data(self.category_config_path)

        self.assertEqual(error, "")
        self.assertEqual(Path(path), self.category_config_path)
        self.assertEqual(config_data.categories, {"方黑名片架": ["方黑名片架"]})
        self.assertEqual(config_data.prefixes, [])

    def test_load_old_category_config_allows_prefixes_category_name(self) -> None:
        self.category_config_path.write_text(
            json.dumps({"prefixes": ["prefix keyword"]}, ensure_ascii=False),
            encoding="utf-8",
        )

        from extract_orders import load_category_config_data

        config_data, _, error = load_category_config_data(self.category_config_path)

        self.assertEqual(error, "")
        self.assertEqual(config_data.categories, {"prefixes": ["prefix keyword"]})
        self.assertEqual(config_data.prefixes, [])

    def test_load_old_category_config_allows_categories_category_name(self) -> None:
        self.category_config_path.write_text(
            json.dumps({"categories": ["category keyword"]}, ensure_ascii=False),
            encoding="utf-8",
        )

        from extract_orders import load_category_config_data

        config_data, _, error = load_category_config_data(self.category_config_path)

        self.assertEqual(error, "")
        self.assertEqual(config_data.categories, {"categories": ["category keyword"]})
        self.assertEqual(config_data.prefixes, [])

    def test_load_structured_category_config_reads_prefixes(self) -> None:
        self.category_config_path.write_text(
            json.dumps(
                {
                    "prefixes": ["WZY", "HAL"],
                    "categories": {"纯木名片架": ["纯木名片架"]},
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        from extract_orders import load_category_config_data

        config_data, _, error = load_category_config_data(self.category_config_path)

        self.assertEqual(error, "")
        self.assertEqual(config_data.prefixes, ["WZY", "HAL"])
        self.assertEqual(config_data.categories, {"纯木名片架": ["纯木名片架"]})

    def test_save_confirmed_category_candidate_adds_prefix_and_category(self) -> None:
        self.category_config_path.write_text(
            json.dumps({"小钢片": ["小钢片"]}, ensure_ascii=False),
            encoding="utf-8",
        )

        from extract_orders import save_confirmed_category_candidate, load_category_config_data

        save_confirmed_category_candidate(
            self.category_config_path,
            prefix="HAL",
            category="方白名片架",
            keyword="方白名片架",
        )
        config_data, _, error = load_category_config_data(self.category_config_path)

        self.assertEqual(error, "")
        self.assertEqual(config_data.prefixes, ["HAL"])
        self.assertEqual(config_data.categories["方白名片架"], ["方白名片架"])
        self.assertIn("小钢片", config_data.categories)

    def test_save_confirmed_category_candidate_merges_without_duplicates(self) -> None:
        self.category_config_path.write_text(
            json.dumps(
                {
                    "prefixes": ["HAL"],
                    "categories": {"方白名片架": ["方白名片架"]},
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        from extract_orders import save_confirmed_category_candidate, load_category_config_data

        save_confirmed_category_candidate(
            self.category_config_path,
            prefix="HAL",
            category="方白名片架",
            keyword="方白名片架",
        )
        config_data, _, _ = load_category_config_data(self.category_config_path)

        self.assertEqual(config_data.prefixes, ["HAL"])
        self.assertEqual(config_data.categories["方白名片架"], ["方白名片架"])

    def test_category_config_window_preserves_prefixes_when_saving_keywords(self) -> None:
        self.category_config_path.write_text(
            json.dumps(
                {
                    "prefixes": ["HAL", "WZY"],
                    "categories": {"方白名片架": ["方白名片架"]},
                },
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )

        import extract_orders_gui
        from extract_orders import load_category_config_data

        window = object.__new__(extract_orders_gui.CategoryConfigWindow)
        window.config_path = self.category_config_path
        window.config = {}
        window.prefixes = []
        window.current_category = None
        window.status_var = DummyStatusVar()
        window.master = DummyMaster()
        window.refresh_categories = lambda: None

        with (
            patch.object(extract_orders_gui.messagebox, "showinfo"),
            patch.object(extract_orders_gui.messagebox, "showwarning"),
            patch.object(extract_orders_gui.messagebox, "showerror"),
        ):
            window.load_config()
            window.config["方白名片架"].append("新关键词")
            window.save_config()
        config_data, _, error = load_category_config_data(self.category_config_path)

        self.assertEqual(error, "")
        self.assertEqual(config_data.prefixes, ["HAL", "WZY"])
        self.assertEqual(config_data.categories["方白名片架"], ["方白名片架", "新关键词"])
        self.assertTrue(window.master.saved)

    def test_save_confirmed_category_candidate_rejects_invalid_existing_config(self) -> None:
        invalid_json = "{ invalid json"
        self.category_config_path.write_text(invalid_json, encoding="utf-8")

        from extract_orders import save_confirmed_category_candidate

        with self.assertRaises(ValueError):
            save_confirmed_category_candidate(
                self.category_config_path,
                prefix="HAL",
                category="方白名片架",
                keyword="方白名片架",
            )

        self.assertEqual(self.category_config_path.read_text(encoding="utf-8"), invalid_json)

    def test_extract_category_candidate_from_folder_name(self) -> None:
        from extract_orders import build_category_candidate_from_name

        candidate = build_category_candidate_from_name("18~19-0620-0625-方黑名片架-20单-3个", [])

        self.assertIsNotNone(candidate)
        self.assertEqual(candidate["raw_candidate"], "方黑名片架")
        self.assertEqual(candidate["prefix"], "")
        self.assertEqual(candidate["category"], "方黑名片架")

    def test_extract_category_candidate_cleans_separated_prefix(self) -> None:
        from extract_orders import build_category_candidate_from_name

        candidate = build_category_candidate_from_name("17-0625-WZY-纯木名片架-5单5个", ["WZY"])

        self.assertIsNotNone(candidate)
        self.assertEqual(candidate["raw_candidate"], "WZY-纯木名片架")
        self.assertEqual(candidate["prefix"], "WZY")
        self.assertEqual(candidate["category"], "纯木名片架")

    def test_extract_category_candidate_cleans_attached_prefix(self) -> None:
        from extract_orders import build_category_candidate_from_name

        candidate = build_category_candidate_from_name("14-0625-HAL小钢片-4单5个", ["HAL"])

        self.assertIsNotNone(candidate)
        self.assertEqual(candidate["raw_candidate"], "HAL小钢片")
        self.assertEqual(candidate["prefix"], "HAL")
        self.assertEqual(candidate["category"], "小钢片")

    def test_extract_category_candidate_keeps_extensionless_dot_date_folder_name(self) -> None:
        from extract_orders import build_category_candidate_from_name

        candidate = build_category_candidate_from_name("33~35-4.18-方黑名片架-20单-3个", [])

        self.assertIsNotNone(candidate)
        self.assertEqual(candidate["category"], "方黑名片架")

    def test_extract_category_candidate_does_not_strip_attached_ascii_prefix_extension(self) -> None:
        from extract_orders import build_category_candidate_from_name

        candidate = build_category_candidate_from_name("17-0625-WZYX小钢片-4单5个", ["WZY"])

        self.assertIsNotNone(candidate)
        self.assertEqual(candidate["prefix"], "")
        self.assertEqual(candidate["category"], "WZYX小钢片")

    def test_extract_category_candidate_ignores_structure_only_name(self) -> None:
        from extract_orders import build_category_candidate_from_name

        candidate = build_category_candidate_from_name("18~19-0620-0625-20单-3个", ["WZY"])

        self.assertIsNone(candidate)

    def test_dry_run_scans_and_reports_without_writing_output_or_backup(self) -> None:
        workbook_path = self.work_dir / "0507-WZY-knife-1order-2pcs.xlsx"
        make_order_workbook(workbook_path, [("ORDER-DRY", "SKU-DRY", 2)])
        make_zip(self.input_dir / "dry_run.zip", [(workbook_path, workbook_path.name)])

        result = self.run_tool(dry_run=True)

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 1)
        self.assertEqual(result["written_rows"], 0)
        self.assertFalse(self.output_path.exists())
        self.assertFalse(self.backups_dir.exists())
        self.assertTrue(Path(result["process_report_path"]).exists())
        self.assertTrue(Path(result["log_file_path"]).exists())
        self.assertTrue(result["stats"]["dry_run"])

    def test_same_order_same_sku_merges_quantities_before_write(self) -> None:
        workbook_path = self.work_dir / "0507-WZY-knife-1order-5pcs.xlsx"
        make_order_workbook(
            workbook_path,
            [
                ("ORDER-MERGE", "SKU-MERGE", 2),
                ("ORDER-MERGE", "SKU-MERGE", 3),
            ],
        )
        make_zip(self.input_dir / "merge.zip", [(workbook_path, workbook_path.name)])

        result = self.run_tool()

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 2)
        self.assertEqual(result["written_rows"], 1)
        rows = read_summary_rows(self.output_path)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["order_id"], "ORDER-MERGE")
        self.assertEqual(rows[0]["sku"], "SKU-MERGE")
        self.assertEqual(rows[0]["quantity"], 5)

    def test_same_order_different_sku_stays_as_two_normal_rows(self) -> None:
        workbook_path = self.work_dir / "0507-WZY-knife-1order-3pcs.xlsx"
        make_order_workbook(
            workbook_path,
            [
                ("ORDER-MULTI", "SKU-A", 1),
                ("ORDER-MULTI", "SKU-B", 2),
            ],
        )
        make_zip(self.input_dir / "different_sku.zip", [(workbook_path, workbook_path.name)])

        result = self.run_tool()

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 2)
        self.assertEqual(result["duplicate_count"], 0)
        rows = sorted(read_summary_rows(self.output_path), key=lambda item: str(item["sku"]))
        self.assertEqual([row["sku"] for row in rows], ["SKU-A", "SKU-B"])
        self.assertEqual([row["quantity"] for row in rows], [1, 2])

    def test_daily_order_summary_counts_unique_orders_by_filename_date(self) -> None:
        day_one_path = self.work_dir / "0507-WZY-knife-2order-6pcs.xlsx"
        day_two_path = self.work_dir / "0508-WZY-knife-1order-4pcs.xlsx"
        make_order_workbook(
            day_one_path,
            [
                ("ORDER-DAILY-A", "SKU-A", 2),
                ("ORDER-DAILY-A", "SKU-B", 3),
                ("ORDER-DAILY-B", "SKU-C", 1),
            ],
        )
        make_order_workbook(day_two_path, [("ORDER-DAILY-C", "SKU-D", 4)])
        make_zip(self.input_dir / "daily.zip", [(day_one_path, day_one_path.name), (day_two_path, day_two_path.name)])

        result = self.run_tool(excel_group_mode="multi")

        self.assertTrue(result["success"])
        report_rows = read_report_sheet_rows(Path(result["process_report_path"]), "每日单量汇总")
        by_date = {row["日期"]: row for row in report_rows}
        self.assertEqual(by_date["5月7日"]["订单数"], 2)
        self.assertEqual(by_date["5月7日"]["数量合计"], 6)
        self.assertEqual(by_date["5月7日"]["明细行数"], 3)
        self.assertEqual(by_date["5月8日"]["订单数"], 1)
        self.assertEqual(by_date["5月8日"]["数量合计"], 4)
        self.assertEqual(by_date["5月8日"]["明细行数"], 1)

    def test_reissue_before_category_is_skipped_and_copied_to_skip_folder(self) -> None:
        workbook_path = self.input_dir / f"13-\u8865\u53d1{DOG_TAG_KEYCHAIN}.xlsx"
        make_order_workbook(workbook_path, [("ORDER-REISSUE-SKIP", "SKU-REISSUE-SKIP", 1)])

        result = self.run_tool(input_mode="folders")

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 0)
        self.assertFalse(self.output_path.exists())
        skip_copy = self.input_dir / SKIP_FOLDER_NAME / workbook_path.name
        self.assertTrue(skip_copy.exists())

    def test_reissue_after_category_is_processed_normally(self) -> None:
        workbook_path = self.input_dir / f"13-{DOG_TAG_KEYCHAIN}-\u8865\u53d1.xlsx"
        make_order_workbook(workbook_path, [("ORDER-REISSUE-KEEP", "SKU-REISSUE-KEEP", 2)])

        result = self.run_tool(input_mode="folders")

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 1)
        rows = read_summary_rows(self.output_path)
        self.assertEqual(rows[0]["sheet"], DOG_TAG_KEYCHAIN)
        self.assertEqual(rows[0]["order_id"], "ORDER-REISSUE-KEEP")
        self.assertEqual(rows[0]["quantity"], 2)

    def test_same_order_empty_sku_still_merges_when_order_id_exists(self) -> None:
        workbook_path = self.work_dir / "0507-WZY-knife-1order-7pcs.xlsx"
        make_order_workbook(
            workbook_path,
            [
                ("ORDER-EMPTY-SKU", "", 3),
                ("ORDER-EMPTY-SKU", "", 4),
            ],
        )
        make_zip(self.input_dir / "empty_sku.zip", [(workbook_path, workbook_path.name)])

        result = self.run_tool()

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 2)
        self.assertEqual(result["written_rows"], 1)
        rows = read_summary_rows(self.output_path)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["order_id"], "ORDER-EMPTY-SKU")
        self.assertIn(rows[0]["sku"], ("", None))
        self.assertEqual(rows[0]["quantity"], 7)

    def test_exact_duplicate_rows_are_skipped_when_setting_is_enabled(self) -> None:
        workbook_path = self.work_dir / "0507-WZY-knife-2rows-2pcs.xlsx"
        make_order_workbook(
            workbook_path,
            [
                ("", "SKU-DUP", 1),
                ("", "SKU-DUP", 1),
            ],
        )
        make_zip(self.input_dir / "exact_duplicate_skip.zip", [(workbook_path, workbook_path.name)])

        result = self.run_tool(skip_exact_duplicates=True)

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 1)
        self.assertEqual(result["exact_duplicate_count"], 1)
        rows = read_summary_rows(self.output_path)
        self.assertEqual(len(rows), 1)

    def test_exact_duplicate_rows_can_be_written_when_setting_is_disabled(self) -> None:
        workbook_path = self.work_dir / "0507-WZY-knife-2rows-2pcs.xlsx"
        make_order_workbook(
            workbook_path,
            [
                ("", "SKU-DUP", 1),
                ("", "SKU-DUP", 1),
            ],
        )
        make_zip(self.input_dir / "exact_duplicate_write.zip", [(workbook_path, workbook_path.name)])

        result = self.run_tool(skip_exact_duplicates=False)

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 2)
        self.assertEqual(result["exact_duplicate_count"], 1)
        rows = read_summary_rows(self.output_path)
        self.assertEqual(len(rows), 2)

    def test_archive_with_multiple_order_subfolders_processes_each_subfolder(self) -> None:
        first = self.work_dir / "0507-WZY-knife-1order-1pc.xlsx"
        second = self.work_dir / "0508-WZY-knife-1order-2pcs.xlsx"
        make_order_workbook(first, [("ORDER-SUB-1", "SKU-SUB-1", 1)])
        make_order_workbook(second, [("ORDER-SUB-2", "SKU-SUB-2", 2)])
        make_zip(
            self.input_dir / "multi_subfolders.zip",
            [
                (first, f"folder-a/{first.name}"),
                (second, f"folder-b/{second.name}"),
            ],
        )

        result = self.run_tool()

        self.assertTrue(result["success"])
        self.assertEqual(result["processed_archives"], 1)
        self.assertEqual(result["written_rows"], 2)
        rows = sorted(read_summary_rows(self.output_path), key=lambda item: str(item["order_id"]))
        self.assertEqual([row["order_id"] for row in rows], ["ORDER-SUB-1", "ORDER-SUB-2"])
        self.assertTrue(Path(result["process_report_path"]).exists())

    def test_archive_multi_subfolder_candidate_uses_subfolder_name(self) -> None:
        workbook_path = self.work_dir / "order.xlsx"
        make_order_workbook(workbook_path, [("ORDER-ARCHIVE-CANDIDATE", "SKU-UNKNOWN", 1)])
        subfolder_name = "1~2-0605-\u65b9\u767d\u540d\u7247\u67b6-1\u5355-1\u4e2a"
        make_zip(
            self.input_dir / "multi_subfolder_candidate.zip",
            [(workbook_path, f"{subfolder_name}/{workbook_path.name}")],
        )

        result = self.run_tool()

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 1)
        self.assertEqual(result["written_rows"], 0)
        self.assertFalse(self.output_path.exists())
        candidates = result["category_candidates"]
        self.assertEqual(len(candidates), 1)
        self.assertEqual(candidates[0]["source_name"], subfolder_name)
        self.assertEqual(candidates[0]["category"], "\u65b9\u767d\u540d\u7247\u67b6")

    def test_archive_candidate_exclusion_keeps_same_basename_classified_subfolder(self) -> None:
        candidate_dir = self.work_dir / "candidate"
        candidate_dir.mkdir()
        candidate_workbook = candidate_dir / "order.xlsx"
        make_order_workbook(candidate_workbook, [("ORDER-CANDIDATE", "SKU-CANDIDATE", 1)])
        classified_dir = self.work_dir / "classified"
        classified_dir.mkdir()
        classified_workbook = classified_dir / "order.xlsx"
        make_order_workbook(classified_workbook, [("ORDER-CLASSIFIED", "SKU-CLASSIFIED", 1)])
        candidate_subfolder = "1~2-0605-\u65b9\u767d\u540d\u7247\u67b6-1\u5355-1\u4e2a"
        classified_subfolder = f"3~4-0605-{SILVER_WING_IMAGE_NECKLACE}-1\u5355-1\u4e2a"
        make_zip(
            self.input_dir / "same_basename_mixed_candidates.zip",
            [
                (candidate_workbook, f"{candidate_subfolder}/{candidate_workbook.name}"),
                (classified_workbook, f"{classified_subfolder}/{classified_workbook.name}"),
            ],
        )

        result = self.run_tool()

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 2)
        self.assertEqual(len(result["category_candidates"]), 1)
        self.assertEqual(result["written_rows"], 1)
        self.assertTrue(self.output_path.exists())
        rows = read_summary_rows(self.output_path)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["order_id"], "ORDER-CLASSIFIED")
        self.assertEqual(rows[0]["sheet"], WING_IMAGE_NECKLACE)

    def test_archive_name_is_used_when_excel_filename_has_no_category_or_date(self) -> None:
        workbook_path = self.work_dir / "order.xlsx"
        make_order_workbook(workbook_path, [("ORDER-ARCHIVE-FALLBACK", "SKU-WING", 10)])
        make_zip(
            self.input_dir / f"13-1216-HC-{SILVER_WING_IMAGE_NECKLACE}-9\u5355-10\u4ef6.zip",
            [(workbook_path, workbook_path.name)],
        )

        result = self.run_tool()

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 1)
        rows = read_summary_rows(self.output_path)
        self.assertEqual(rows[0]["sheet"], WING_IMAGE_NECKLACE)
        self.assertEqual(rows[0]["date"], "12\u670816\u65e5")
        self.assertEqual(result["stats"]["category_counts"], {WING_IMAGE_NECKLACE: 1})

    def test_reissue_before_category_in_archive_name_is_skipped(self) -> None:
        workbook_path = self.work_dir / "order.xlsx"
        archive_path = self.input_dir / f"13-\u8865\u53d1{DOG_TAG_KEYCHAIN}.zip"
        make_order_workbook(workbook_path, [("ORDER-ARCHIVE-REISSUE-SKIP", "SKU-ARCHIVE-REISSUE-SKIP", 1)])
        make_zip(archive_path, [(workbook_path, workbook_path.name)])

        result = self.run_tool()

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 0)
        self.assertFalse(self.output_path.exists())
        self.assertTrue((self.input_dir / SKIP_FOLDER_NAME / archive_path.name).exists())

    def test_folder_mode_recursively_extracts_excel_files(self) -> None:
        nested = self.input_dir / "orders" / "day-1"
        nested.mkdir(parents=True)
        workbook_path = nested / "0507-WZY-knife-1order-4pcs.xlsx"
        make_order_workbook(workbook_path, [("ORDER-FOLDER", "SKU-FOLDER", 4)])

        result = self.run_tool(input_mode="folders")

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 1)
        self.assertEqual(result["written_rows"], 1)
        rows = read_summary_rows(self.output_path)
        self.assertEqual(rows[0]["order_id"], "ORDER-FOLDER")

    def test_unclassified_folder_excel_records_category_candidate(self) -> None:
        folder = self.input_dir / "0607" / "1~2-0605-\u65b9\u767d\u540d\u7247\u67b6-1\u5355-1\u4e2a"
        folder.mkdir(parents=True)
        workbook_path = folder / "random-order.xlsx"
        make_order_workbook(workbook_path, [("ORDER-UNCLASSIFIED", "SKU-UNKNOWN", 1)])

        result = self.run_tool(input_mode="folders")

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 1)
        self.assertEqual(result["written_rows"], 0)
        self.assertFalse(self.output_path.exists())
        self.assertTrue((self.input_dir / UNCLASSIFIED_FOLDER_NAME / workbook_path.name).exists())
        candidates = result["category_candidates"]
        self.assertEqual(len(candidates), 1)
        self.assertEqual(candidates[0]["source_name"], "1~2-0605-\u65b9\u767d\u540d\u7247\u67b6-1\u5355-1\u4e2a")
        self.assertEqual(candidates[0]["raw_candidate"], "\u65b9\u767d\u540d\u7247\u67b6")
        self.assertEqual(candidates[0]["category"], "\u65b9\u767d\u540d\u7247\u67b6")
        self.assertEqual(candidates[0]["status"], "\u5f85\u786e\u8ba4")

    def test_process_report_contains_new_category_candidate_sheet(self) -> None:
        folder = self.input_dir / "0607" / "1~2-0605-\u65b9\u767d\u540d\u7247\u67b6-1\u5355-1\u4e2a"
        folder.mkdir(parents=True)
        workbook_path = folder / "random-order.xlsx"
        make_order_workbook(workbook_path, [("ORDER-UNCLASSIFIED", "SKU-UNKNOWN", 1)])

        result = self.run_tool(input_mode="folders")

        report_path = Path(result["stats"]["process_report_path"])
        workbook = load_workbook(report_path, data_only=True)
        try:
            self.assertIn("\u65b0\u54c1\u7c7b\u5019\u9009", workbook.sheetnames)
            rows = list(workbook["\u65b0\u54c1\u7c7b\u5019\u9009"].iter_rows(values_only=True))
        finally:
            workbook.close()
        self.assertEqual(
            rows[0][:6],
            ("\u5e8f\u53f7", "\u539f\u59cb\u540d\u79f0", "\u539f\u59cb\u5019\u9009", "\u8bc6\u522b\u524d\u7f00", "\u5019\u9009\u54c1\u7c7b", "\u72b6\u6001"),
        )
        self.assertEqual(rows[1][1], "1~2-0605-\u65b9\u767d\u540d\u7247\u67b6-1\u5355-1\u4e2a")
        self.assertEqual(rows[1][4], "\u65b9\u767d\u540d\u7247\u67b6")

    def test_folder_mode_skips_generated_unclassified_folder_on_later_runs(self) -> None:
        folder = self.input_dir / "0607" / "1~2-0605-\u65b9\u767d\u540d\u7247\u67b6-1\u5355-1\u4e2a"
        folder.mkdir(parents=True)
        workbook_path = folder / "random-order.xlsx"
        make_order_workbook(workbook_path, [("ORDER-UNCLASSIFIED-RERUN", "SKU-UNKNOWN", 1)])

        first_result = self.run_tool(input_mode="folders")

        self.assertTrue(first_result["success"])
        self.assertEqual(first_result["total_rows"], 1)
        self.assertEqual(len(first_result["category_candidates"]), 1)
        self.assertTrue((self.input_dir / UNCLASSIFIED_FOLDER_NAME / workbook_path.name).exists())

        second_result = self.run_tool(input_mode="folders")

        self.assertTrue(second_result["success"])
        self.assertEqual(second_result["total_rows"], 1)
        self.assertEqual(len(second_result["category_candidates"]), 1)

    def test_folder_name_is_used_when_excel_filename_has_no_category_or_date(self) -> None:
        folder = self.input_dir / f"13-1216-HC-{SILVER_WING_IMAGE_NECKLACE}-9\u5355-10\u4ef6"
        folder.mkdir()
        workbook_path = folder / "order.xlsx"
        make_order_workbook(workbook_path, [("ORDER-FOLDER-FALLBACK", "SKU-WING", 10)])

        result = self.run_tool(input_mode="folders")

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 1)
        rows = read_summary_rows(self.output_path)
        self.assertEqual(rows[0]["sheet"], WING_IMAGE_NECKLACE)
        self.assertEqual(rows[0]["date"], "12\u670816\u65e5")
        self.assertEqual(result["stats"]["category_counts"], {WING_IMAGE_NECKLACE: 1})

    def test_folder_mode_copies_modify_excel_file_to_skip_folder(self) -> None:
        workbook_path = self.input_dir / f"{MODIFY_PREFIX}-0507-WZY-knife-1order-1pc.xlsx"
        make_order_workbook(workbook_path, [("ORDER-MODIFY-FOLDER", "SKU-MODIFY", 1)])

        result = self.run_tool(input_mode="folders")

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 0)
        self.assertEqual(result["written_rows"], 0)
        self.assertFalse(self.output_path.exists())
        self.assertTrue((self.input_dir / SKIP_FOLDER_NAME / workbook_path.name).exists())

    def test_folder_mode_modify_excel_dry_run_does_not_create_skip_folder(self) -> None:
        workbook_path = self.input_dir / f"{MODIFY_PREFIX}-0507-WZY-knife-1order-1pc.xlsx"
        make_order_workbook(workbook_path, [("ORDER-MODIFY-DRY", "SKU-MODIFY", 1)])

        result = self.run_tool(input_mode="folders", dry_run=True)

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 0)
        self.assertEqual(result["written_rows"], 0)
        self.assertFalse(self.output_path.exists())
        self.assertFalse((self.input_dir / SKIP_FOLDER_NAME).exists())

    def test_mixed_mode_extracts_archive_and_folder_excel_files(self) -> None:
        folder_excel = self.input_dir / "folder-orders" / "0507-WZY-knife-1order-2pcs.xlsx"
        folder_excel.parent.mkdir(parents=True)
        make_order_workbook(folder_excel, [("ORDER-FOLDER-MIX", "SKU-FOLDER", 2)])
        archive_excel = self.work_dir / "0508-WZY-knife-1order-3pcs.xlsx"
        make_order_workbook(archive_excel, [("ORDER-ARCHIVE-MIX", "SKU-ARCHIVE", 3)])
        make_zip(self.input_dir / "archive_orders.zip", [(archive_excel, archive_excel.name)])

        result = self.run_tool(input_mode="mixed")

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 2)
        rows = sorted(read_summary_rows(self.output_path), key=lambda item: str(item["order_id"]))
        self.assertEqual([row["order_id"] for row in rows], ["ORDER-ARCHIVE-MIX", "ORDER-FOLDER-MIX"])

    def test_mixed_mode_single_excel_input_has_no_archive_format_error(self) -> None:
        workbook_path = self.input_dir / "0507-WZY-knife-1order-3pcs.xlsx"
        make_order_workbook(workbook_path, [("ORDER-SINGLE-XLSX", "SKU-SINGLE", 3)])

        result = run_extract(
            str(workbook_path),
            str(self.output_path),
            clear=True,
            workers=1,
            report_dir=str(self.logs_dir),
            backup_dir=str(self.backups_dir),
            input_mode="mixed",
        )

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 1)
        self.assertEqual(result["error_count"], 0)
        self.assertEqual(result["exception_logs"], [])
        rows = read_summary_rows(self.output_path)
        self.assertEqual(rows[0]["order_id"], "ORDER-SINGLE-XLSX")

    def test_mixed_mode_ignores_archives_inside_generated_skip_folder(self) -> None:
        skipped_dir = self.input_dir / SKIP_FOLDER_NAME
        skipped_dir.mkdir()
        ignored_excel = self.work_dir / "0506-WZY-knife-1order-9pcs.xlsx"
        make_order_workbook(ignored_excel, [("ORDER-IGNORED-SKIP", "SKU-IGNORED", 9)])
        make_zip(skipped_dir / "ignored.zip", [(ignored_excel, ignored_excel.name)])
        real_excel = self.input_dir / "real.xlsx"
        make_order_workbook(real_excel, [("ORDER-REAL-MIX", "SKU-REAL", 1)])

        result = self.run_tool(input_mode="mixed")

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 1)
        rows = read_summary_rows(self.output_path)
        self.assertEqual([row["order_id"] for row in rows], ["ORDER-REAL-MIX"])

    def test_single_file_mode_skips_multiple_excel_files_in_one_folder_unit(self) -> None:
        folder = self.input_dir / "same-unit"
        folder.mkdir()
        make_order_workbook(folder / "0507-WZY-knife-1order-1pc.xlsx", [("ORDER-A", "SKU-A", 1)])
        make_order_workbook(folder / "0508-WZY-knife-1order-2pcs.xlsx", [("ORDER-B", "SKU-B", 2)])

        result = self.run_tool(input_mode="folders", excel_group_mode="single")

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 0)
        self.assertEqual(result["written_rows"], 0)
        self.assertFalse(self.output_path.exists())
        self.assertGreaterEqual(result["skipped_archives"], 1)
        skip_details = "\n".join(
            str(item) for item in result.get("exception_logs", []) + result.get("error_report_rows", [])
        )
        self.assertIn("多个正式 Excel", skip_details)
        self.assertIn("0507-WZY-knife-1order-1pc.xlsx", skip_details)
        self.assertIn("0508-WZY-knife-1order-2pcs.xlsx", skip_details)

    def test_multi_file_summary_mode_extracts_multiple_excel_files_in_one_folder_unit(self) -> None:
        folder = self.input_dir / "same-unit"
        folder.mkdir()
        make_order_workbook(folder / "0507-WZY-knife-1order-1pc.xlsx", [("ORDER-A", "SKU-A", 1)])
        make_order_workbook(folder / "0508-WZY-knife-1order-2pcs.xlsx", [("ORDER-B", "SKU-B", 2)])

        result = self.run_tool(input_mode="folders", excel_group_mode="multi")

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 2)
        rows = sorted(read_summary_rows(self.output_path), key=lambda item: str(item["order_id"]))
        self.assertEqual([row["order_id"] for row in rows], ["ORDER-A", "ORDER-B"])

    def test_multi_file_summary_mode_applies_inside_archives(self) -> None:
        first = self.work_dir / "0507-WZY-knife-1order-1pc.xlsx"
        second = self.work_dir / "0508-WZY-knife-1order-2pcs.xlsx"
        make_order_workbook(first, [("ORDER-ZIP-A", "SKU-A", 1)])
        make_order_workbook(second, [("ORDER-ZIP-B", "SKU-B", 2)])
        make_zip(self.input_dir / "multi_excel.zip", [(first, first.name), (second, second.name)])

        result = self.run_tool(excel_group_mode="multi")

        self.assertTrue(result["success"])
        self.assertEqual(result["written_rows"], 2)
        rows = sorted(read_summary_rows(self.output_path), key=lambda item: str(item["order_id"]))
        self.assertEqual([row["order_id"] for row in rows], ["ORDER-ZIP-A", "ORDER-ZIP-B"])

    def test_multi_file_summary_mode_returns_validation_for_each_excel(self) -> None:
        first = self.work_dir / "0507-WZY-knife-1order-1pc.xlsx"
        second = self.work_dir / "0508-WZY-knife-1order-2pcs.xlsx"
        make_order_workbook(first, [("ORDER-VALIDATION-A", "SKU-A", 1)])
        make_order_workbook(second, [("ORDER-VALIDATION-B", "SKU-B", 2)])
        make_zip(self.input_dir / "multi_validation.zip", [(first, first.name), (second, second.name)])

        result = self.run_tool(excel_group_mode="multi")

        self.assertTrue(result["success"])
        validation_names = {item["Excel文件名"] for item in result["filename_validations"]}
        self.assertEqual(validation_names, {first.name, second.name})
        archive_excel_names = result["archive_details"][0]["正式Excel文件"]
        self.assertIn(first.name, archive_excel_names)
        self.assertIn(second.name, archive_excel_names)

    def test_modify_archive_is_copied_to_skip_folder_and_excluded_from_output(self) -> None:
        workbook_path = self.work_dir / "0507-WZY-knife-1order-1pc.xlsx"
        make_order_workbook(workbook_path, [("ORDER-SKIP", "SKU-SKIP", 1)])
        archive_name = f"{MODIFY_PREFIX}-order.zip"
        make_zip(self.input_dir / archive_name, [(workbook_path, workbook_path.name)])

        result = self.run_tool()

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 0)
        self.assertEqual(result["written_rows"], 0)
        self.assertFalse(self.output_path.exists())
        self.assertTrue((self.input_dir / SKIP_FOLDER_NAME / archive_name).exists())

    def test_modify_archive_dry_run_does_not_create_skip_folder(self) -> None:
        workbook_path = self.work_dir / "0507-WZY-knife-1order-1pc.xlsx"
        make_order_workbook(workbook_path, [("ORDER-SKIP-DRY", "SKU-SKIP", 1)])
        archive_name = f"{MODIFY_PREFIX}-dry.zip"
        make_zip(self.input_dir / archive_name, [(workbook_path, workbook_path.name)])

        result = self.run_tool(dry_run=True)

        self.assertTrue(result["success"])
        self.assertEqual(result["total_rows"], 0)
        self.assertEqual(result["written_rows"], 0)
        self.assertFalse((self.input_dir / SKIP_FOLDER_NAME).exists())

    def test_reports_and_returned_paths_exist_after_run(self) -> None:
        workbook_path = self.work_dir / "0507-WZY-knife-1order-1pc.xlsx"
        make_order_workbook(workbook_path, [("ORDER-REPORT", "SKU-REPORT", 1)])
        make_zip(self.input_dir / "reports.zip", [(workbook_path, workbook_path.name)])

        result = self.run_tool()

        self.assertTrue(result["success"])
        self.assertTrue(Path(result["process_report_path"]).exists())
        self.assertTrue(Path(result["debug_report_path"]).exists())
        self.assertTrue(Path(result["log_file_path"]).exists())


if __name__ == "__main__":
    unittest.main()
